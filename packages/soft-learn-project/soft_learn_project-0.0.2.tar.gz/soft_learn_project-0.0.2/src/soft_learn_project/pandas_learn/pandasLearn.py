import pandas as pd
import os


class PandasLearn():
  def __init__(self) -> None:
    """
      * 两种数据结构:
      * Series 是一列，DataFrame是多列
      * 可以理解为把数据变成excel表格型数据，Series 是一列，DataFrame是多列
    """
    pass

  def install(self):
    string = """默认也安装在 conda环境里面了
    conda install pandas
    """
    print(string)
    return None

  def set_option(self):
    """解决pandas中打印 DataFrame 行列显示不全的问题
    """
    # 最多显示50列，可以设置为None 表示显示所有列
    pd.set_option('display.max_columns', 50)
    # 每列显示150个字符  # 调整显示宽度，以便整行显示
    pd.set_option('display.width', 150)
    # 设置显示格式为科学计数法，并保留 6 位有效数字
    pd.options.display.float_format = '{: .3e}'.format
    pass

  def styled_df_example(self):
    """显示带有样式的 DataFrame
    """
    # 创建示例 DataFrame
    data = {
        'A': [1, 2, 3, 4],
        'B': [5, 6, 7, 8],
        'C': [9, 10, 11, 12]
    }
    df = pd.DataFrame(data)

    # 使用 lambda 函数定义颜色映射
    styled_df = df.style.apply(lambda s: ['color: red' if v > 6 else '' for v in s],
                               axis=1, subset=['B'])
    return styled_df

  def df_search(self, df: pd.DataFrame,
                column='Name',
                is_index=False,
                key_str='xiaohu'):
    """
    * 查找关键字 所在的行
    * df = mypaper.PaperReferees.get_all_df()
    ---
    - 查找 Name 列中 'Xiaohu Yu' 所在的行
    - df[df[column] == 'Xiaohu Yu']
    """
    if is_index:
      df_result = df[df.index.str.contains(pat=key_str,
                                           case=False,
                                           na=False,
                                           )]
    else:
      df_result = df[df[column].str.contains(pat=key_str,
                                             case=False,
                                             na=False)]
    return df_result

  def df_search_universal(self, df: pd.DataFrame,
                          key_str='王金龙',
                          ):
    """给出关键词所带的行
    """
    df_result = df[df.apply(lambda row: any(key_str.lower() in str(cell).lower()
                            for cell in row), axis=1)]
    return df_result

  def get_exel_obj(self,
                   fname='/Users/wangjinlong/job/teach/course_课程/clsjymo_材料设计与模拟/xgcl_相关材料/课堂花名册.xlsx',):

    suffix = os.path.splitext(fname)[1][1:]
    if suffix == 'xls':
      engine = 'xlrd'
    elif suffix == 'xlsx':
      engine = 'openpyxl'
    else:
      raise ValueError(f'不支持的文件格式: {suffix}')
    excel = pd.ExcelFile(path_or_buffer=fname,
                         engine=engine,)
    return excel

  def get_df_from_excel(self,
                        excel_obj: pd.ExcelFile,
                        skiprows=1,
                        header=0,
                        sheet_num=0,
                        ):
    """
      * 或者使用
      df = pd.read_excel(io=fname, skiprows=skiprows)
      ---
      excel_boj -> get_exel_obj
    """

    # print(f'当前默认读取第一个表单, 表单: {excel_obj.sheet_names}')
    df = excel_obj.parse(sheet_name=excel_obj.sheet_names[sheet_num],
                         header=header,
                         skiprows=skiprows,)
    return df

  def get_df_from_excel_wrapper(self, fname='xxx.xlsx',
                                skiprows=0,
                                sheet_num=0,):
    excel_obj = self.get_exel_obj(fname,)
    df = self.get_df_from_excel(excel_obj=excel_obj,
                                skiprows=skiprows,
                                sheet_num=sheet_num,)
    return df

  def get_df_cleaned_dropna(self,
                            df: pd.DataFrame,
                            to_replace='-',
                            ):
    """去掉无效数据的行
    """

    df = df.replace(to_replace=to_replace, value=pd.NA)  # inplace=True
    df_cleaned = df.dropna()
    # 或者这个
    # df = df[~df['得分\n\n（试卷总分：100.0）'].str.contains('-')]
    return df_cleaned

  def get_df_from_dict(self, data: dict,
                       orient: str = 'index',
                       ) -> pd.DataFrame:
    """data={'lc':3.14, 'E_patom': -8.9}
    pd.DataFrame({'name': list(data.keys()), 'value': list(data.values())})
    """
    df = pd.DataFrame.from_dict(data, orient=orient)
    return df

  def concat_df(self, df_list) -> pd.DataFrame:
    df: pd.DataFrame = pd.concat(df_list, axis=0)
    # 避免重复
    df = df[~df.index.duplicated(keep='first')]  # 或 keep='last'
    return df

  def write_df2excel(self, df: pd.DataFrame,
                     fname='/Users/wangjinlong/job/teach/course_课程/clsjymo_材料设计与模拟/xgcl_相关材料/课堂花名册.xlsx',
                     sheet_name='sheet_new',
                     index=False):
    """
      直接用df.to_excel(writer, sheet_name=sheet_name)会覆盖所有其他sheet
    """

    if os.path.exists(fname):
      excel_obj = self.get_exel_obj(fname=fname)
      with pd.ExcelWriter(path=fname, mode="a", if_sheet_exists="new") as writer:
        if sheet_name not in excel_obj.sheet_names:
          df.to_excel(excel_writer=writer,
                      sheet_name=sheet_name,
                      index=index)
        else:
          print(f'{fname} 已存在表单: -> {sheet_name}')
    else:
      with pd.ExcelWriter(path=fname, mode="w",) as writer:
        df.to_excel(excel_writer=writer,
                    sheet_name=sheet_name,
                    index=index)

    print(f'df 写入{fname} 的表单{sheet_name}')

    return None

  def get_df_date_index_range(self,
                              start_date='2025-2-17',
                              periods=12*7):
    """获取 以日期为索引的 df
    """
    # 设置起始日期
    start = pd.to_datetime(start_date)
    # 创建一个日期范围，包含12周的日期
    dates = pd.date_range(start=start,
                          periods=periods, freq='D')
    # 创建 DataFrame，日期作为索引
    df = pd.DataFrame(dates, columns=["Date"])
    # 设置日期为索引
    df.set_index("Date", inplace=True)
    return df

  def get_df_calendar(self,
                      start_date='2025-6-1',
                      periods=30,
                      select_weekday_list=[5, 6],
                      strftime_format='%d/%m'
                      ):
    """用来修改教学日历, 获得日期 比如六月份的周六周日的日期
    """
    df = self.get_df_date_index_range(start_date=start_date,
                                      periods=periods)
    # 获取选择的 weekday
    select_df: pd.DataFrame = df[df.index.weekday.isin(
        select_weekday_list)].copy()  # 0是星期一，4是星期五
    select_df['日期'] = select_df.index.strftime(strftime_format)
    return select_df

  def read_pickle(self, fname_pyobj):
    """反序列化"""
    pyobj = pd.read_pickle(fname_pyobj)
    print(f'fname_pyobj -> {fname_pyobj}')
    return pyobj

  def to_pickle(self, df: pd.DataFrame, fname_obj):
    """序列化
    """
    df.to_pickle(fname_obj)
    print(f'文件保存为 -> {fname_obj}')
    return None

  def example_notes(self):
    string = """ * 一些例子
    import pandas as pd  # 导入包

    从ndarrays 创建 DataFrame
    #  df = pd.DataFrame(data, columns=data[0,:], index=labels_int)  # data是数据表，创建的时候就规定列名(columns)和行名(index)
    nme = ["Google", "Runoob", "Taobao", "Wiki"]
    st = ["www.google.com", "www.runoob.com",
        "www.taobao.com", "www.wikipedia.org"]
    ag = [90, 40, 80, 98]
    dic = {'name': nme, 'site': st, 'age': ag}
    df = pd.DataFrame(dic)
    df.astype("str")

    print(df["name"])  # print(df.name) 显示某一列也可以这么用 .属性的方式
    a = df.values  # 直接从df(pandas数据表对象)获得数组 变成np.ndarray对象
    # print(a, type(a))
    # --- 这种用法和matlab中的矩阵提取类似,可以用: 表示所有
    # print(df.loc[1:3, ['name', 'age']])  # 按照名称选取块
    # print(df.iloc[:, :-1])  # 按索引取选取块  # i 表示按照索引 取1:2 行，0:2列 -1 表示倒数第二行或列

    # print(df.index)
    # print(df.columns[1:3])
    # print(df.loc[[0, 1], ['site', 'age']])  # 显示块
    # print(df.loc[[1, 2]][['site', 'age']])
    # print(df.loc[[1, 2]])  # 显示多行 # print(df.loc[0:2])
    # print(df[['age', 'site']])  # 显示多列
    # print("行索引：{}".format(list(df.index)))
    # print("列索引：{}".format(list(df.columns)))


    # 读取csv文件
    # print(df.to_string())  # 返回所有数据
    df = df.loc[:, ['Name', 'Age', 'Weight']]  # 取出三列法1
    print(df)
    age_ave = df.loc[:, 'Age'].mean()  # 求平均值
    weight_ave = df['Weight'].mean()
    print(weight_ave)

    n = 0
    for line in df.index:  # 根据条件查找数据
        if df.loc[line, 'Age'] > age_ave:
            n += 1
            # print(df.loc[[line]])
    print(n)

    # 新增一列
    # df['score'] = [80, 98, 67, 90, 66, 70]  # 增加列的元素个数要跟原数据列的个数一样,
    # df.loc[:, 'score'] = [80, 98, 67, 90, 66, 70]
    df.loc[1:2, 'score'] = [34, 88]  # 新增部分列
    print(df)
    m = df.loc[:, 'score'].mean()
    print(m)

    # 新增一行
    # df.loc[6, :] = ['-', age_ave, weight_ave, 'unCalculate']
    df.loc['Average', ['Age', 'Weight']] = [age_ave, weight_ave]  # 新增部分行
    print(df)
    fname = 'dealed_nba.csv'
    df.to_csv(fname)
    print('*' * 40 + '\n' + f'{fname}保存完毕')

    # 插入一行数据， pandas 中没有插入行数据的方法，只有插入列数据的方法 df.insert(),
    # 插入行，只能借助于numpy 中的插入方法，如下
    df = pd.DataFrame(np.insert(arr=df.values, obj=n_line,
                      values="Selective dynamics", axis=0))


    df = df.rename(columns={"index": "sigma"})  # 修改列名
    df = df.rename(index={"index": "sigma"})  # 修改行名


    # 创建 pandas.Series 对象， pandas.Series(data, index, dtype, name, copy)
    # data： 可以是列表，字典, 元组
    # index: 给data数据自定义索引(默认是从0开始的数字作为索引)
    # dtype：数据类型，默认会自己判断。
    # name：设置名称。
    # copy：拷贝数据，默认为 False。

    a = (3, 4, 6, 'aa')
    sa = pandas.Series(a, index=['x', 'y', 'c', 'z'], name='mytest')
    print(sa)
    # a = ["Google", "Runoob", "Wiki"]
    # myvar1 = pd.Series(a)  # 通过列表创建
    # print(myvar1[0])
    # myvar2 = pd.Series(a, index=["x", "y", "z"])  # 指定索引值创建
    # print(myvar2['y'])  # 通过指定的索引或者是0，1，2都可以给出需要的内容
    sites = {'name': "Google", 2: "Runoob", 3: "Wiki"}  # 通过字典创建
    # myvar3 = pd.Series(sites, index=[1, 2])  # 需要字典中的一部分数据，只需要指定需要数据的索引即可
    myvar3 = pd.Series(sites)
    print(myvar3)
    # myvar4 = pd.Series(sites, index=[1, 2], name="RUNOOB-Series-TEST")  # 设置 Series 名称参数
    # print(myvar3)


    # 创建DataFrame 对象， 类似二维数组。 pandas.DataFrame(data, index, columns, dtype, copy)
    # 1. 从二维数组
    data = [['Google', 10], ['Runoob', 12], ['Wiki', 13]]
    # 使用列表创建, 也可以不指定列的名称 pd.DataFrame(data)
    df1 = pd.DataFrame(data, columns=['Site', 'Age'], index=['a', 'b', 'c'])
    # df1 = pd.DataFrame(data)
    # print(df1)

    # 2. 从 ndarrays 创建
    data = {"calories": [420, 380, 390], "duration": [50, 40, 45]}
    # df2 = pd.DataFrame(data)
    df2 = pd.DataFrame(data, index=["day1", "day2", "day3"])  # 指定索引值
    print(df2)

    # 3. 从字典创建
    data = [{'a': 1, 'b': 2}, {'a': 5, 'b': 10, 'c': 20}]
    df3 = pd.DataFrame(data)
    # print(df3)  # 没有对应的部分数据为 NaN


    # dataframe 对象的方法(显示方法)  查找
    data = {"calories": [420, 380, 390], "duration": [50, 40, 45]}
    df = pd.DataFrame(data)
    # print(df.to_string())  # 返回所有数据
    # print(df[['calories', 'duration']])  # 显示两列
    print(df.calories)   # 显示calories 这一列
    # print(df['calories'].head(3))  # 读取前面的 n 行，如果不填参数 n ，默认返回 5 行
    # print(df)  # 最多显示数据的前面 5 行和末尾 5 行，中间部分以 ... 代替。
    # print(df.loc[[0, 1], ['name', 'age']])  # 显示第1，2行，'name,'age'列的数据
    # print(df.loc[1, 'age'])  # 显示第2行'age'列的数据
    # print(df.loc[[0, 2]])  # 显示第一行和第三行数据 # 有索引的话必须指定索引值 print(df.loc[['day1', 'day2']])
    # print(df['age'].loc[1])  # 显示第2行'age'列的数据
    # print(df.tail(3))  # tail( n ) 方法用于读取尾部的 n 行
    # print(df.info())  # 返回表格的一些基本信息

    # 查找包含特定值的行内容
    df = pd.DataFrame({'A': 'foo bar foo bar foo bar foo foo'.split(),
                      'B': 'one one two three two two one three'.split(),
                      'C': np.arange(8), 'D': np.arange(8) * 2})
    # print(df)
    print(df.query('A=="foo" & B=="three"'))  # 多条件查找  | 或 & 与
    # 多种条件限制时使用&，&的优先级高于>=或<=，所以要注意括号的使用
    print(df.loc[(df['C'] >= 3) & (df['D'] <= 14)])
    print(df.loc[df["C"].isin([3, 5])])  # 除了可以在同一列使用多个 OR，你还可以使用.isin() 函数。
    # print(df[df['C'] == 6]) # 判断等式是否成立
    # print(df[df["B"] == "three"])
    # print(df.loc[df["B"] == "three"])  # 查找C列包含three的行
    # print(df.set_index('A', append=True, drop=False).xs('foo', level=1).set_index("B", drop=False).xs("two"))  # xs方法适用于多重索引DataFrame的数据筛选

    # 解决pandas中打印DataFrame行列显示不全的问题
    # pandas设置最大显示行和列
    pd.set_option('display.max_columns', 50)  # 最多显示50列，可以设置为None 表示显示所有列
    pd.set_option('display.width', 150)  # 每列显示150个字符  # 调整显示宽度，以便整行显示

    # pd.set_option('display.max_rows', 300)  # 可以设置为None 表示显示所有行
    # 设置value的显示长度为100，默认为50
    pd.set_option('max_colwidth', 100)  # 设置单列的显示最大显示宽度

    # 清洗数据
    # -- 查找和确定空数据的行
    df = pd.read_csv('property-data.csv')
    print(df)
    df.replace(to_replace="--", value=np.nan,
              inplace=True)  # 用value的值替换to_replace的值
    print(df)
    print(df.dropna())

    # print(df['SQ_FT'].isnull())  # 该列第二行'--'没有被认为空
    df = pd.read_csv('property-data.csv', na_values=['--', 'n/a'])  # 指定空数据类型
    # print(df['SQ_FT'])

    # print(df['SQ_FT'].isnull())  # 该列第二行'--'被认为空

    # -- 删除包含空数据的行
    # new_df = df.dropna()  # 删除包含空数据的行。dropna() 方法默认返回一个新的 DataFrame。
    # print(new_df.to_string())
    # df.dropna(inplace = True) # 如果你要修改源数据使用 inplace = True 参数
    df = pd.read_csv('property-data.csv')
    new_df:pd.DataFrame = df.dropna(
        subset=['ST_NUM', 'SQ_FT'])  # 移除 ST_NUM 和SQ_FT 列中字段值为空的行
    print(new_df)
    print(new_df.fillna(0))
    # -- 替换空字段
    # df['PID'].fillna(145, inplace=True)  # 使用 12345 替换PID列的空字段
    # print(df)

    # Pandas使用 mean()、median() 和 mode() 方法计算列的均值、中位数值（排序后排在中间的数）和众数（出现频率最高的数）。
    x = df['ST_NUM'].mean()
    df["ST_NUM"].fillna(int(x), inplace=True)  # 使用 mean() 方法计算列的均值并替换空单元格：
    # print(df.to_string())

    # -- 去除重复数据  df.duplicated()
    person = {
        "name": ['Google', 'Runoob', 'Runoob', 'Taobao', 'Runoob'],
        "age": [50, 40, 40, 23, 40]
    }
    df = pd.DataFrame(person)
    # print(df.duplicated())  # 如果对应的数据是重复的，duplicated() 会返回 True，否则返回 False。
    df.drop_duplicates(inplace=True)  # 删除重复数据
    # print(df)

    # 将列中的所有单元格转换为相同格式的数据。日期处理
    data = {
        "Date": ['2020/12/01', '2020/12/02', '20201226'],  # 第三个日期格式错误
        "duration": [50, 40, 45]
    }
    df = pd.DataFrame(data) # , index=["day1", "day2", "day3"])
    print(df)
    # print(pd.to_datetime(df['Date']))
    df['Date'] = pd.to_datetime(df['Date'])  # 更改Date列
    print(df)

    # 根据条件替换数据
    person = {
        "name": ['Google', 'Runoob', 'Taobao'],
        "age": [50, 200, 12345]
    }
    df = pd.DataFrame(person)
    for x in df.index:  #
        if df.loc[x, "age"] > 120:
            df.loc[x, "age"] == 120
    # print(df.to_string())

    # 调整列的顺序
    order = ['date', 'time', 'open', 'high',
        'low', 'close', 'volumefrom', 'volumeto']
    df = df[order]


    # 直接用df.to_excel(writer, sheet_name=sheet_name)会覆盖所有其他sheet
    with pd.ExcelWriter(path=self.outfile, mode="a", if_sheet_exists="new") as writer:
        if self.out_sheet_name not in pd.ExcelFile(self.filepath).sheet_names:
            df.to_excel(excel_writer=writer,
                            sheet_name=self.out_sheet_name)
            print(f"保存到{self.outfile}-->{self.out_sheet_name}")
        else:
            print(f"已经存在sheet_name: {self.out_sheet_name}")
    """
    print(string)
    return None

  def drop(self, df: pd.DataFrame):
    df.drop(labels=['B_N2_graphene', 'B2_N4_graphene'])  # 删除多行
    df.drop(labels='N2_B3_graphene')  # 删除一行
    pass

  def to_excel_指定宽度(self, df: pd.DataFrame,
                    fname='xxx.xlsx',
                    col_width=11):
    """
      保存为pdf时, 指定列宽
    """
    writer = pd.ExcelWriter(fname, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    worksheet = writer.sheets['Sheet1']
    worksheet.column_dimensions['B'].width = col_width  # 指定'B' 列的宽度
    for col in 'defghijklmnop'.upper():
      worksheet.column_dimensions[col].width = col_width
    writer.close()

  def to_excel_插入标题行(self, df,
                     fname_out,
                     sheet_name='Sheet1',
                     l1="铜陵学院学生平时成绩登记表",
                     l2="开课学期: 2023-2024-2  课程: 大学物理(上)   教师: 玉金龙   班级: 23材料成型及控制工程班",
                     ):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # 写入数据（从第3行开始）
    with pd.ExcelWriter(fname_out, engine="openpyxl") as writer:
      df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

    # 用 openpyxl 编辑前两行标题
    wb = load_workbook(fname_out)
    ws = wb[sheet_name]

    # 合并单元格
    max_col = ws.max_column
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    # 写入标题
    ws["A1"] = l1
    ws["A2"] = l2
    # 设置居中格式
    align_center = Alignment(horizontal='center', vertical='center')
    ws["A1"].alignment = align_center
    ws["A2"].alignment = align_center

    # 自动设置列宽（可选）
    for i, col in enumerate(ws.iter_cols(min_row=3, max_row=3), start=1):
      max_len = max(len(str(cell.value)) for cell in col)
      col_letter = get_column_letter(i)
      ws.column_dimensions[col_letter].width = max_len + 2

    # 保存
    wb.save(fname_out)
