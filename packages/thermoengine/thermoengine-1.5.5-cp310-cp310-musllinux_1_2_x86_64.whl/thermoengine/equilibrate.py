'''
The equilibrate module implements a Python interface to the Equilibrate and
EquilState classes. It also implements a Python class called MELTSmodel that
wraps the objective-C classes: EquilibrateUsingMELTSv102,
EquilibrateUsingMELTSv110, EquilibrateUsingMELTSv120,
EquilibrateUsingpMELTSv561, EquilibrateUsingMELTSwithDEW,
and EquilibrateUsingStixrude.

The Equilibrate class provides methods to calculate an equilibrium phase
assemblage given a list of Phase class instances.

You can calculate equilibrium in closed thermodynamic systems under a
variety of constraints:

- Temperature and pressure (Gibbs free energy minimization)

- Entropy and pressure (enthalpy minimization)

- Temperature and volume (Helmholtz free energy minimization)

- Entropy and volume (internal energy minimization)

You can also calculate equilibrium under constraints of fixed temperature and
pressure in open thermodynamic systems by specifying one or more fixed elemental
chemical potentials.

For details of the underlying theory and algorithms implemented in the Equilibrate
class, see the notebooks in the PublicNotebooks/Equilibrate folder on
the `ENKI server <https://enki.ofm-research.org/hub/login>`_.
'''
from __future__ import annotations  # Enable Python 4 type hints in Python 3

from collections import OrderedDict

from thermoengine_utils.core import chem, UnorderedList
from thermoengine.samples import SampleMesh, SampleMaker, _PhaseSample, \
    SampleLibrary, Assemblage, MonophaseAssemblage
from thermoengine import core
from thermoengine import phases as phs
from thermoengine import chemistry
from thermoengine.const import units
from thermoengine import model



import collections
import abc
from typing import Type, List, Optional, Dict, Union
import collections
from nptyping import NDArray, Shape, Float


import numpy as np
import scipy as sp
import sympy as sym
import pandas as pd
import scipy.optimize as opt
import scipy.interpolate as interp
import warnings
import copy



with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    import statsmodels.api as sm

import xml.etree.ElementTree as ET
import locale
import sys

from enum import Enum, auto


# Excel imports
from openpyxl import Workbook

import time

__all__ = ['Equilibrate', 'EquilState', 'MELTSmodel']

# array length for elements
NE = 106

class EquilState:
    """Class for holding the phase state for an equilibrium model determined by Equilibrate

    Parameters
    ----------
    elements_l : []
        A list of strings identifying the chemical elements present in the
        system
    phases_l : []
        A list of class instances for phases that are in the system.  These
        instances must derive from the *PurePhase* or *SolutionPhase* classes,
        which are defined in the *phases* module.

    Attributes
    ----------
    c_matrix
    c_omni
    c_omni_qr
    c_qr_decomp
    element_l
    phase_d
    pressure
    temperature

    """

    def __init__(self, elements_l, phases_l):
        elements_present = [False for i in range(0, NE)]
        for elm in elements_l:
            index = core.chem.PERIODIC_ORDER.tolist().index(elm)
            elements_present[index] = True
        self._element_l = [i for i, e in enumerate(elements_present) if e]

        self._phase_d = OrderedDict()
        for phase in phases_l:
            entry = dict()
            entry['obj'] = phase
            nc = int(phase.props['endmember_num']) \
                if 'endmember_num' in phase.props else 1
            entry['name'] = phase.props['phase_name']
            entry['abbrev'] = phase.props['abbrev']
            entry['nc'] = nc
            entry['end_name'] = phase.props['endmember_name']
            entry['end_formula'] = phase.props['formula']
            entry['end_molwts'] = phase.props['molwt']
            entry['moles'] = np.zeros(nc)
            entry['mole_frac'] = np.zeros(nc)
            entry['mole_nz'] = np.ones(nc)
            entry['grams'] = 0.0
            entry['affinity'] = 0.0
            # create a list of length nc where each entry is a numpy array
            # that contains stoichiometric coefficients converting each
            # endmember to system elements (in the order _element_l)
            conv_to_elm = []
            conv_to_elm_sum = np.zeros(len(self._element_l))
            conv_to_oxd = []
            for i in range(0, nc):
                mol = np.zeros(nc)
                mol[i] = 1.0
                if nc == 1:
                    mol_elm = phase.props['element_comp'][0]
                else:
                    mol_elm = entry['obj'].convert_endmember_comp(
                        mol, output='moles_elements')
                conv_entry = np.zeros(len(self._element_l))
                for i in range(0, NE):
                    if mol_elm[i] != 0.0:
                        conv_entry[self._element_l.index(i)] = mol_elm[i]
                conv_to_elm_sum = np.add(conv_to_elm_sum, conv_entry)
                conv_to_elm.append(conv_entry)
                conv_to_oxd.append(core.chem.calc_mol_oxide_comp(mol_elm))
            entry['conv_to_elm'] = conv_to_elm
            entry['conv_to_oxd'] = conv_to_oxd
            entry['omnicomponent'] = True if np.count_nonzero(
                conv_to_elm_sum) == len(conv_to_elm_sum) else False
            entry['allow_unmixing'] = True if not entry['omnicomponent'] else False
            # phase_module = entry['obj'].module
            # if phase_module is not None:
            #     method_match = [key for key in phase_module.__dict__.keys() if 'allow_unmixing' in key]
            #     if len(method_match) > 0:
            #         method = getattr(phase_module, method_match[0], None)
            #         entry['allow_unmixing'] = method()
            name = phase.props['phase_name']
            self._phase_d[name] = entry

        first = True
        for entry in self._phase_d.values():
            nc = entry['nc']
            for row in entry['conv_to_elm']:
                if first:
                    c_matrix = np.reshape(row, (1, row.size))
                    first = False
                else:
                    c_matrix = np.vstack((c_matrix, np.reshape(
                        row, (1, row.size))))
        self._c_matrix = c_matrix
        self._c_qr_decomp = np.linalg.qr(self.c_matrix)
        self._c_omni = None
        self._c_omni_qr = None
        self._temperature = 1000.0
        self._pressure = 1000.0

    ############################
    # Properties of the system #
    ############################

    @property
    def phase_d(self):
        """
        A dictionary of dictionaries that holds properties of system phases

        Returns
        -------
        Dictionary of phases in the system (dict)
        """
        return self._phase_d

    @property
    def element_l(self):
        """
        List of atomic numbers of elements in the system

        Returns
        -------
        A Python list
        """
        return self._element_l

    @property
    def c_matrix(self):
        """
        Numpy matrix that maps elements to moles of endmembers

        Moles of endmembers are indexed by row, and moles of elements
        are indexed by column.

        Returns
        -------
        Numpy 2-D array
        """
        return self._c_matrix

    @property
    def c_qr_decomp(self):
        """
        Tuple of Numpy matrices of the Q,R decomposition of the c_matrix

        Returns
        -------
        tuple
        """
        return self._c_qr_decomp

    @property
    def c_omni(self):
        """
        Numpy conversion matrix for the omnicomponent phase (if one exists)

        Moles of endmembers are indexed by row, and moles of elements
        are indexed by column.

        Returns
        -------
        np 2-D array
        """
        return self._c_omni

    @property
    def c_omni_qr(self):
        """
        Tuple of Numpy matrices of the Q,R decomposition of c_omni

        Returns
        -------
        tuple
        """
        return self._c_omni_qr

    @property
    def temperature(self):
        """
        Temperature of the assemblage, in Kelvins

        Returns
        -------
        float
        """
        return self._temperature

    @temperature.setter
    def temperature(self, value):
        assert isinstance(value, (int, float))
        assert value > 0
        self._temperature = value

    @property
    def pressure(self):
        """
        Pressure of the assemblage, in bars

        Returns
        -------
        float
        """
        return self._pressure

    @pressure.setter
    def pressure(self, value):
        assert isinstance(value, (int, float))
        assert value > 0
        self._pressure = value

    ################################
    # Methods for phases in system #
    ################################

    def omni_phase(self):
        """
        Returns first omnicomponent phase in the phase dictionary
        """
        for entry in self.phase_d.values():
            if entry['omnicomponent']:
                return entry['name']
        return None

    def c_matrix_for_phase(self, phase_name):
        """
        Returns a slice of c_matrix relevant to the specified phase_name

        Parameters
        ----------
        phase_name : str
            Name of a system phase
        """
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return None
        row_min = 0
        for key, entry in self.phase_d.items():
            if key == phase_name:
                row_max = row_min + entry['nc']
                break
            else:
                row_min += entry['nc']
        return self.c_matrix[row_min:row_max, :]

    def set_phase_comp(self, phase_name, cmp, input_as_elements=False):
        """
        Sets the endmember moles and total moles a phase in the system

        Parameters
        ----------
        phase_name : str
            Name of a system phase
        cmp : ndarray
            1-D Numpy array with compositional data
        input_as_elements : bool, def False
            If True, convert input array from moles of elements to moles of
            endmember components.

        Returns
        -------
        valid : bool
            True if input composition is valid for phase
        """
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return False
        assert type(cmp) is np.ndarray, \
            'bulk_comp must be set to an numpy array'
        if not input_as_elements:
            assert cmp.size == self.phase_d[phase_name]['nc'], \
                'cmp array must have length ' \
                + str(self.phase_d[phase_name]['nc'])
        else:
            assert cmp.size == len(self.element_l), \
                'cmp array must have length ' + str(len(self.element_l))
            cT = self.c_matrix_for_phase(phase_name).T
            cTinv = np.linalg.inv(cT) if cT.shape[0] == cT.shape[1] else np.linalg.pinv(cT)
            cmp = np.matmul(cTinv, cmp)
        self.phase_d[phase_name]['moles'] = cmp
        np.put(self.phase_d[phase_name]['mole_nz'],
               np.flatnonzero(np.abs(cmp) < 10.0 * np.finfo(float).eps), 0.0)
        return self.phase_d[phase_name]['obj'].test_endmember_comp(cmp)

    def tot_moles_phase(self, phase_name):
        """
        Returns total moles of phase

        Parameters
        ----------
        phase_name : str
            Name of a system phase
        """
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return 0.0
        return np.sum(self.phase_d[phase_name]['moles'])

    def tot_grams_phase(self, phase_name):
        """
        Returns total grams of phase

        Parameters
        ----------
        phase_name : str
            Name of a system phase
        """
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return 0.0
        grams = np.matmul(self.phase_d[phase_name]['moles'],
                          self.phase_d[phase_name]['end_molwts'])
        return np.sum(grams)

    def moles_elements(self, phase_name):
        """
        Returns array of moles of elements in phase

        Parameters
        ----------
        phase_name : str
            Name of a system phase

        Returns
        -------
        result : np_array
            Numpy array of mole numbers of the reduced set of elements in the
            phase
        """
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return None
        elm = np.zeros(len(self.element_l))
        for moles, coeff in zip(self.phase_d[phase_name]['moles'],
                                self.phase_d[phase_name]['conv_to_elm']):
            if moles != 0:
                elm = np.add(elm, moles * coeff)
        return elm

    def tot_moles_elements(self):
        """
        Returns array of moles of elements in system

        Returns
        -------
        result : np_array
            Numpy array of mole numbers of the reduced set of elements in the
            system
        """
        elm = np.zeros(len(self.element_l))
        for phase in self.phase_d.keys():
            if self.tot_moles_phase(phase) > 0:
                elm += self.moles_elements(phase)
        return elm

    def oxide_comp(self, phase_name, output='wt%', prune_list=True):
        """
        Returns a dictionary of concentrations of oxides in phase

        Parameters
        ----------
        phase_name : str
            Name of a system phase
        output : str, default "wt%"
            Units of output: 'wt%' (default), 'moles', 'grams'
        prune_list : bool, default True
            Remove zeroed valued entries

        Returns
        -------
        result : collections.OrderedDict
            Dictionary of oxide concentration values, keyed by oxide
        """
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return None
        oxd = np.zeros(core.chem.oxide_props['oxide_num'])
        for moles, coeff in zip(self.phase_d[phase_name]['moles'],
                                self.phase_d[phase_name]['conv_to_oxd']):
            if moles != 0:
                oxd = np.add(oxd, moles * coeff)
        result = OrderedDict()
        for key, value in zip(core.chem.oxide_props['oxides'], oxd):
            result[key] = value
        if output == 'wt%' or output == 'grams':
            mod_result = OrderedDict()
            sum_oxd = 0.0
            for index, (key, value) in enumerate(result.items()):
                mod_result[key] = value * chem.oxide_props['molwt'][index]
                sum_oxd += mod_result[key]
            result = mod_result
            if output == 'wt%' and sum_oxd > 0:
                mod_result = OrderedDict()
                for (key, value) in result.items():
                    mod_result[key] = 100.0 * value / sum_oxd
                result = mod_result
        if prune_list:
            mod_result = OrderedDict()
            for (key, value) in result.items():
                if value != 0:
                    mod_result[key] = value
            result = mod_result
        return result

    def properties(self, phase_name=None, props=None, units=False):
        """
        Returns properties of phases in the system

        Parameters
        ----------
        phase_name : str, default None
            Name of phase in system. If None, returns a dictionary of system
            phases, with names as keys and 'stable' or 'unstable' as values.
        props : str, default None
            Name of property to be retrieved. If None, a list of valid properties
            is returned.
        units : bool, default False
            Property units are provided as the second entry of a returned tuple.
        """
        prop_d = OrderedDict([('Mass', 'g'), ('GibbsFreeEnergy', 'J'),
                              ('Enthalpy', 'J'), ('Entropy', 'J/K'), ('HeatCapacity', 'J/K'),
                              ('DcpDt', 'J/K^2'), ('Volume', 'J/bar'), ('DvDt', 'J/bar-K'),
                              ('DvDp', 'J/bar^2'), ('D2vDt2', 'J/bar-K^2'), ('D2vDtDp', 'J/bar^2-K'),
                              ('D2vDp2', 'J/bar^3'), ('Density', 'g/cm^3'), ('Alpha', '1/K'),
                              ('Beta', '1/bar'), ('K', 'GPa'), ("K'", 'none'), ('Gamma', 'none')])

        if phase_name is None:
            phase_d = {}
            for key in self.phase_d.keys():
                phase_d[key] = 'stable' if self.tot_moles_phase(
                    key) > 0 else 'unstable'
            phase_d['System'] = 'stable'
            return phase_d
        if phase_name not in self.phase_d and phase_name != 'System':
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return None
        if props is None:
            if units:
                return set(zip(list(prop_d.keys()), list(prop_d.values())))
            else:
                return list(prop_d.keys())
        if props not in prop_d:
            print('Error: ' + props + ' is not a valid property.')
            return None

        t = self.temperature
        p = self.pressure
        if phase_name != 'System':
            entry = self.phase_d[phase_name]
            moles = entry['moles']
            if np.sum(moles) == 0:
                result = 0.0
                if units:
                    result = (0.0, None)
                return result
            nc = entry['nc']
            sys = False
        else:
            sys = True
        result = None
        if units:
            result = (None, None)

        if props == 'Mass':
            if sys:
                result = 0.0
                for phs in self.phase_d.keys():
                    grams = np.matmul(self.phase_d[phs]['moles'],
                                      self.phase_d[phs]['end_molwts'])
                    result += np.sum(grams)
            else:
                grams = np.matmul(self.phase_d[phase_name]['moles'],
                                  self.phase_d[phase_name]['end_molwts'])
                result = np.sum(grams)
            if units:
                result = (result, prop_d['Mass'])
        elif props == 'GibbsFreeEnergy':
            if sys:
                result = self.G(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].gibbs_energy(t, p)
                else:
                    result = entry['obj'].gibbs_energy(t, p, mol=moles)
            if units:
                result = (result, prop_d['GibbsFreeEnergy'])
        elif props == 'Enthalpy':
            if sys:
                result = self.G(t, p) - t * self.dGdT(t, p)
            else:
                if nc == 1:
                    result = moles[0] * (entry['obj'].gibbs_energy(t, p)
                                         + t * entry['obj'].entropy(t, p))
                else:
                    result = (entry['obj'].gibbs_energy(t, p, mol=moles)
                              + t * entry['obj'].entropy(t, p, mol=moles))
            if units:
                result = (result, prop_d['Enthalpy'])
        elif props == 'Entropy':
            if sys:
                result = -self.dGdT(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].entropy(t, p)
                else:
                    result = entry['obj'].entropy(t, p, mol=moles)
            if units:
                result = (result, prop_d['Entropy'])
        elif props == 'HeatCapacity':
            if sys:
                result = -t * self.d2GdT2(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].heat_capacity(t, p)
                else:
                    result = entry['obj'].heat_capacity(t, p, mol=moles)
            if units:
                result = (result, prop_d['HeatCapacity'])
        elif props == 'DcpDt':
            if sys:
                result = -t * self.d3GdT3(t, p) - self.d2GdT2(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].heat_capacity(
                        t, p, deriv={'dT': 1})
                else:
                    result = entry['obj'].heat_capacity(
                        t, p, mol=moles, deriv={'dT': 1})
            if units:
                result = (result, prop_d['DcpDt'])
        elif props == 'Volume':
            if sys:
                result = self.dGdP(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].volume(t, p)
                else:
                    result = entry['obj'].volume(t, p, mol=moles)
            if units:
                result = (result, prop_d['Volume'])
        elif props == 'DvDt':
            if sys:
                result = self.d2GdTdP(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].volume(t, p, deriv={'dT': 1})
                else:
                    result = entry['obj'].volume(t, p, mol=moles, deriv={'dT': 1})
            if units:
                result = (result, prop_d['DvDt'])
        elif props == 'DvDp':
            if sys:
                result = self.d2GdP2(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].volume(t, p, deriv={'dP': 1})
                else:
                    result = entry['obj'].volume(t, p, mol=moles, deriv={'dP': 1})
            if units:
                result = (result, prop_d['DvDp'])
        elif props == 'D2vDt2':
            if sys:
                result = self.d3GdT2dP(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].volume(t, p, deriv={'dT': 2})
                else:
                    result = entry['obj'].volume(t, p, mol=moles, deriv={'dT': 2})
            if units:
                result = (result, prop_d['D2vDt2'])
        elif props == 'D2vDtDp':
            if sys:
                result = self.d3GdTdP2(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].volume(
                        t, p, deriv={'dT': 1, 'dP': 1})
                else:
                    result = entry['obj'].volume(t, p, mol=moles,
                                                 deriv={'dT': 1, 'dP': 1})
            if units:
                result = (result, prop_d['D2vDtDp'])
        elif props == 'D2vDp2':
            if sys:
                result = self.d3GdP3(t, p)
            else:
                if nc == 1:
                    result = moles[0] * entry['obj'].volume(t, p, deriv={'dP': 2})
                else:
                    result = entry['obj'].volume(t, p, mol=moles, deriv={'dP': 2})
            if units:
                result = (result, prop_d['D2vDp2'])
        elif props == 'Density':
            v = self.properties(phase_name=phase_name, props='Volume',
                                units=False)
            m = self.properties(phase_name=phase_name, props='Mass',
                                units=False)
            result = m / v / 10.0
            if units:
                result = (result, prop_d['Density'])
        elif props == 'Alpha':
            v = self.properties(phase_name=phase_name, props='Volume',
                                units=False)
            dvdt = self.properties(phase_name=phase_name, props='DvDt',
                                   units=False)
            result = 10.0 * dvdt / v
            if units:
                result = (result, prop_d['Alpha'])
        elif props == 'Beta':
            v = self.properties(phase_name=phase_name, props='Volume',
                                units=False)
            dvdp = self.properties(phase_name=phase_name, props='DvDp',
                                   units=False)
            result = -10.0 * dvdp / v
            if units:
                result = (result, prop_d['Beta'])
        elif props == 'K':
            beta = self.properties(phase_name=phase_name, props='Beta',
                                   units=False)
            result = 1.0 / beta / 10000.0
            if units:
                result = (result, prop_d['K'])
        elif props == "K'":
            v = self.properties(phase_name=phase_name, props='Volume',
                                units=False)
            dvdp = self.properties(phase_name=phase_name, props='DvDp',
                                   units=False)
            d2vdp2 = self.properties(phase_name=phase_name, props='D2vDp2',
                                     units=False)
            result = v * d2vdp2 / dvdp / dvdp - 1
            if units:
                result = (result, prop_d["K'"])
        elif props == 'Gamma':
            v = self.properties(phase_name=phase_name, props='Volume',
                                units=False)
            dvdp = self.properties(phase_name=phase_name, props='DvDp',
                                   units=False)
            dvdt = self.properties(phase_name=phase_name, props='DvDt',
                                   units=False)
            cp = self.properties(phase_name=phase_name, props='HeatCapacity',
                                 units=False)
            result = (-v * dvdt / dvdp) / (cp + t * dvdt * dvdt / dvdp)
            if units:
                result = (result, prop_d['Gamma'])
        return result

    def compositions(self, phase_name=None, ctype='components', units='moles'):
        """
        Returns compositions of phases in the system

        Parameters
        ----------
        phase_name : str, default None
            Name of phase in system
        ctype : str, default 'components'
            Compositional descriptor. Permitted: 'components', 'oxides',
            'elements'
        units : str, default 'moles'
            Units of composition. Permitted: 'moles', 'mole_frac', 'wt%'

        Returns
        -------
        result : numpy array
            One dimensional vector

        Notes
        -----
        If units is set to 'oxides' or 'elements', results are returned in an
        array of standard length and order.
        """
        if phase_name is None:
            print('Error: Method requires a phase_name.')
            return None
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return None
        if ctype not in ['components', 'oxides', 'elements']:
            print('Error: ' + ctype + ' is not permitted. See doc string.')
            return None
        if units not in ['moles', 'mole_frac', 'wt%']:
            print('Error: ' + units + ' is not permitted. See doc string.')
            return None

        entry = self.phase_d[phase_name]
        if ctype == 'components':
            if units == 'moles' or units == 'mole_frac':
                result = entry['moles']
                if units == 'mole_frac':
                    result = result / np.sum(result)
            elif units == 'wt%':
                result = entry['moles'] * entry['end_molwts']
                result = 100.0 * result / np.sum(result)
        elif ctype == 'oxides':
            if units == 'moles' or units == 'mole_frac':
                result = self.oxide_comp(phase_name, output='moles',
                                         prune_list=False)
                result = np.array(list(result.values()))
                if units == 'mole_frac':
                    result = result / np.sum(result)
            elif units == 'wt%':
                result = self.oxide_comp(phase_name, output='wt%',
                                         prune_list=False)
                result = np.array(list(result.values()))
        elif ctype == 'elements':
            moles = entry['moles']
            if entry['nc'] == 1:
                result = moles[0] * entry['obj'].props['element_comp'][0]
            else:
                result = np.zeros(NE)
                for i in range(0, entry['nc']):
                    mol = np.zeros(entry['nc'])
                    mol[i] = 1.0
                    mol_elm = entry['obj'].covert_endmember_comp(
                        mol, output='moles_elements')
                    for j in range(1, NE):
                        result[j] += moles[i] * mol_elm[j]
            if units == 'mole_frac':
                result = result / np.sum(result)
            elif units == 'wt%':
                for i in range(1, NE):
                    result[i] *= chem.PERIODIC_WEIGHTS[i]
                result = 100.0 * result / np.sum(result)
        return result

    def affinities(self, phase_name=None):
        """
        Returns affinity of phases in the system

        Parameters
        ----------
        phase_name : str, default None
            Name of phase in system. If None, returns a dictionary of system
            phases, with names as keys and 'stable' or 'unstable' as values.
        """
        if phase_name is None:
            print('Error: Please specify a phase name.')
            return None
        if phase_name not in self.phase_d:
            print('Error: ' + phase_name + ' is not in phase dictionary.')
            return None
        tot_mol = self.tot_moles_phase(phase_name)
        if tot_mol > 0:
            return 0.0
        else:
            return self.phase_d[phase_name]['affinity']

    def print_state(self, level='summary', wt_as_oxides=True):
        """
        Prints results about the system state

        Parameters
        ----------
        level : str
            Level of detail to be printed
        wt_as_oxides : bool, default True
            Print wt% values on an oxide basis; otherwise print wt% of
            endmember components.
        """
        print(' ')
        print('T = {0:10.2f} °C, P = {1:10.1f} MPa'.format(
            self.temperature - 273.15, self.pressure / 10.0))
        for key, entry in self.phase_d.items():
            tot_mol = self.tot_moles_phase(key)
            if tot_mol > 0:
                tot_grm = self.tot_grams_phase(key)
                print('{0:<15s}'.format(entry['name']), end=' ')
                print('moles: {0:10.6f}'.format(tot_mol), end=' ')
                print('grams: {0:7.3f}'.format(tot_grm))
                if wt_as_oxides:
                    oxd = self.oxide_comp(key)
                    oxd_s = list(oxd.keys())
                    oxd_v = list(oxd.values())
                    n_oxd = len(oxd)
                nc = entry['nc']
                if nc > 1:
                    for i in range(0, nc):
                        gram = entry['moles'][i] * entry['end_molwts'][i]
                        print('{0:>15s}'.format(
                            entry['end_name'][i][:15]), end=' ')
                        print('form:  {0:<10s}'.format(
                            entry['end_formula'][i][:10]), end='     ')
                        print('X: {0:7.4f}'.format(
                            entry['moles'][i] / tot_mol), end='  ')
                        if wt_as_oxides:
                            if i < n_oxd:
                                print('wt% {0:>7s} {1:7.2f}'.format(
                                    oxd_s[i], oxd_v[i]))
                            else:
                                print('')
                        else:
                            print('wt% {0:7.2f}'.format(100.0 * gram / tot_grm))
                    if wt_as_oxides and nc < n_oxd:
                        for i in range(nc, n_oxd):
                            print('{0:49s} wt% {1:>7s} {2:7.2f}'.format(
                                ' ', oxd_s[i], oxd_v[i]))
            else:
                print('{0:<15s}'.format(entry['name']), end=' ')
                print(entry['affinity'])
                # print('affn: {0:10.2f}'.format(entry['affinity']))
                nc = entry['nc']
                if nc > 1:
                    for i in range(0, nc):
                        print('{0:>15s}'.format(entry['end_name'][i][:15]), end=' ')
                        print('form:  {0:<10s}'.format(
                            entry['end_formula'][i][:10]), end='     ')
                        print('X: {0:7.4f}'.format(entry['mole_frac'][i]))

    ###################################################
    # System potentials and compositional derivatives #
    ###################################################

    def moles_v(self, reaction_v):
        """
        Computes of a reaction product

        moles (scalar) = (mole vector of all endmembers of all active phases
        in the system) x (input mole vector of reaction coefficients)

        Parameters
        ----------
        reaction_v : ndarray
            Array of reaction coefficients

        Returns
        -------
        moles : float
            Moles of reaction product
        """
        moles = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                for x in entry['moles']:
                    moles.append(x)
        moles = np.array(moles)
        moles = np.matmul(reaction_v, moles)
        return moles

    def G(self, t=1000.0, p=1000.0):
        """
        Returns Gibbs free energy of the system; a scalar
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(t, p)
                else:
                    result += entry['obj'].gibbs_energy(t, p, mol=moles)
        return result

    def dGdT(self, t=1000.0, p=1000.0):
        """
        Returns the temperature derivative of the Gibbs free energy of the
        system; a scalar, and the negative of the system entropy
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += -moles[0] * entry['obj'].entropy(t, p)
                else:
                    result += -entry['obj'].entropy(t, p, mol=moles)
        return result

    def dGdP(self, t=1000.0, p=1000.0):
        """
        Returns the pressure derivative of the Gibbs free energy of the
        system; a scalar, and the system volume
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(
                        t, p, deriv={'dP': 1})
                else:
                    result += entry['obj'].gibbs_energy(
                        t, p, mol=moles, deriv={'dP': 1})
        return result

    def dGdn(self, t=1000.0, p=1000.0, element_basis=True, full_output=False,
             use_omni_phase=False):
        """
        Returns the first molar derivative of G of the system; a 1-D numpy array

        Parameters
        ----------
        element_basis : bool, def True
            If True, returns a projected array of element chemical potentials
        full_output : bool, def False
            If True, returns a tuple with full output from numpy lstlq method
        use_omni_phase : bool, def False
            If True, returns a projected array of element chemical potentials
            using the omnicomponent phase as a basis.  The system must have
            an omnicomponent phase, and element_basis must be set to True for
            this keyword to have effect.
        """
        if use_omni_phase:
            omni_phase_name = self.omni_phase()
            assert omni_phase_name, \
                'ERROR in dGdn method: use omni_phase keyword specified' + \
                ' when the system has no omnicomponent phase'
            element_basis = True

        if use_omni_phase and element_basis:
            # computes chemical potentials for all component endmembers
            entry = self.phase_d[omni_phase_name]
            result = []
            nc = entry['nc']
            if nc == 1:
                row = np.array([entry['obj'].gibbs_energy(t, p)])
            else:
                moles = entry['moles']
                row = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                deriv={'dmol': 1})[0]
                np.put(row, np.flatnonzero(entry['mole_nz'] == 0), 0.0)
            result.append(np.reshape(row, (nc, 1)))
            result = np.vstack(tuple(result))
        else:
            # computes chemical potentials for all component endmembers in a
            # system phase if the total moles of that phase is > 0
            result = []
            for entry in self.phase_d.values():
                if self.tot_moles_phase(entry['name']) > 0.0:
                    nc = entry['nc']
                    if nc == 1:
                        row = np.array([entry['obj'].gibbs_energy(t, p)])
                    else:
                        moles = entry['moles']
                        row = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                        deriv={'dmol': 1})[0]
                    result.append(np.reshape(row, (nc, 1)))
            result = np.vstack(tuple(result))

        if not element_basis:
            return result
        else:
            if use_omni_phase:
                if self.c_omni is None:
                    self._c_omni = self.c_matrix_for_phase(omni_phase_name)
                    self._c_omni_qr = np.linalg.qr(self.c_omni)
                Q, R = self.c_omni_qr
            else:
                first = True
                for key, entry in self._phase_d.items():
                    if self.tot_moles_phase(key) > 0.0:
                        nc = entry['nc']
                        for row in entry['conv_to_elm']:
                            if first:
                                c_matrix = np.reshape(row, (1, row.size))
                                first = False
                            else:
                                c_matrix = np.vstack((c_matrix, np.reshape(
                                    row, (1, row.size))))
                Q, R = np.linalg.qr(c_matrix)
            Qb = np.matmul(Q.T, result)
            result, residuals, rank, S = np.linalg.lstsq(R, Qb, rcond=None)
            return result if not full_output else (result, residuals, rank, S)

    def d2GdT2(self, t=1000.0, p=1000.0):
        """
        Returns the second temperature derivative of the Gibbs free energy of
        the system; a scalar, and the negative of the temperature times the
        system heat capacity
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += -moles[0] * entry['obj'].heat_capacity(t, p) / t
                else:
                    result += -entry['obj'].heat_capacity(t, p, mol=moles) / t
        return result

    def d2GdTdP(self, t=1000.0, p=1000.0):
        """
        Returns the cross temperature-pressure derivative of the Gibbs free
        energy of the system; a scalar, and the volume times the system's
        coefficient of isothermal expansion
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(
                        t, p, deriv={'dT': 1, 'dP': 1})
                else:
                    result += entry['obj'].gibbs_energy(
                        t, p, mol=moles, deriv={'dT': 1, 'dP': 1})
        return result

    def d2GdP2(self, t=1000.0, p=1000.0):
        """
        Returns the second pressure derivative of the Gibbs free energy of
        the system; a scalar, and the negative of the volume times the
        system's coefficient of isothermal compressibility
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(
                        t, p, deriv={'dP': 2})
                else:
                    result += entry['obj'].gibbs_energy(
                        t, p, mol=moles, deriv={'dP': 2})
        return result

    def d2GdTdn(self, t=1000.0, p=1000.0):
        """
        Returns the second derivative of G of the system with respect to
        temperature and mole numbers; a 1-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                nc = entry['nc']
                if nc == 1:
                    row = np.array([-entry['obj'].entropy(t, p)])
                else:
                    moles = entry['moles']
                    row = -entry['obj'].entropy(t, p, mol=moles,
                                                deriv={'dmol': 1})[0]
                result.append(np.reshape(row, (nc, 1)))
        result = np.vstack(tuple(result))
        return result

    def d2GdPdn(self, t=1000.0, p=1000.0):
        """
        Returns the second derivative of G of the system with respect to
        pressure and mole numbers; a 1-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                nc = entry['nc']
                if nc == 1:
                    row = np.array([entry['obj'].gibbs_energy(
                        t, p, deriv={'dP': 1})])
                else:
                    moles = entry['moles']
                    row = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                    deriv={'dP': 1, 'dmol': 1})[0]
                result.append(np.reshape(row, (nc, 1)))
        result = np.vstack(tuple(result))

        return result

    def _sym_to_dense(self, n, sym_m):
        den_m = np.empty((n, n))
        ind = 0
        for i in range(0, n):
            for j in range(i, n):
                entry = sym_m[ind]
                den_m[i, j] = entry
                den_m[j, i] = entry
                ind += 1
        return den_m

    def d2Gdn2(self, t=1000.0, p=1000.0, element_basis=True):
        """
        Returns the second molar derivative of G of system; a 2-D numpy array

        Parameters
        ----------
        element_basis : bool, def True
            If True, returns a projected Hessian on an element basis
        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                if entry['nc'] == 1:
                    mat = np.array([[0.0]])
                else:
                    moles = entry['moles']
                    mat = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                    deriv={'dmol': 2})[0]
                    if len(mat.shape) == 1:
                        mat = self._sym_to_dense(entry['nc'], mat)
                result.append(mat)
        result = sp.linalg.block_diag(*result)
        if not element_basis:
            return result
        else:
            first = True
            for key, entry in self._phase_d.items():
                if self.tot_moles_phase(key) > 0.0:
                    nc = entry['nc']
                    for row in entry['conv_to_elm']:
                        if first:
                            c_matrix = np.reshape(row, (1, row.size))
                            first = False
                        else:
                            c_matrix = np.vstack((c_matrix, np.reshape(
                                row, (1, row.size))))
            Q, R = np.linalg.qr(c_matrix)
            # Qb = np.matmul(Q.T, np.matmul(result, Q))
            Qb = np.matmul(Q.T, result)
            result, residuals, rank, S = np.linalg.lstsq(R, Qb, rcond=None)
            return result

    def d3GdT2dn(self, t=1000.0, p=1000.0):
        """
        Returns the third derivative of G of the system twice with respect to
        temperature and once with respect to mole numbers; a 1-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                nc = entry['nc']
                if nc == 1:
                    row = np.array([-entry['obj'].heat_capacity(t, p) / t])
                else:
                    moles = entry['moles']
                    row = -entry['obj'].heat_capacity(t, p, mol=moles,
                                                      deriv={'dmol': 1})[0] / t
                result.append(np.reshape(row, (nc, 1)))
        result = np.vstack(tuple(result))

        return result

    def d3GdP2dn(self, t=1000.0, p=1000.0):
        """
        Returns the third derivative of G of the system with respect to
        pressure twice and mole numbers once; a 1-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                nc = entry['nc']
                if nc == 1:
                    row = np.array([entry['obj'].gibbs_energy(
                        t, p, deriv={'dP': 2})])
                else:
                    moles = entry['moles']
                    row = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                    deriv={'dP': 2, 'dmol': 1})[0]
                result.append(np.reshape(row, (nc, 1)))
        result = np.vstack(tuple(result))

        return result

    def d3GdTdPdn(self, t=1000.0, p=1000.0):
        """
        Returns the third derivative of G of the system with respect to
        temperature, pressure and mole numbers; a 1-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                nc = entry['nc']
                if nc == 1:
                    row = np.array([entry['obj'].gibbs_energy(
                        t, p, deriv={'dT': 1, 'dP': 1})])
                else:
                    moles = entry['moles']
                    row = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                    deriv={'dT': 1, 'dP': 1, 'dmol': 1})[0]
                result.append(np.reshape(row, (nc, 1)))
        result = np.vstack(tuple(result))

        return result

    def d3GdTdn2(self, t=1000.0, p=1000.0):
        """
        Returns the first temperature and second molar derivative of G of
        system; a 2-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                if entry['nc'] == 1:
                    mat = np.array([[0.0]])
                else:
                    moles = entry['moles']
                    mat = -entry['obj'].entropy(t, p, mol=moles,
                                                deriv={'dmol': 2})[0]
                    if len(mat.shape) == 1:
                        mat = self._sym_to_dense(entry['nc'], mat)
                result.append(mat)
        result = sp.linalg.block_diag(*result)
        return result

    def d3GdPdn2(self, t=1000.0, p=1000.0):
        """
        Returns the first pressure and second molar derivative of G of system;
        a 2-D numpy array

        """
        result = []
        for entry in self.phase_d.values():
            if self.tot_moles_phase(entry['name']) > 0.0:
                if entry['nc'] == 1:
                    mat = np.array([[0.0]])
                else:
                    moles = entry['moles']
                    mat = entry['obj'].gibbs_energy(t, p, mol=moles,
                                                    deriv={'dP': 1, 'dmol': 2})[0]
                    if len(mat.shape) == 1:
                        mat = self._sym_to_dense(entry['nc'], mat)
                result.append(mat)
        result = sp.linalg.block_diag(*result)
        return result

    def d3GdT3(self, t=1000.0, p=1000.0):
        """
        Returns the third temperature derivative of the Gibbs free energy of
        the system; a scalar
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    cp = entry['obj'].heat_capacity(t, p)
                    dcpdt = entry['obj'].heat_capacity(t, p, deriv={'dT': 1})
                    result += moles[0] * (cp / t / t - dcpdt / t)
                else:
                    cp = entry['obj'].heat_capacity(t, p, mol=moles)
                    dcpdt = entry['obj'].heat_capacity(t, p, mol=moles,
                                                       deriv={'dT': 1})
                    result += cp / t / t - dcpdt / t
        return result

    def d3GdT2dP(self, t=1000.0, p=1000.0):
        """
        Returns the third pressure derivative of the Gibbs free energy of
        the system; a scalar
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(
                        t, p, deriv={'dT': 2, 'dP': 1})
                else:
                    result += entry['obj'].gibbs_energy(
                        t, p, mol=moles, deriv={'dT': 2, 'dP': 1})
        return result

    def d3GdTdP2(self, t=1000.0, p=1000.0):
        """
        Returns the third pressure derivative of the Gibbs free energy of
        the system; a scalar
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(
                        t, p, deriv={'dT': 1, 'dP': 2})
                else:
                    result += entry['obj'].gibbs_energy(
                        t, p, mol=moles, deriv={'dT': 1, 'dP': 2})
        return result

    def d3GdP3(self, t=1000.0, p=1000.0):
        """
        Returns the third pressure derivative of the Gibbs free energy of
        the system; a scalar
        """
        result = 0.0
        for entry in self.phase_d.values():
            moles = entry['moles']
            if np.sum(moles) > 0.0:
                if entry['nc'] == 1:
                    result += moles[0] * entry['obj'].gibbs_energy(
                        t, p, deriv={'dP': 3})
                else:
                    result += entry['obj'].gibbs_energy(
                        t, p, mol=moles, deriv={'dP': 3})
        return result

    def _sym_ten_to_dense(self, n, sym_t):
        den_t = np.empty((n, n, n))
        ind = 0
        for i in range(0, n):
            for j in range(i, n):
                for k in range(j, n):
                    entry = sym_t[ind]
                    den_t[i, j, k] = entry
                    den_t[j, i, k] = entry
                    den_t[i, k, j] = entry
                    den_t[j, k, i] = entry
                    den_t[k, i, j] = entry
                    den_t[k, j, i] = entry
                    ind += 1
        return den_t

    def d3Gdn3(self, t=1000.0, p=1000.0, phase=None, cmp=None):
        """
        Returns the third molar derivative of G system for the cmp component
        of the phase named phase as a 2-D numpy array

        Parameters
        ----------
        phase : str, def None
        cmp : int, def None
            for the phase named phase d3Gdn3[cmp][*][*] is returned
        """
        if phase is None or cmp is None:
            return None
        if self.tot_moles_phase(phase) <= 0.0:
            return None
        entry = self.phase_d[phase]
        if entry['nc'] == 1:
            mat = np.array([[0.0]])
        else:
            moles = entry['moles']
            mat = entry['obj'].gibbs_energy(t, p, mol=moles,
                                            deriv={'dmol': 3})[0]
            if len(mat.shape) == 1:
                mat = self._sym_ten_to_dense(entry['nc'], mat)
        nc = mat.shape[0]
        result = np.zeros((nc, nc))
        for i in range(0, nc):
            for j in range(i, nc):
                for k in range(j, nc):
                    if i == cmp:
                        result[j][k] = mat[i][j][k]
                        result[k][j] = mat[i][j][k]
                    elif j == cmp:
                        result[i][k] = mat[i][j][k]
                        result[k][i] = mat[i][j][k]
                    elif k == cmp:
                        result[i][j] = mat[i][j][k]
                        result[j][i] = mat[i][j][k]
        return result

class Equilibrate:
    """Class for minimizing a generic thermodynamic potential in order to
    calculate an equilibrium phase assemblage.

    The default potential is the Gibbs free energy.

    Parameters
    ----------
    element_l : [], default None
        See documentation for element_list attribute.
    phase_l : [], default None
        See documentation for phase_list attribute.
    lagrange_l : [], default None
        See documentation for lagrange_list attribute.

    Attributes
    ----------
    A_omni_inv
    bulk_comp
    CTf
    element_list
    entropy
    eps_linear
    eps_minimal_energy
    eps_quad_optimal
    eps_quad_suboptimal
    eps_rank
    equil_cycle_max
    equil_linear_min
    lagrange_list
    lagrange_moles
    lagrange_no_mol_deriv
    lagrange_use_omni
    max_linear_iters
    max_quad_iters
    moles_in
    moles_out
    phase_list
    phase_separ_threshold
    reactions
    rotate_orthog_proj
    use_numpy_lstsq
    volume
    VT_null

    Notes
    -----
    Alternate potentials are specified by stipulating appropriate Lagrange
    transformations of the Gibbs potential.
    """

    def __init__(self, element_l=None, phase_l=None, lagrange_l=None):
        self._element_list = element_l
        self._phase_list = phase_l
        self._lagrange_list = lagrange_l
        self._bulk_comp = None
        self._moles_in = 1.0e-5
        self._moles_out = 1.0e-8
        self._max_quad_iters = 100
        self._max_linear_iters = 100
        self._phase_separ_threshold = -0.1
        self._entropy = False
        self._volume = False
        self._CTf = None
        self._VT_null = None
        self._reactions = None
        self._lagrange_moles = None
        self._lagrange_use_omni = True
        self._rotate_orthog_proj = False
        self._A_omni_inv = None
        self._lagrange_no_mol_deriv = False
        self._eps_linear = 10.0 * np.finfo(float).eps
        self._eps_quad_optimal = pow(np.finfo(float).eps, 2. / 3.)
        self._eps_quad_suboptimal = np.sqrt(np.finfo(float).eps)
        self._eps_rank = 10.0 * np.finfo(float).eps
        self._equil_cycle_max = 3
        self._equil_linear_min = 5
        self._eps_minimal_energy = np.sqrt(np.finfo(float).eps)
        self._use_numpy_lstsq = True
        # This line causes phase tests to fail if run after instantiating an Equilibrate instance:
        # phs.Phase.MINVAL = np.finfo(float).eps

    @property
    def element_list(self):
        """
        A list of strings that identify the element basis for the system

        For example,['H','C','O','Na','Mg','Al','Si','P','K','Ca','Ti','Cr',
        'Mn','Fe','Co','Ni']

        readonly

        Returns
        -------
        A Python list
        """
        return self._element_list

    @property
    def phase_list(self):
        """
        A list of class instances for phases that are permitted to form

        These instances must derive from the *PurePhase* or *SolutionPhase* classes,
        which are defined in the *phases* module.

        readonly

        Returns
        -------
        A Python list
        """
        return self._phase_list

    @property
    def lagrange_list(self):
        """
        A list of tuples characterizing the Lagrange transformation of the
        Gibbs free energy to form the operative thermodynamic potential

        Each tuple has two terms:

        1) Either of the following:

            - A string with content 'T' or 'P'

            - A dictionary keyed on element symbols with values corresponding to
              stoichiometric coefficients of element chemical potentials.  These
              linear combinations of potentials are fixed by the constraint.

        2) A function to compute values for the constraints specified in (1)

        The function has a signature func(T, P, state), where *T* is temperature
        in K, *P* is pressure in bars, and *state* is an instance of the EquilState
        class.

        readonly

        Returns
        -------
        A Python list of tuples
        """
        return self._lagrange_list

    @property
    def bulk_comp(self):
        """
        Bulk composition of the system as moles of elements

        readwrite

        Returns
        -------
        numpy array
        """
        return self._bulk_comp

    @bulk_comp.setter
    def bulk_comp(self, value):
        assert type(value) is np.ndarray, \
            'bulk_comp must be set to an numpy array'
        assert value.size == len(self.element_list), \
            'bulk_comp array must have same length as element_list'
        self._bulk_comp = value

    @property
    def moles_in(self):
        """
        Moles of phase added to the system on detection of saturation

        readwrite

        Returns
        -------
        Number of moles added (number)
        """
        return self._moles_in

    @moles_in.setter
    def moles_in(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: moles_in must a a number.'
        self._moles_in = value

    @property
    def moles_out(self):
        """
        Minimum total moles of phase allowed in system

        If there is less than this quantity, the phase is discarded.

        readwrite

        Returns
        -------
        Minimum moles allowed (number)
        """
        return self._moles_out

    @moles_out.setter
    def moles_out(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: moles_in must a a number.'
        self._moles_out = value

    @property
    def max_quad_iters(self):
        """
        Maximum number of Newton quadratic minimization steps allowed in
        computation

        readwrite

        Returns
        -------
        Number of steps allowed (int)
        """
        return self._max_quad_iters

    @max_quad_iters.setter
    def max_quad_iters(self, value):
        assert isinstance(value, int), \
            'ERROR: max_quad_iters must be an integer.'
        self._max_quad_iters = value

    @property
    def max_linear_iters(self):
        """
        Maximum number of linear search steps associated with each quadratic
        iteration

        readwrite

        Returns
        -------
        Number of linear search steps (int)
        """
        return self._max_linear_iters

    @max_linear_iters.setter
    def max_linear_iters(self, value):
        assert isinstance(value, int), \
            'ERROR: max_linear_iters must be an integer.'
        self._max_linear_iters = value

    @property
    def phase_separ_threshold(self):
        """
        Minimum energy in Joules for phase separation

        Gibbs free energy threshold that must be attained in order to add
        another instance of a phase to the equilibrium assemblage. In the phase
        unmixing (immiscibility) algorithm, candidate compositions are tested
        and if one is found with a Gibbs free energy below the projected
        tangent plane to the input composition by this amount, then the new
        phase instance is added to the system. The value is in Joules and
        should always be negative. Values larger than the default will likely
        encourage the detection of false unmixing events; values more
        negative will likely prevent the detection of unmixing and result in
        metastable assemblages. Use caution when altering the default value.

        Default is -0.1

        readwrite

        Returns
        -------
        Threshold value (number)
        """
        return self._phase_separ_threshold

    @phase_separ_threshold.setter
    def phase_separ_threshold(self, value):
        self._phase_separ_threshold = value

    @property
    def entropy(self):
        """
        Indicates if entropy is an independent variable of the minimal potential
        and if temperature is a dependent variable

        True if yes; False if reverse.

        readonly

        Returns
        -------
        Flag (bool)
        """
        return self._entropy

    @property
    def volume(self):
        """
        Indicates if volume is an independent variable of the minimal potential
        and if pressure is a dependent variable

        True if yes; False if reverse.

        readonly

        Returns
        -------
        Flag (bool)
        """
        return self._volume

    @property
    def CTf(self):
        """
        Stoichiometric vectors of element concentrations that embody the
        Lagrange constraints

        Default is None

        readonly

        Returns
        -------
        np_array or None
        """
        return self._CTf

    @property
    def VT_null(self):
        """
        Orthogonal projection operator obtained from constraints on chemical
        potentials derived from the lagrange_list entries

        Default is None

        readonly

        Returns
        -------
        np array or None
        """
        return self._VT_null

    @property
    def reactions(self):
        """
        Balanced reactions pertaining to the Lagrange constraints

        Default is None

        readonly

        Returns
        -------
        np.array or None
        """
        return self._reactions

    @property
    def lagrange_moles(self):
        """
        Moles of chemical potential entities stipulated in dictionaries,
        which are contained in the lagrange_list constraints

        Default is None

        readonly

        Returns
        -------
        np.array or None
        """
        return self._lagrange_moles

    @property
    def lagrange_use_omni(self):
        """
        If set to True, the algorithm uses the omnicomponent phase exclusively
        to balance non-linear constraint reactions.

        Default is True

        readwrite

        Returns
        -------
        Flag (bool)
        """
        return self._lagrange_use_omni

    @lagrange_use_omni.setter
    def lagrange_use_omni(self, value):
        self._lagrange_use_omni = value

    @property
    def rotate_orthog_proj(self):
        """
        Prevents creation of a search direction for potential
        minimization that has unwanted coupling of oxygen-bearing components

        The basis vectors of the orthogonal projection matrix, VT_null,
        generated from the lagrange_list attribute, are rotated to zero to
        minimize the contribution of oxygen to the null space. This flag is
        enabled to avoid creating a search direction for potential
        minimization that has unwanted coupling of oxygen-bearing components
        (for more information, see method _compute_null_space(...)). This
        option is seldom needed and can be applied only if no omnicomponent
        phase is present in the system.

        Default is False

        readwrite

        Returns
        -------
        Flag (bool)
        """
        return self._rotate_orthog_proj

    @rotate_orthog_proj.setter
    def rotate_orthog_proj(self, value):
        self._rotate_orthog_proj = value

    @property
    def A_omni_inv(self):
        """
        A matrix that transforms an array/matrix of mole numbers of elements to
        an array/matrix of mole numbers of components of the omnicomponent
        phase

        This matrix is the inverse of a matrix that maps elemental
        abundances to component mole numbers for the omnicomponent phase.
        This property is available only if chemical potential constraints
        are specified in lagrange_list and if there is an omnicomponent phase
        in the system.

        Default is None

        readonly

        Returns
        -------
        2-d nd_array
        """
        return self._A_omni_inv

    @property
    def lagrange_no_mol_deriv(self):
        """
        Produces a simplified gradient and hessian of the generalized
        Khorzhinskii potential

        Assume that the generalized Khorzhinkii potential is constructed so
        that the imposed chemical potential is not a function of mole numbers
        of any phase in the system. Alternatively, the imposed potential is
        stoichiometrically assembled using reaction coefficients involving
        equilibrium phases in the assemblage. This option produces a simplified
        gradient and hessian of the generalized Khorzhinskii potential, but
        does not affect construction of the equality constraint matrix nor the
        Lagrange multiplier terms in the hessian of the Lagrangian function.
        It is seldom necessary to set this flag to True.

        Default is False

        readwrite

        Returns
        -------
        Flag (bool)
        """
        return self._lagrange_no_mol_deriv

    @lagrange_no_mol_deriv.setter
    def lagrange_no_mol_deriv(self, value):
        self._lagrange_no_mol_deriv = value

    @property
    def eps_linear(self):
        """
        Convergence criteria for the norm of the linear projection phase of the
        equilibrium calculation

        readwrite

        Returns
        -------
        Convergence tolerance (number)
        """
        return self._eps_linear

    @eps_linear.setter
    def eps_linear(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: eps_linear must be a number.'
        assert value > np.finfo(float).eps, \
            'The value of eps_linear should be greater than machine ' \
            'precision, ' + str(np.finfo(float).eps)
        self._eps_linear = value

    @property
    def eps_quad_optimal(self):
        """
        Convergence criteria for the norm of the quadratic projection phase of
        the equilibrium calculation

        readwrite

        Returns
        -------
        Convergence tolerance (number)
        """
        return self._eps_quad_optimal

    @eps_quad_optimal.setter
    def eps_quad_optimal(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: eps_quad_optimal must be a number.'
        assert value > np.finfo(float).eps, \
            'The value of eps_quad_optimal should be greater than machine ' \
            'precision, ' + str(np.finfo(float).eps)
        self._eps_quad_optimal = value

    @property
    def eps_quad_suboptimal(self):
        """
        Relaxed convergence criteria for the norm of the quadratic projection
        phase of the equilibrium calculation

        readwrite

        Returns
        -------
        Convergence tolerance (number)
        """
        return self._eps_quad_suboptimal

    @eps_quad_suboptimal.setter
    def eps_quad_suboptimal(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: eps_quad_suboptimal must be a number.'
        assert value > np.finfo(float).eps, \
            'The value of eps_quad_suboptimal should be greater than machine ' \
            'precision, ' + str(np.finfo(float).eps)
        self._eps_quad_suboptimal = value

    @property
    def eps_rank(self):
        """
        Tolerance for establishing the rank of the projected Hessian, which is
        needed to compute a valid quadratic search direction

        readwrite

        Returns
        -------
        Tolerance (number)
        """
        return self._eps_rank

    @eps_rank.setter
    def eps_rank(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: eps_rank must be a number.'
        assert value > np.finfo(float).eps, \
            'The value of eps_rank should be greater than machine ' \
            'precision, ' + str(np.finfo(float).eps)
        self._eps_rank = value

    @property
    def equil_cycle_max(self):
        """
        Number of addition/removal cycles allowed for a phase before it is
        suppressed from the equilibrium assemblage

        readwrite

        Returns
        -------
        Iteration limit (int)
        """
        return self._equil_cycle_max

    @equil_cycle_max.setter
    def equil_cycle_max(self, value):
        assert isinstance(value, int), \
            'ERROR: equil_cycle_max must be an integer.'
        assert value > 0, \
            'The value of equil_cycle_max should be greater than zero'
        self._equil_cycle_max = value

    @property
    def equil_linear_min(self):
        """
        Number of successful linear minimization attempts that do not result in
        a subsequent dimunition of the quadratic norm after which convergence is
        accepted as a "minimal energy" condition.

        readwrite

        Returns
        -------
        Iteration limit (int)
        """
        return self._equil_linear_min

    @equil_linear_min.setter
    def equil_linear_min(self, value):
        assert isinstance(value, int), \
            'ERROR: equil_linear_min must be an integer.'
        assert value > 0, \
            'The value of equil_linear_min should be greater than zero'
        self._equil_linear_min = value

    @property
    def eps_minimal_energy(self):
        """
        Tolerance for establishing criteria for a minimum in the system
        potential.

        Successive linear minimization attempts (equil_linear_min) that do not
        result in a subsequent dimunition of the quadratic norm within this
        tolerance result in convergence as a "minimal energy" condition.

        readwrite

        Returns
        -------
        Tolerance (number)
        """
        return self._eps_minimal_energy

    @eps_minimal_energy.setter
    def eps_minimal_energy(self, value):
        assert isinstance(value, (int, float)), \
            'ERROR: eps_rank must be a number.'
        assert value > np.finfo(float).eps, \
            'The value of eps_rank should be greater than machine ' \
            'precision, ' + str(np.finfo(float).eps)
        self._eps_minimal_energy = value

    @property
    def use_numpy_lstsq(self):
        """
        Flag to toggle method used to solve for quadratic search direction

        The quadratic search direction, dn, is computed by solving a system of
        equations given by H dn = -g, where H is the projected Hessian and g the
        projected gradient of the system potential (e.g., the system Gibbs free
        energy). If *use_numpy_lstsq* is True (default), the solution method is
        numpy's linalg.lstsq method. If False, a *QR* decomposition method from
        the Python package statsmodels is utilized. The numpy method uses SVD
        decomposition and is slower, but more tolerant of rank deficient
        solutions (e.g., near phase rule violations).

        Returns
        -------
        Flag (bool)
        """
        return self._use_numpy_lstsq

    @use_numpy_lstsq.setter
    def use_numpy_lstsq(self, value):
        assert isinstance(value, bool), 'Value must be True or False.'
        self._use_numpy_lstsq = value

    ################################
    # Class private helper methods #
    ################################

    def _print_matrix_repr(self, M):
        if M.shape[0] == 0 or M.shape[1] == 0:
            return
        for i in range(0, M.shape[0]):
            print('   ', end=' ')
            for j in range(0, M.shape[1]):
                print('-', end=' ') if M[i][j] == 0 else print('x', end=' ')
            print(' ')

    def _calc_sat_state_affinities(self, state, t, p, mu_elm, omni_phase=None,
                                   debug=0):
        """
        Determine saturation state affinities for all system phases with zero
        moles.
        """
        A_min = sys.float_info.max
        A_min_name = None
        for entry in state.phase_d.values():
            name = entry['name']
            if (name != omni_phase) and (state.tot_moles_phase(name) == 0.0):
                if debug > 0:
                    print('******************************** ')
                    print('Calculating saturation state for ' + name)
                c_mat = state.c_matrix_for_phase(name)
                nc = entry['nc']
                mu_end = np.zeros(nc)
                for i in range(0, nc):
                    c_row = c_mat[i, :]
                    c_mask = np.not_equal(c_row, 0)
                    valid = np.count_nonzero(c_row) == np.count_nonzero(
                        np.extract(c_mask, self.bulk_comp))
                    mu_end[i] = np.matmul(c_row, mu_elm) if valid else 0.0
                if debug > 1:
                    print('Target chemical potentials: ', mu_end)
                # Now call a phase method that converts mu_elm to A and X
                (A, X) = entry['obj'].affinity_and_comp(t, p, mu_end,
                                                        debug=(True if debug > 2 else False), method='special')
                if debug > 0:
                    print('Affinity, mole fraction', A, X)
                    print(' ')
                entry['affinity'] = A
                entry['mole_frac'] = X
                if A < A_min:
                    A_min = A
                    A_min_name = name
        return (A_min, A_min_name)

    def _determine_phase_stability(self, state, t, p, phase_name, debug=0):
        """
        Calculate if the specified phase is stable with respect to unmixing.

        Returns:
        --------
        result: tuple
            bool, ndarray - unmixing detected; None or composition of second phase
        """
        #if phase_name == state.omni_phase():
        #    return False, None
        phase = state.phase_d[phase_name]
        if not phase['allow_unmixing']:
            return False, None
        moles = phase['moles']
        if moles.size == 1:
            return False, None
        tot_moles = np.sum(moles)
        if tot_moles <= 0:
            return False, None
        if debug > 0:
            print('')
            print('Unmixing calculation for', phase_name, moles / tot_moles)

        nc = phase['nc']
        mu0 = np.full(nc, 0.0)
        moles_nz = np.array([i for i in range(0,nc) if moles[i] != 0])  
        x = np.full(nc, 0.0)
        np.put(x, moles_nz, 0.000001)
        for i in moles_nz:
            x[i] = 1.0 - np.sum(x)
            mu0[i] = phase['obj'].gibbs_energy(t, p, mol=x)
            x[i] = 0.000001
        gHat = phase['obj'].gibbs_energy(t, p, mol=moles)
        dgdn = phase['obj'].gibbs_energy(t, p, mol=moles, deriv={'dmol': 1})[0]
        for i in moles_nz:
            gHat -= mu0[i] * moles[i]
            dgdn[i] -= mu0[i]
        gHat /= tot_moles

        def deltaG(x, *args):
            moles_nz, t, p, phase, moles, gHat, dgdn, mu0 = args
            f = phase['obj'].gibbs_energy(t, p, mol=x)
            df = phase['obj'].gibbs_energy(t, p, mol=x, deriv={'dmol': 1})[0]
            for i in moles_nz:
                f -= mu0[i] * x[i]
                df[i] -= mu0[i]
            f /= np.sum(x)
            f -= gHat + np.matmul(x / np.sum(x) - moles / np.sum(moles), dgdn)
            df -= dgdn
            return f, df

        bound_set = [(0.0, 0.0) for i in range(0,nc)]
        for i in moles_nz:
            bound_set[i] = (0.01, 0.99)

        lowest_fun = 0.0
        lowest_comp = None
        debug_flag = True if debug > 1 else False
        for i in moles_nz:
            x = np.full(nc, 0.0)
            np.put(x, moles_nz, 0.05)
            x[i] = 1.0 - np.sum(x)
            result = sp.optimize.minimize(deltaG, x,
                                          args=(moles_nz, t, p, phase, moles, gHat, dgdn, mu0),
                                          method='L-BFGS-B', jac=True,
                                          bounds=bound_set,
                                          options={'maxiter': 250, 'disp': debug_flag})
            if debug > 0:
                print(result)
                print('')
            if result.fun < lowest_fun:
                lowest_fun = result.fun
                lowest_comp = result.x
        if lowest_fun < self.phase_separ_threshold:
            return True, lowest_comp
        else:
            return False, None

    def _add_phase_to_system(self, state, name):
        """
        Add a small amount of phase 'name' to the system and maintain mass
        balance by substracting amount added from the omnicomponent phase.
        """
        omni_phase_name = state.omni_phase()
        omni_phase = state.phase_d[omni_phase_name]
        new_phase = state.phase_d[name]
        for i in range(0, new_phase['nc']):
            new_phase['moles'][i] = new_phase['mole_frac'][i] * self.moles_in
        np.put(new_phase['mole_nz'],
               np.flatnonzero(np.abs(new_phase['moles']) < 10.0 * np.finfo(float).eps),
               0.0)
        c_new_phase = state.c_matrix_for_phase(name)
        # determine elements associated with new_phase
        elm = np.matmul(c_new_phase.T, new_phase['moles'])
        # cast these elements into moles of the omnicomponent phase
        cT = state.c_matrix_for_phase(omni_phase_name).T
        cTinv = np.linalg.inv(cT) if cT.shape[0] == cT.shape[1] else np.linalg.pinv(cT)
        comp = np.matmul(cTinv, elm)
        # now subtract this quanity from the omnicomponent phase
        omni_phase['moles'] = np.subtract(omni_phase['moles'], comp)
        # For almost missing components, a small negative component fraction
        # may be obtained, overwrite this with a small positive mole fraction:
        for i in range(len(omni_phase["moles"])):
            if omni_phase["moles"][i] < 0.0 and np.abs(omni_phase["moles"][i])<1e-5:
                omni_phase["moles"][i] = 0.0
        return omni_phase['obj'].test_endmember_comp(omni_phase['moles'])

    def _remove_phase_from_system(self, state, name):
        """
        Remove a phase 'name' from the system and maintain mass balance by
        adding amount removed to the omnicomponent phase.
        """
        omni_phase_name = state.omni_phase()
        omni_phase = state.phase_d[omni_phase_name]
        old_phase = state.phase_d[name]
        c_old_phase = state.c_matrix_for_phase(name)
        # determine elements associated with old_phase
        elm = np.matmul(c_old_phase.T, old_phase['moles'])
        # cast these elements into moles of the omnicomponent phase
        cT = state.c_matrix_for_phase(omni_phase_name).T
        cTinv = np.linalg.inv(cT) if cT.shape[0] == cT.shape[1] else np.linalg.pinv(cT)
        comp = np.matmul(cTinv, elm)
        # now add this quanity to the omnicomponent phase
        omni_phase['moles'] = np.add(omni_phase['moles'], comp)
        # finally, zero mole numbers of the phase removed from the system
        for i in range(0, old_phase['nc']):
            old_phase['moles'][i] = 0.0
        return omni_phase['obj'].test_endmember_comp(omni_phase['moles'])

    def _apply_non_linear_constraints(self, tp_l, state, debug):
        """
        Enforce the non-linear constraints (entropy, volume, chemical
        potential(s)) on the system.
        """

        def linear_search(step, state, t, p, n_ref, row, f):
            index = 0
            for name, entry in state.phase_d.items():
                if state.tot_moles_phase(name) > 0:
                    for i in range(0, entry['nc']):
                        corr = step * self.reactions[row][index]
                        entry['moles'][i] = n_ref[index] + corr
                        index += 1
            dgdn = state.dGdn(t, p, element_basis=False)
            result = np.matmul(self.reactions[row], dgdn)[0]
            result -= f(t, p, state)
            return np.sqrt(result * result)

        def test_bounds_for_search(step, state, t, p, row):
            index = 0
            for name, entry in state.phase_d.items():
                if state.tot_moles_phase(name) > 0:
                    nc = entry['nc']
                    moles = np.empty(nc)
                    for i in range(0, nc):
                        corr = step * self.reactions[row][index]
                        moles[i] = entry['moles'][i] + corr
                        index += 1
                    if nc == 1:
                        if moles[0] < 0.0:
                            return False
                    else:
                        if not entry['obj'].test_endmember_comp(moles):
                            return False
            return True

        def root_target(x, state, t, p, c, f):
            if c == 'T':
                result = state.dGdT(x, p) - f(t, p, state)
                Dresult = state.d2GdT2(x, p)
                return result, Dresult
            elif c == 'P':
                result = state.dGdP(t, x) - f(t, p, state)
                Dresult = state.d2GdP2(t, x)
                return result, Dresult
            else:
                return None, None

        t = tp_l[0]
        p = tp_l[1]
        row = 0
        func_for_row = []
        for c, f in self.lagrange_list:
            if type(c) is dict:
                if self.reactions.shape[0] > 1:
                    func_for_row.append(f)
                    continue

                step_max = 0.5
                while not test_bounds_for_search(step_max, state, t, p, row):
                    step_max *= 0.9
                    if step_max < sys.float_info.epsilon:
                        break
                if debug > 1:
                    print('Maximum linear step search length; ', step_max)

                step_min = -0.5
                while not test_bounds_for_search(step_min, state, t, p, row):
                    step_min *= 0.9
                    if step_min > -sys.float_info.epsilon:
                        break
                if debug > 1:
                    print('Minimum linear step search length; ', step_min)

                n_ref = []
                for name, entry in state.phase_d.items():
                    if state.tot_moles_phase(name) > 0:
                        for i in range(0, entry['nc']):
                            n_ref.append(entry['moles'][i])
                n_ref = np.array(n_ref)

                result = sp.optimize.minimize_scalar(linear_search,
                                                     bracket=(0, 1), bounds=(step_min, step_max),
                                                     args=(state, t, p, n_ref, row, f), method='Bounded',
                                                     options={'maxiter': 250, 'disp': debug, 'xatol': 1e-8})
                if debug > 1:
                    print('Optimal linear search length:', result)
                if np.absolute(result.fun) > 1.0:
                    print('Imposed chemical potential constraints', end=' ')
                    print('cannot be reconciled within 1 Joule.')
                    print('Consider setting the "lagrange_use_omni"', end=' ')
                    print('property to True.')
                self._lagrange_moles[row] += result.x
                row += 1
            elif type(c) is str:
                if self.entropy and self.volume:
                    if c == 'T':
                        tFun = f
                    elif c == 'P':
                        pFun = f
                    continue
                result = sp.optimize.root_scalar(root_target,
                                                 args=(state, t, p, c, f),
                                                 method='newton',
                                                 x0=t if c == 'T' else p,
                                                 fprime=True,
                                                 maxiter=50)
                if debug > 0:
                    print(result)
                if c == 'T':
                    tp_l[0] = result.root
                elif c == 'P':
                    tp_l[1] = result.root
        if self.entropy and self.volume:
            # tFun and pFun from above
            # TBD
            def roots_fun(x, state, t, p, tFun, pFun):
                resT = state.dGdT(x[0], x[1]) - tFun(t, p, state)
                DresTT = state.d2GdT2(x[0], x[1])
                resP = state.dGdP(x[0], x[1]) - pFun(t, p, state)
                DresPP = state.d2GdP2(x[0], x[1])
                DresTP = state.d2GdTdP(x[0], x[1])
                return [resT, resP], np.array([[DresTT, DresTP], [DresTP, DresPP]])

            result = sp.optimize.root(roots_fun,
                                      [t, p],
                                      args=(state, t, p, tFun, pFun),
                                      method='hybr',
                                      jac=True)
            if debug > 0:
                print(result)
            tp_l[0] = result.x[0]
            tp_l[1] = result.x[1]

        if len(func_for_row) > 0:
            if debug > 0:
                print('Solving with multiple chemical potential constraints.')

            def multi_search(step, state, t, p, n_ref, func_for_row):
                index = 0
                for name, entry in state.phase_d.items():
                    if state.tot_moles_phase(name) > 0:
                        for i in range(0, entry['nc']):
                            corr = 0.0
                            for j in range(0, len(func_for_row)):
                                corr += step[j] * self.reactions[j][index]
                            entry['moles'][i] = n_ref[index] + corr
                            index += 1
                dgdn = state.dGdn(t, p, element_basis=False)

                result = []
                for row, f in enumerate(func_for_row):
                    result.append(np.matmul(self.reactions[row], dgdn)[0]
                                  - f(t, p, state))
                return result

            def test_bounds_for_multi_search(step, state, t, p, func_for_row):
                index = 0
                for name, entry in state.phase_d.items():
                    if state.tot_moles_phase(name) > 0:
                        nc = entry['nc']
                        moles = np.empty(nc)
                        for i in range(0, nc):
                            corr = 0.0
                            for j in range(0, len(func_for_row)):
                                corr += step[j] * self.reactions[j][index]
                            moles[i] = entry['moles'][i] + corr
                            index += 1
                        if nc == 1:
                            if moles[0] < 0.0:
                                return False
                        else:
                            if not entry['obj'].test_endmember_comp(moles):
                                return False
                return True

            n_ref = []
            for name, entry in state.phase_d.items():
                if state.tot_moles_phase(name) > 0:
                    for i in range(0, entry['nc']):
                        n_ref.append(entry['moles'][i])
            n_ref = np.array(n_ref)

            if debug > 0:
                comp = state.oxide_comp('Liquid')
                for key, value in comp.items():
                    print("{0:<6s} {1:6.2f}".format(key, value), end=' ')
                print('')
            result = sp.optimize.root(multi_search,
                                      np.zeros(len(func_for_row)),
                                      args=(state, t, p, n_ref, func_for_row),
                                      method='hybr',
                                      jac=False)
            if debug > 0:
                comp = state.oxide_comp('Liquid')
                for key, value in comp.items():
                    print("{0:<6s} {1:6.2f}".format(key, value), end=' ')
                print('')
                print(result)
                state.print_state()
            for i in range(0, len(func_for_row)):
                self._lagrange_moles[i] += result.x[i]

    def _compute_a_and_qr(self, t, p, state, P_nz, debug):
        """
        From the system state, calculate the mass balance constraint matrix and
        its QR decomposition.
        """
        filtr = lambda x: x if abs(x) > float(1000 * sys.float_info.epsilon) else 0
        vfiltr = np.vectorize(filtr, otypes=[float])

        first = True
        for name, entry in state.phase_d.items():
            if state.tot_moles_phase(name) > 0.0:
                if first:
                    c_mat = state.c_matrix_for_phase(name)
                    first = False
                else:
                    c_mat = np.vstack((c_mat, state.c_matrix_for_phase(name)))
        A = c_mat.T

        if self.lagrange_list is not None:
            if debug > 1:
                print('A before projection and augmentation', A.shape)
                if debug > 2:
                    print(A)
                else:
                    self._print_matrix_repr(A)
            A = np.matmul(self.A_omni_inv, A)
            if debug > 1:
                print('A after projection and before augmentation', A.shape)
                if debug > 2:
                    print(A)
                else:
                    self._print_matrix_repr(A)
            if self.CTf.shape[0] > 0:
                ne = len(self.element_list)
                omni_phase_name = state.omni_phase()
                if self.lagrange_use_omni and (omni_phase_name is not None):
                    first = True
                    for phase in state.phase_d.values():
                        if state.tot_moles_phase(phase['name']) > 0:
                            if phase['name'] == omni_phase_name:
                                entry = np.zeros((ne, phase['nc']))
                            else:
                                entry = np.ones((ne, phase['nc']))
                            A_mask = entry if first else np.hstack((A_mask, entry))
                            first = False
                    A_react = np.ma.array(A, mask=A_mask, fill_value=0)
                    A_react = A_react.filled()
                else:
                    A_react = np.copy(A)
                if debug > 1:
                    print('A reaction matrix', A_react.shape)
                    if debug > 2:
                        print(A_react)
                    else:
                        self._print_matrix_repr(A_react)
                # Create a mass balance constraint on constrained entity
                react, res, rank, s = np.linalg.lstsq(A_react,
                                                      np.matmul(self.A_omni_inv, self.CTf.T), rcond=None)
                self._reactions = vfiltr(react.T)
                if not self.entropy and not self.volume:
                    if debug > 0:
                        print(".....")
                        print('Enforcing chemical potential constraints.')
                    self._apply_non_linear_constraints([t, p], state, debug)
                    if debug > 0:
                        print(".....")
                if debug > 1:
                    print('Balanced Lagrange reactions', self._reactions.shape)
                    print(self._reactions)
                    if debug > 2:
                        print('... residuals', res.shape)
                        print(res)
                        print('... rank', rank)
                        print('... singular values', s.shape)
                        print(s)
                if debug > 1:
                    dgdn = state.dGdn(t, p, element_basis=False)
                    deltaG = np.matmul(self.reactions, dgdn)
                    row = 0
                    for c, f in self.lagrange_list:
                        if type(c) is dict:
                            deltaG[row] -= f(t, p, state)
                            row += 1
                    print('Lagrange chemical potential constraints',
                          deltaG.shape)
                    print(deltaG)
                # Project the A  matrix into the NULL space
                Aproj = np.matmul(self.VT_null, A)
                if debug > 1:
                    print('Aproj', Aproj.shape)
                    if debug > 2:
                        print(Aproj)
                    else:
                        self._print_matrix_repr(Aproj)
                # Add rows for the constraint
                ###### d2gdn2 = state.d2Gdn2(t, p, element_basis=True)
                ###### con_rows = np.matmul(self.CTf, d2gdn2)
                d2gdn2 = state.d2Gdn2(t, p, element_basis=False)
                con_rows = np.matmul(self._reactions, d2gdn2)
                if debug > 2:
                    print('Constraint rows', con_rows.shape)
                    print(con_rows)
                Aproj = np.vstack((Aproj, con_rows))
                if debug > 1:
                    print('A proj augmented with constraints', Aproj.shape)
                    if debug > 2:
                        print(Aproj)
                    else:
                        self._print_matrix_repr(Aproj)
                A = Aproj
            if self.entropy and self.volume:
                d2GdTdn = state.d2GdTdn(t, p)
                d2GdT2 = state.d2GdT2(t, p)
                d2GdPdn = state.d2GdPdn(t, p)
                d2GdP2 = state.d2GdP2(t, p)
                d2GdTdP = state.d2GdTdP(t, p)
                A = np.vstack((A, d2GdTdn.T, d2GdPdn.T))
                rows, cols = A.shape
                col = np.zeros((rows, 2))
                col[-2][0] = d2GdT2
                col[-1][1] = d2GdP2
                col[-2][1] = d2GdTdP
                col[-1][0] = d2GdTdP
                A = np.hstack((A, col))
            elif self.entropy:
                d2GdTdn = state.d2GdTdn(t, p)
                d2GdT2 = state.d2GdT2(t, p)
                A = np.vstack((A, d2GdTdn.T))
                rows, cols = A.shape
                col = np.zeros((rows, 1))
                col[-1][0] = d2GdT2
                A = np.hstack((A, col))
            elif self.volume:
                d2GdPdn = state.d2GdPdn(t, p)
                d2GdP2 = state.d2GdP2(t, p)
                A = np.vstack((A, d2GdPdn.T))
                rows, cols = A.shape
                col = np.zeros((rows, 1))
                col[-1][0] = d2GdP2
                A = np.hstack((A, col))
            if debug > 1:
                print('A augmented with constraints', A.shape)
                if debug > 2:
                    print(A)
                else:
                    self._print_matrix_repr(A)

        # delete rows that have zero concentrations of elements in the system
        if self.lagrange_list is not None:
            bc = np.matmul(self.A_omni_inv, self.bulk_comp)
            if self.CTf.shape[0] > 0:
                bc = np.matmul(self.VT_null, bc)
            A = np.delete(A, np.argwhere(bc == 0), axis=0)
        else:
            A = np.delete(A, np.argwhere(self.bulk_comp == 0), axis=0)
        # delete columns that have zero concentrations of solution components
        # print(f"A: {A}")
        # print(f"P_nz: {P_nz}")
        A = np.matmul(A, P_nz)

        row, col = A.shape
        rref = np.array(sym.Matrix(A).rref()[0]).astype(np.float64)
        row_rank = np.count_nonzero(np.linalg.norm(rref, axis=1))
        if row == row_rank:
            df = col - row
        else:
            if debug > 1:
                print('A matrix has reduced row space by:', row - row_rank)
            df = col - row_rank
        R, Q = sp.linalg.rq(A, mode='full')
        if debug > 1:
            print('RQ = A, R matrix', R.shape)
            if debug > 2:
                print(vfiltr(R))
            else:
                self._print_matrix_repr(vfiltr(R))
        if debug > 1:
            print('RQ = A, Q matrix', Q.shape)
            if debug > 2:
                print(vfiltr(Q))
            else:
                self._print_matrix_repr(vfiltr(Q))
        R11 = vfiltr(R[:, df:])
        Q1 = vfiltr(Q[df:, :])
        Q2 = vfiltr(Q[0:df, :])
        return (A, df, Q1, Q2, R11)

    def _compute_null_space(self, state, debug):
        """
        Construct a projection operator to account for open system constraints
        """
        ne = len(self.element_list)
        CTf = np.empty((0, ne))
        for c, f in self.lagrange_list:
            if type(c) is str:
                if c == 'T':
                    self._entropy = True
                elif c == 'P':
                    self._volume = True
                else:
                    assert True, 'Legendre transform must specify "T", "P", ' \
                        + 'or a dictionary of chemical potential constraints.'
            elif type(c) is dict:
                row = np.zeros(ne)
                for key, value in c.items():
                    index = self.element_list.index(key)
                    row[index] = value
                CTf = np.vstack((CTf, row))
            else:
                assert True, 'Legendre transform must specify "T", "P", ' \
                    + 'or a dictionary of chemical potential constraints.'
        self._CTf = CTf
        self._lagrange_moles = np.zeros(len(self.lagrange_list))
        if state.omni_phase() is not None:
            self._A_omni_inv = np.linalg.inv(
                state.c_matrix_for_phase(state.omni_phase()).T)
        else:
            self._A_omni_inv = np.eye(ne)
        if debug > 1:
            print('CTf', CTf.shape)
            print(CTf)
            print('A_omni_inv', self.A_omni_inv.shape)
            if debug > 2:
                print(self.A_omni_inv)
            else:
                self._print_matrix_repr(self.A_omni_inv)
        CTf_pad = np.pad(CTf, ((0, CTf.shape[1] - CTf.shape[0]), (0, 0)), mode='constant')
        U, S, VT = np.linalg.svd(np.matmul(CTf_pad, self.A_omni_inv.T))
        if debug > 1:
            print('U', U.shape)
            if debug > 2:
                print(U)
            else:
                self._print_matrix_repr(U)
            print('S', S.shape)
            if debug > 2:
                print(S)
            print('VT', VT.shape)
            if debug > 2:
                print(VT)
            else:
                self._print_matrix_repr(VT)
        rank = np.count_nonzero(S)
        self._VT_null = VT[rank:, :]
        if debug > 1:
            print('VT null space', self._VT_null.shape)
            if debug > 2:
                print(self._VT_null)
            else:
                self._print_matrix_repr(self._VT_null)
        if self.rotate_orthog_proj and state.omni_phase() is None:
            try:
                row_ind = self.element_list.index("O")
            except ValueError:
                if debug > 1:
                    print("ERROR! Oxygen is not present in the system and ")
                    print("       a null-space rotation on O is specified.")
                return
            if debug > 1:
                print("----------")
                print("Adjusting the null space basis for oxygen.")
            ns = self._VT_null.T
            filtr = lambda x: x if abs(x) > float(
                1000 * sys.float_info.epsilon) else 0
            vfiltr = np.vectorize(filtr, otypes=[float])

            # zero the null space bassis coefficient at col_ind and row_ind
            def zeros(deg_a, v_a, u, ns, col_ind_a, row_ind, ret_matrix):
                ns_p = ns
                for (deg, v, col_ind) in zip(deg_a, v_a, col_ind_a):
                    theta = (deg / 180.0) * np.pi
                    P = np.eye(v.shape[0]) + np.sin(theta) * (np.outer(u, v)
                                                              - np.outer(v, u)) + (np.cos(theta) - 1) * (np.outer(u, u)
                                                                                                         - np.outer(v,
                                                                                                                    v))
                    ns_p = np.matmul(P, ns_p)
                if ret_matrix:
                    return ns_p
                else:
                    result = []
                    for col_ind in col_ind_a:
                        result.append(ns_p[row_ind][col_ind])
                    return np.array(result)

            u = np.zeros(ns.shape[0])
            u[row_ind] = 1
            v_a = []
            col_ind_a = []
            guess = []
            for col in range(0, ns.shape[1]):
                if ns[row_ind][col] != 0.0:
                    v_a.append(ns[:, col])
                    col_ind_a.append(col)
                    guess.append(0)
            result = sp.optimize.root(zeros,
                                      args=(v_a, u, ns, col_ind_a, row_ind, False),
                                      x0=np.array(guess))
            if debug > 1:
                print("... rotation algoritm output:")
                print(result)
            ns_p = vfiltr(zeros(result.x, v_a, u, ns, col_ind_a, row_ind, True))
            self._VT_null = ns_p.T
            if debug > 1:
                print('rotated null space basis', self._VT_null.shape)
                if debug > 2:
                    print(self._VT_null)
                else:
                    self._print_matrix_repr(self._VT_null)
                print("----------")

    def _augment_gradient_hessian(self, t, p, g, H, state, debug):
        """
        Ammend the gradient and Hessian of the Gibbs free energy to account for
        the Khorzhinskii corections.

        Returns
        -------
        result: tuple
            corrected (g, H)
        """
        dgdn = np.copy(g)
        d2gdn2 = np.copy(H)
        rank_H = H.shape[0]
        idx = 0
        for c, f in self.lagrange_list:
            if type(c) is dict:
                # determine molar equivalent of open system reactant
                moles = state.moles_v(self.reactions[idx])
                # determine the first order correction to the gradient
                gAdd_1 = np.matmul(self.reactions[idx], dgdn)[0]
                gAdd_1 *= self.reactions[idx]
                gAdd_1 = np.reshape(gAdd_1, (gAdd_1.shape[0], 1))
                if debug > 2:
                    print('moles', moles)
                    print('1st gradient addition', gAdd_1.shape)
                    print(-gAdd_1)
                # determine the second order correction to the gradient
                if not self.lagrange_no_mol_deriv:
                    gAdd_2 = np.matmul(self.reactions[idx], d2gdn2)
                    gAdd_2 = np.reshape(gAdd_2, (gAdd_2.shape[0], 1))
                    if debug > 2:
                        print('2nd gradient addition', gAdd_2.shape)
                        print(-moles * gAdd_2)
                else:
                    gAdd_2 = np.zeros(g.shape)
                # add contributions to gradient
                g = g - gAdd_1 - moles * gAdd_2
                # add contributions to Hessian
                # ... first, index the structure of the hessian
                phase_info = []
                offset = 0
                for entry in state.phase_d.values():
                    if state.tot_moles_phase(entry['name']) > 0.0:
                        for idx2 in range(0, entry['nc']):
                            phase_info.append((entry['name'], offset, idx2))
                        offset += entry['nc']
                if debug > 1:
                    print('phase_info', len(phase_info), 'reactions',
                          self.reactions.shape)
                    print(phase_info)
                # ... second, build the Hessian additions
                if not self.lagrange_no_mol_deriv:
                    hAdd_1 = np.outer(self.reactions[idx], gAdd_2)
                    hAdd_2 = np.outer(gAdd_2, self.reactions[idx])
                    hAdd_3 = np.zeros(H.shape)
                    for idx2, coeff in enumerate(self.reactions[idx]):
                        if coeff != 0:
                            name, offset, col = phase_info[idx2]
                            HAdd = coeff * state.d3Gdn3(t, p, name, cmp=col)
                            idx3 = rank_H - HAdd.shape[0] - offset
                            hAdd_3 += np.pad(HAdd, ((offset, idx3), (offset, idx3)),
                                             'constant', constant_values=0)
                    if debug > 2:
                        print('1st hessian addition', hAdd_1.shape)
                        print(-hAdd_1)
                        print('2nd hessian addition', hAdd_2.shape)
                        print(-hAdd_2)
                        print('3rd hessian addition', hAdd_3.shape)
                        print(-moles * hAdd_3)
                    H += -hAdd_1 - hAdd_2 - moles * hAdd_3
                idx += 1

        if self.entropy and self.volume:
            g = g - t * state.d2GdTdn(t, p)
            g = g - p * state.d2GdPdn(t, p)
            d2gdtdp = state.d2GdTdP(t, p)
            g = np.vstack((g, -t * state.d2GdT2(t, p) - p * d2gdtdp))
            g = np.vstack((g, -p * state.d2GdP2(t, p) - t * d2gdtdp))
            H = H - t * state.d3GdTdn2(t, p) - p * state.d3GdPdn2(t, p)
            d3gdt2dn = state.d3GdT2dn(t, p)
            d3gdp2dn = state.d3GdP2dn(t, p)
            d3gdtdpdn = state.d3GdTdPdn(t, p)
            colT = -t * d3gdt2dn - p * d3gdtdpdn
            colP = -p * d3gdp2dn - t * d3gdtdpdn
            H = np.hstack((H, colT, colP))
            d3gdt2dp = state.d3GdT2dP(t, p)
            d3gdtdp2 = state.d3GdTdP2(t, p)
            termTT = -state.d2GdT2(t, p) - t * state.d3GdT3(t, p) - p * d3gdt2dp
            termPP = -state.d2GdP2(t, p) - p * state.d3GdP3(t, p) - t * d3gdtdp2
            termTP = -state.d2GdTdP(t, p) - t * d3gdt2dp - p * d3gdtdp2
            rowT = np.hstack((colT.T, np.array([termTT], ndmin=2),
                              np.array([termTP], ndmin=2)))
            rowP = np.hstack((colP.T, np.array([termTP], ndmin=2),
                              np.array([termPP], ndmin=2)))
            H = np.vstack((H, rowT, rowP))
        elif self.entropy:
            g = g - t * state.d2GdTdn(t, p)
            g = np.vstack((g, -t * state.d2GdT2(t, p)))
            H = H - t * state.d3GdTdn2(t, p)
            col = -t * state.d3GdT2dn(t, p)
            H = np.hstack((H, col))
            term = -state.d2GdT2(t, p) - t * state.d3GdT3(t, p)
            row = np.hstack((col.T, np.array([term], ndmin=2)))
            H = np.vstack((H, row))
        elif self.volume:
            g = g - p * state.d2GdPdn(t, p)
            g = np.vstack((g, -p * state.d2GdP2(t, p)))
            H = H - p * state.d3GdPdn2(t, p)
            col = -p * state.d3GdP2dn(t, p)
            H = np.hstack((H, col))
            term = -state.d2GdP2(t, p) - p * state.d3GdP3(t, p)
            row = np.hstack((col.T, np.array([term], ndmin=2)))
            H = np.vstack((H, row))

        if debug > 1:
            print('Augmented g', g.shape)
            if debug > 2:
                print(g)
            print('Augmented H', H.shape)
            if debug > 2:
                print(H)
            else:
                self._print_matrix_repr(H)
        return (g, H)

    def _compute_lagrange_multipliers(self, g, A, debug):
        """
        Estimate Lagrange multipliers from the gradient and equality constraint
        derivative matrix.
        """
        soln, res, rank, s = np.linalg.lstsq(A.T, g, rcond=None)
        if debug > 1:
            print('Lagrange multiplier solution', rank)
            print('... solution', soln.shape)
            print(soln)
            if debug > 2:
                print('... residuals', res.shape)
                print(res)
                print('... singular values', s.shape)
                print(s)
        return soln

    def _augment_hessian_return_wronskian(self, t, p, H, l_mult, state, debug):
        """
        Augment the Hessian matrix with Lagrange multiplier terms for the
        non-linear constraints.
        """
        if self.reactions is not None:
            rows = self.reactions.shape[0]
            row = 0
            for i in range(0, rows):
                lagrange = l_mult[l_mult.size - rows + row]
                react = self.reactions[i]
                col = 0
                result = []
                if debug > 1:
                    print('Lagrange multiplier:', lagrange)
                for name, entry in state.phase_d.items():
                    if state.tot_moles_phase(name) > 0.0:
                        nc = entry['nc']
                        if entry['nc'] == 1:
                            coeff = react[col]
                            mat = coeff * np.array([[0.0]])
                            col += 1
                        else:
                            mat = np.zeros((nc, nc))
                            for j in range(0, nc):
                                if react[col] != 0.0:
                                    coeff = react[col]
                                    mat_r = state.d3Gdn3(t, p, phase=name, cmp=j)
                                    mat = np.add(mat, coeff * mat_r)
                                col += 1
                        result.append(mat)
                result = -lagrange * sp.linalg.block_diag(*result)
                if debug > 1:
                    print('Hessian addition:', result.shape)
                    if debug > 2:
                        print(result)
                H = np.add(H, result)
                row += 1
        if self.entropy and self.volume:
            lam_T = l_mult[-2]
            lam_P = l_mult[-1]
            result = -lam_T * state.d3GdTdn2(t, p) - lam_P * state.d3GdPdn2(t, p)
            d3gdtdpdn = state.d3GdTdPdn(t, p)
            col_T = -lam_T * state.d3GdT2dn(t, p) - lam_P * d3gdtdpdn
            col_P = -lam_T * d3gdtdpdn - lam_P * state.d3GdP2dn(t, p)
            result = np.hstack((result, col_T, col_P))
            d3gdt2dp = state.d3GdT2dP(t, p)
            d3gdtdp2 = state.d3GdTdP2(t, p)
            term_TT = -lam_T * state.d3GdT3(t, p) - lam_P * d3gdt2dp
            term_TP = -lam_T * d3gdt2dp - lam_P * d3gdtdp2
            term_PP = -lam_T * d3gdtdp2 - lam_P * state.d3GdP3(t, p)
            row_T = np.hstack((col_T.T, np.array([term_TT], ndmin=2),
                               np.array([term_TP], ndmin=2)))
            row_P = np.hstack((col_P.T, np.array([term_TP], ndmin=2),
                               np.array([term_PP], ndmin=2)))
            result = np.vstack((result, row_T, row_P))
            if debug > 1:
                print('Hessian addition (S,V):', result.shape)
                if debug > 2:
                    print(result)
            H = np.add(H, result)
        elif self.entropy:
            lam_T = l_mult[-1]
            result = -lam_T * state.d3GdTdn2(t, p)
            col_T = -lam_T * state.d3GdT2dn(t, p)
            result = np.hstack((result, col_T))
            term_TT = -lam_T * state.d3GdT3(t, p)
            row_T = np.hstack((col_T.T, np.array([term_TT], ndmin=2)))
            result = np.vstack((result, row_T))
            if debug > 1:
                print('Hessian addition (S):', result.shape)
                if debug > 2:
                    print(result)
            H = np.add(H, result)
        elif self.volume:
            lam_P = l_mult[-1]
            result = -lam_P * state.d3GdPdn2(t, p)
            col_P = -lam_P * state.d3GdP2dn(t, p)
            result = np.hstack((result, col_P))
            term_PP = -lam_P * state.d3GdP3(t, p)
            row_P = np.hstack((col_P.T, np.array([term_PP], ndmin=2)))
            result = np.vstack((result, row_P))
            if debug > 1:
                print('Hessian addition (V):', result.shape)
                if debug > 2:
                    print(result)
            H = np.add(H, result)
        return H

    ########################
    # Class public methods #
    ########################

    def mu0O2(self, t, p):
        """
        Calculates the chemical potential of oxygen gas in the standard state
        of unit fugacity at one bar and any tenperature

        Parameters
        ----------
        t : float
            Temperature in Kelvins
        p : float
            Pressure in bars

        Returns
        -------
        result : float
            standard state chemical potential in Joules/mole of O2 gas

        Notes
        -----
        Algorithm from Berman (1988).
        """
        tr = 298.15
        hs = 23.10248 * (t - tr) + 2.0 * 804.8876 * (np.sqrt(t) - np.sqrt(tr)) \
             - 1762835.0 * (1.0 / t - 1.0 / tr) - 18172.91960 * np.log(t / tr) \
             + 0.5 * 0.002676 * (t * t - tr * tr)
        ss = 205.15 + 23.10248 * np.log(t / tr) \
             - 2.0 * 804.8876 * (1.0 / np.sqrt(t) - 1.0 / np.sqrt(tr)) \
             - 0.5 * 1762835.0 * (1.0 / (t * t) - 1.0 / (tr * tr)) \
             + 18172.91960 * (1.0 / t - 1.0 / tr) + 0.002676 * (t - tr)
        return hs - t * ss

    def log10NNO(self, t, p):
        """
        Calculates the base 10 logarithm of oxygen fugacity along the nickel-
        nickel oxide buffer

        Parameters
        ----------
        t : float
            Temperature in Kelvins
        p : float
            Pressure in bars

        Returns
        -------
        result : float
            log (10) f O2

        Notes
        -----
        Algorithm from O'Neill and Pownceby (1993, Contributions to Mineralogy
        and Petrology, 114, 296-314) using the pressure correction suggested by
        Frost (1991, Mineralogical Society of America, Reviews in Mineralogy,
        v. 25, 1-9)
        """
        return -25018.7 / t + 12.981 + 0.046 * (p - 1.0) / t - 0.5117 * np.log(t)

    def muNNO(self, t, p, delta=0.0):
        """
        Calculates the excess chemical potential of oxygen along the nickel-
        nickel oxide (+ delta offset) buffer

        Parameters
        ----------
        t : float
            Temperature in Kelvins
        p : float
            Pressure in bars
        delta : float, default 0.0
            Offset in base 10 log units relative to the nickel-nickel oxide
            oxygen buffer

        Returns
        -------
        result : float
            Excess chemical potential of oxygen in Joules/mole of O2

        Notes
        -----
        See definition of function log10NNO(t, p)
        """
        return 8.3144598 * t * np.log(10.0) * (self.log10NNO(t, p) + delta)

    def kc_ferric_ferrous(self, t, p, m, mtype='components', compute='logfO2',
                          deltaNNO=0.0, debug=0):
        """
        Calculates oxygen fugacity or ferric-ferrous ratio for silicate melts
        using the Kress and Carmichael (1991) calibration

        Parameters
        ----------
        t : float
            Temperature in Kelvins
        p : float
            Pressure in bars
        m : numpy ndarray
            An array of moles of endmember liquid components correcponding to
            the model of Ghiorso and Sack (1995) (m has length 15) or the
            model of Ghiorso and Gualda (2015) (m has length 16)
        mtype : str, default 'components'
            Type of values in m.
        compute : str, default 'logfO2'
            Type of output requested, see Returns.
        deltaNNO : float, default 0.0
            If ferric-ferrous computation if requested (see compute), the ratio
            will be computed at log 10 f O2 = nickel-nickel oxide + deltaNNO
        debug : int, default 0
            Level of detail printed by the method:

            - 0, no information

            - 1, minimal progress information

            - 2, normal debugint output level

            - 3, verbose debuging output level

        Returns
        -------
        result : float
            log 10 f O2, if compute is set to 'logfO2' (default)
            excess chemical potential of O2, if compute is set to 'chem_pot'
        result : numpy ndarray
            array of oxide values, with the computed ferric-ferrous ratio
            appropriate to the imposed log f O2. The oxides are in standard
            order: SiO2, TiO2, Al2O3, Fe2O3, Cr2O3, FeO, MnO, MgO, NiO, CoO,
            CaO, Na2O, K2O, P2O5, H2O, [CO2]
        """
        assert isinstance(m, np.ndarray), 'm must be an numpy array.'
        assert m.size == 15 or m.size == 16, \
            'm must be a 1-d numpy array of length 15 or 16'
        assert mtype == 'components', 'the value of mtype must be "components"'
        assert compute == 'logfO2' or compute == 'chem_pot' or \
               compute == 'oxides', \
            'the value of compute must be "logfO2" or "chem_pot" or "oxides"'
        t0 = 1673.15
        a = 0.196
        b = 1.1492e4
        c = -6.675
        e = -3.364
        f = -7.01e-7 * 1.0e5
        g = -1.54e-10 * 1.0e5
        h = 3.85e-17 * 1.0e5 * 1.0e5
        d = np.array([0, 0, -2.243, 0, 0, -1.828, 0, 0, 0, 0, 3.201, 5.854,
                      6.215, 0, 0, 0])
        # 'SiO2', 'TiO2', 'Al2O3', 'Fe2O3', 'MgCr2O4', 'Fe2SiO4', 'MnSi0.5O2',
        # 'Mg2SiO4', 'NiSi0.5O2', 'CoSi0.5O2', 'CaSiO3', 'Na2SiO3', 'KAlSiO4',
        # 'Ca3(PO4)2', 'H2O', 'CO2'
        ox = np.array([m[0] + m[5] + m[6] / 2. + m[7] + m[8] / 2. + m[9] / 2. + m[10] + m[11] + m[12],  # SiO2
                       m[1],  # TiO2
                       m[2] + m[12] / 2.,  # Al2O3
                       m[3],  # Fe2O3
                       m[4],  # Cr2O3
                       m[5] * 2.,  # FeO
                       m[6],  # MnO
                       m[4] + m[7] * 2.0,  # MgO
                       m[8],  # NiO
                       m[9],  # CoO
                       m[10] + m[13] * 3.,  # CaO
                       m[11],  # Na2O
                       m[12] / 2.,  # K2O
                       m[13],  # P2O5
                       m[14]  # H2O
                       ])
        if debug > 0:
            print('kc Fe3+/Fe2+ input grams Fe2O3, FeO',
                  ox[3] * core.chem.oxide_props['molwt'][3],
                  ox[5] * core.chem.oxide_props['molwt'][5])
        if m.size == 15:
            ox = np.append(ox, 0.0)
        elif m.size == 16:
            ox = np.append(ox, m[15])  # CO2
        tot = np.sum(ox) + m[3]
        if (m[3] == 0.0 and m[5] == 0.0) or (tot == 0.0):
            if debug > 0:
                print('kc Fe3+/Fe2+ no iron in input composition.')
            if compute == 'logfO2' or compute == 'chem_pot':
                return 0.0
            else:
                return ox
        if compute == 'logfO2' or compute == 'chem_pot':
            temp = b / t + c + e * (1.0 - t0 / t - np.log(t / t0)) + f * p / t \
                   + g * (t - t0) * p / t + h * p * p / t
            temp += (np.dot(ox, d) + 2.0 * d[5] * ox[3] - d[3] * ox[3]) / tot
            logfo2 = (np.log(ox[3] / ox[5]) - temp) / (a * np.log(10.0))
            if debug > 0:
                print('kc Fe3+/Fe2+ log fO2', logfo2)
            return logfo2 if compute == 'logfO2' \
                else 8.3144598 * t * np.log(10.0) * logfo2
        elif compute == 'oxides':
            ox[5] += 2.0 * ox[3]
            ox[3] = 0.0
            logfO2 = self.log10NNO(t, p) + deltaNNO
            temp = a * np.log(10.0) * logfO2 + b / t + c + e * (1.0 - t0 / t \
                                                                - np.log(t / t0)) + f * p / t + g * (
                           t - t0) * p / t + h * p * p / t
            temp += np.dot(ox, d) / tot
            temp = np.exp(temp)
            ox[3] = temp * ox[5] / (1.0 + 2.0 * temp)
            ox[5] -= 2.0 * ox[3]
            if debug > 0:
                print('kc Fe3+/Fe2+ comp  grams Fe2O3, FeO',
                      ox[3] * core.chem.oxide_props['molwt'][3],
                      ox[5] * core.chem.oxide_props['molwt'][5])
            return ox
        else:
            return None

    def kc_print_state(self, state=None, silent=False):
        """
        Prints information about the  current system state in terms of the Kress
        and Carmichael (1991) ferrous-ferric equilibrium model

        Parameters
        ----------
        state : EquilState, default None
            An instance of the EquilState class generated by this method. This
            parameter is only specified for a sequence of equilibrium
            calculations where the state of the system is initialized from a
            previous computation.
        silent : bool, default False
            Flag to silence printing

        Returns
        -------
        result : tuple or None
            Returns None if state is None or the system does not contain an
            omnicomponent phase name 'Liquid', else returns a tuple with the
            computer log 10 fO2 of the liquid phase relative to the NNO oxygen
            buffer, the value of log 10 fO2 on the buffer, and the total number
            of moles of oxygen in the system
        """
        if state is None:
            print('Current system "state" must be provided.')
            return None
        if 'Liquid' not in state.phase_d:
            print('Current system must contain the phase "Liquid".')
            return None
        t = state.temperature
        p = state.pressure
        moles = state.phase_d['Liquid']['moles']
        logfO2 = self.kc_ferric_ferrous(t, p, moles)
        NNOfO2 = self.log10NNO(t, p)
        moles_O = state.tot_moles_elements()[state.element_l.index(8)]
        if not silent:
            print('System log 10 fO2', logfO2 - NNOfO2, 'realtive to NNO.')
            print('System log 10 NNO', NNOfO2)
            print('Moles of O in system', moles_O)
        return (logfO2 - NNOfO2, NNOfO2, moles_O)

    def execute(self, t=1000.0, p=1000.0, bulk_comp=None, state=None,
                con_deltaNNO=None, debug=0, stats=False) -> EquilState:
        """
        Calculates an equilibrium assemblage, returning the results as an
        instance of the EquilState class

        Parameters
        ----------
        t : float, default 1000.0
            Temperature in Kelvins
        p : float, default 1000.0
            Pressure in bars
        bulk_comp : numpy array, default None
            Bulk composition of the system. Array of element concentrations,
            of the length and order of self.element_list
        state : EquilState, default None
            An instance of the EquilState class generated by this method. This
            parameter is only specified for a sequence of equilibrium
            calculations where the state of the system is initialized from a
            previous computation.
        con_deltaNNO : float or None
            A non-None value for this parameter is ignored unless the system
            contains an omnicomponnet liquid phase of either 15 or 16 endmember
            components modeled after Ghiorso and Sack (1995) or Ghiorso and
            Gualda (2015), i.e. a MELTS silicate liquid thermodynamic model. If
            a value is set, the liquid is constrained to follow the nickel-
            nickel oxide oxygen buffer plus the offset value, con_deltaNNO,
            which is specified in log 10 fO2 units. For example, a value of 1.0
            will force the system to equilibrate on the NNO+1 oxygen buffer by
            adjusting the ratio of ferric and ferrous iron according to the
            Kress and Carmichael (1991) calibration. Setting a value permits
            the system to be open to oxygen transfer; the bulk moles of
            oxygen in the system is no longer constrained by input bulk
            composition.
        debug : int, default 0
            Level of detail printed by the method:

            - 0, no information

            - 1, minimal progress information

            - 2, normal debugint output level

            - 3, verbose debuging output level

        stats : bool, default False
            Toggle printing of statistics associated with calculation progress

        Returns
        -------
        state: EquilState
            An instance of the EquilState class that contains the equilibrium
            state of the system

        Notes
        -----
        Both the bulk_comp and state parameters cannot be set simultaneously.
        """
        if state is not None and bulk_comp is not None:
            assert False, \
                'Both "bulk_comp" and "state" parameters cannot be specified.'
        if bulk_comp is not None:
            try:
                self.bulk_comp = bulk_comp.astype(float)
            except:
                assert False, 'bulk_comp must be set to an numpy array'
        if self.bulk_comp is None and state is None:
            print('Bulk composition of the system must be set.')
            return None

        # Initialize the EquilState class instance if one is not specified
        require_quad_pass_before_phase_addition = False
        if state is None:
            state = EquilState(self.element_list, self.phase_list)
            # Currently, one phase must be omnicomponent
            omni_phase = state.omni_phase()
            if omni_phase is None:
                print("There is no omnicomponent phase in the system.")
                return None
            # Assume the system inititally has just the omnicomponent phase
            state.set_phase_comp(omni_phase, self.bulk_comp,
                                 input_as_elements=True)
            force_once_quad_pass = False

            state.temperature = t
            state.pressure = p

            if self.lagrange_list is not None:
                self._compute_null_space(state, debug)
                force_once_quad_pass = True
                if not self.entropy and not self.volume:
                    # imposed chemical potential constraints
                    require_quad_pass_before_phase_addition = True

        else:
            omni_phase = state.omni_phase()
            force_once_quad_pass = True
            state.temperature = t
            state.pressure = p
        if debug > 0:
            print(omni_phase + ' is the omnicomponent phase.')

        ne = len(self.element_list)

        if self.entropy or self.volume:
            tp_l = [t, p]
            self._apply_non_linear_constraints(tp_l, state, debug)
            (t, p) = tuple(tp_l)
            state.temperature = t
            state.pressure = p
            if debug > 0 and self.entropy:
                print('Final corrected temperature', t)
            if debug > 0 and self.volume:
                print('Final corrected pressure', p)

        if con_deltaNNO is not None:
            assert omni_phase is not None, \
                'Specifying a value for con_deltaNNO requires an ' \
                + 'omnicomponent phase in the assemblage.'
            assert omni_phase == 'Liquid', \
                'The omnicomponent phase must be named "Liquid" ' \
                + 'and must inherit from the silicate liquid model of ' \
                + 'Ghiorso and Sack (1995) or Ghiorso and Gualda (2015).'
            omni_nc = state.phase_d['Liquid']['nc']
            assert omni_nc == 15 or omni_nc == 16, \
                'Liquid phase must have 15 or 16 endmember components.'

        ###########################
        # Calculation starts here #
        ###########################

        while True:
            # Special case: adjust ferric-ferrous of silicate liquid
            if con_deltaNNO is not None:
                omni_moles = state.phase_d['Liquid']['moles']
                omni_oxide = self.kc_ferric_ferrous(t, p, omni_moles,
                                                    compute='oxides', deltaNNO=con_deltaNNO, debug=debug)
                omni_moles, oxide_res = \
                    state.phase_d['Liquid']['obj'].calc_endmember_comp(
                        mol_oxide_comp=omni_oxide, method='intrinsic',
                        output_residual=True)
                state.set_phase_comp(omni_phase, omni_moles)

            # Obtain chemical potentials of elements in the system
            mu_elm = state.dGdn(t, p, element_basis=True, use_omni_phase=True)

            # If a phase has zero moles, calculate its chemical affinity
            A_min, A_min_name = self._calc_sat_state_affinities(state, t, p,
                                                                mu_elm, omni_phase=omni_phase, debug=debug)

            # if a phase is supersaturated, add it to the system
            if A_min >= 0.0 and not force_once_quad_pass:
                if debug > 0:
                    print('Stable phase assemblage computed. Exiting.')
                break
            force_once_quad_pass = False

            if A_min < 0.0 and not require_quad_pass_before_phase_addition:
                if debug > 0:
                    print('Adding phase ', A_min_name,
                          ' to system. with affinity = ', A_min)
                if stats:
                    print('Add:', A_min_name)
                okay = self._add_phase_to_system(state, A_min_name)
                if okay and debug > 1:
                    print('... phase successfully added.')
                elif not okay:
                    print('... phase addition unsuccessful. Exiting.')
                    break

            require_quad_pass_before_phase_addition = False
            quad_loop = True
            quad_iter = 0
            quad_energies = []
            while quad_loop:
                # Deal with equality/inequality constraints
                if (quad_iter == 0) or self.lagrange_list is not None:
                    """
                    Create projection operator to strip columns from to
                    account for missing composition variables
                    """
                    result = []
                    for entry in state.phase_d.values():
                        if state.tot_moles_phase(entry['name']) > 0.0:
                            nc = entry['nc']
                            result.append(np.reshape(entry['mole_nz'], (nc, 1)))
                    P_nz = np.vstack(tuple(result))
                    if self.entropy:
                        P_nz = np.vstack((P_nz, np.ones(1)))
                    if self.volume:
                        P_nz = np.vstack((P_nz, np.ones(1)))
                    P_nz = np.diag(P_nz[:, 0])
                    """
                    now remove zero columns of the projection matrix
                    taken from: https://stackoverflow.com/questions/51769962/
                    find-and-delete-all-zero-columns-from-numpy-array-using-fancy-indexing/51770365
                    """
                    P_nz = np.delete(P_nz, np.argwhere(np.all(P_nz[..., :] == 0,
                                                              axis=0)), axis=1)
                    if debug > 1:
                        print('Zero element projection matrix', P_nz.shape)
                        if debug > 2:
                            print(P_nz)
                        else:
                            self._print_matrix_repr(P_nz)
                    A, df, Q1, Q2, R11 = self._compute_a_and_qr(t, p, state,
                                                                P_nz, debug)
                    if debug > 0:
                        print('Exit constraint matrix and RQ decomposition.')
                        if debug > 1:
                            print('... df', df, 'A', A.shape)
                            if debug > 2:
                                print(A)
                            else:
                                self._print_matrix_repr(A)
                        if debug > 1:
                            print('... Q1', Q1.shape)
                            if debug > 2:
                                print(Q1)
                            else:
                                self._print_matrix_repr(Q1)
                        if debug > 1:
                            print('... Q2', Q2.shape)
                            if debug > 2:
                                print(Q2)
                            else:
                                self._print_matrix_repr(Q2)
                        if debug > 1:
                            print('... R11', R11.shape)
                            if debug > 2:
                                print(R11)
                            else:
                                self._print_matrix_repr(R11)

                # Compute gradient and Hessian
                g = state.dGdn(t, p, element_basis=False)
                H = state.d2Gdn2(t, p, element_basis=False)
                if debug > 1:
                    print('Gradient', g.shape)
                    if debug > 2:
                        print(g)
                if debug > 1:
                    print('Hessian', H.shape)
                    if debug > 2:
                        print(H)
                    else:
                        self._print_matrix_repr(H)
                if self.lagrange_list is not None:
                    g, H = self._augment_gradient_hessian(t, p, g, H, state,
                                                          debug)
                    l_mult = self._compute_lagrange_multipliers(
                        g if P_nz.shape[0] == P_nz.shape[1] else np.matmul(
                            P_nz.T, g), A, debug)
                    H = self._augment_hessian_return_wronskian(t, p, H,
                                                               l_mult, state, debug)
                    if debug > 1:
                        print('Final Hessian', H.shape)
                        if debug > 2:
                            print(H)
                        else:
                            self._print_matrix_repr(H)
                # projection gradient and Hessian into the null space
                if P_nz.shape[0] > P_nz.shape[1]:
                    g = np.matmul(P_nz.T, g)
                    H = np.matmul(P_nz.T, np.matmul(H, P_nz))
                g_p = np.matmul(Q2, g)
                H_p = np.matmul(np.matmul(Q2, H), Q2.T)
                if debug > 1:
                    print('Projected gradient:', g_p.shape)
                    if debug > 2:
                        print(g_p)
                if debug > 1:
                    print('Projected Hessian:', H_p.shape)
                    if debug > 2:
                        print(H_p)
                    else:
                        self._print_matrix_repr(H_p)

                # Exit if the problem is trivial
                if H_p.size == 0 and g_p.size == 0:
                    quad_normal_exit = True
                    break

                """
                Solve the null space problem. Could use self.eps_rank here for
                the condition number, but the default value computed setting
                rcond=None in linalg.lstsq is usually optimal
                """
                if self.use_numpy_lstsq:
                    result, residuals, rank, S = np.linalg.lstsq(H_p, -g_p,
                                                                 rcond=None)
                    if debug > 0:
                        print('')
                        print('Quadratic solution, SVD algorithm, rank', rank)
                    if debug > 1:
                        print(result)
                        if debug > 2:
                            print('... residuals:')
                            print(residuals)
                            print('... S:')
                            print(S)
                else:
                    result = sm.OLS(-g_p, H_p, hasconst=False).fit(method='qr')
                    if debug > 0:
                        print('')
                        print('Quadratic solution, QR algorithm.')
                        if debug > 1:
                            print(result.params.shape)
                            print(result.params)
                            if debug > 2:
                                print(result.summary())
                    result = np.reshape(result.params, (result.params.shape[0], 1))
                n2 = np.matmul(Q2.T, result)
                if P_nz.shape[0] > P_nz.shape[1]:
                    n2 = np.matmul(P_nz, n2)
                if debug > 0:
                    print('Re-inflated delta n solution vector:', n2.shape)
                    for ind in range(0, n2.shape[0]):
                        if n2[ind][0] == 0.0:
                            print('0.0', end=' ')
                        else:
                            print("{0:10.3e}".format(n2[ind][0]), end=' ')
                    print('')
                norm = np.linalg.norm(n2)
                if debug > 0:
                    print('Norm: ', norm, 'Op, sub-Op:', self.eps_quad_optimal,
                          self.eps_quad_suboptimal)
                if stats:
                    print('Quad ({0:03d}) norm: {1:20.13e}'.format(quad_iter,
                                                                   norm), end='')
                if norm < self.eps_quad_optimal:
                    quad_normal_exit = True
                    break
                elif quad_iter == self.max_quad_iters:
                    if norm < self.eps_quad_suboptimal:
                        print(' Quadratic norm is acceptable, but not optimal.')
                        quad_normal_exit = True
                        break

                # Perform a linear search along the quadratic search direction
                n_ref = []
                for name, entry in state.phase_d.items():
                    if state.tot_moles_phase(name) > 0:
                        for i in range(0, entry['nc']):
                            n_ref.append(entry['moles'][i])
                if self.lagrange_list is not None:
                    row = 0
                    for c, f in self.lagrange_list:
                        if type(c) is dict:
                            n_ref.append(self.lagrange_moles[row])
                        elif type(c) is str:
                            n_ref.append(t if c == 'T' else p)
                        row += 1
                n_ref = np.array(n_ref)
                if debug > 2:
                    print('Reference phase/component moles', n_ref.shape,
                          np.linalg.norm(n_ref))
                    print(n_ref)

                # Potential function to be minimized
                def linear_search(step, state, t, p, n_ref, n2):
                    index = 0
                    for name, entry in state.phase_d.items():
                        if state.tot_moles_phase(name) > 0:
                            for i in range(0, entry['nc']):
                                entry['moles'][i] = n_ref[index] + step * n2[index]
                                index += 1
                            if entry['nc'] > 1:
                                assert entry['obj'].test_endmember_comp(
                                    entry['moles']), 'Bounds failure for ' + name
                    if self.entropy:
                        t = n_ref[index] + step * n2[index]
                        index += 1
                    if self.volume:
                        p = n_ref[index] + step * n2[index]
                        index += 1
                    if self.entropy or self.volume:
                        tp_l = [t, p]
                        self._apply_non_linear_constraints(tp_l, state, debug)
                        (t, p) = tuple(tp_l)
                    result = state.G(t, p)
                    if self.lagrange_list is not None:
                        dgdn = state.dGdn(t, p, element_basis=False)
                        row = 0
                        for c, f in self.lagrange_list:
                            if type(c) is dict:
                                moles = state.moles_v(self.reactions[row])
                                contrib = moles * (
                                    np.matmul(self.reactions[row], dgdn)[0])
                                if debug > 1:
                                    print('... linear search. ', end=' ')
                                    print('ref = ', n_ref[index], end=' ')
                                    print('index = ', index, 'i = ', row, end=' ')
                                    print('moles = ', moles, end=' ')
                                    print('contrib ', contrib)
                                result -= contrib
                                self.lagrange_moles[row] = moles
                                index += 1
                                row += 1
                        if self.entropy:
                            result -= t * state.dGdT(t, p)
                        if self.volume:
                            result -= p * state.dGdP(t, p)
                    return result

                # Determine domain bounds on the potential linear search
                def test_bounds_for_search(step, state, t, p, n_ref, n2):
                    index = 0
                    for name, entry in state.phase_d.items():
                        if state.tot_moles_phase(name) > 0:
                            nc = entry['nc']
                            moles = np.empty(nc)
                            for i in range(0, nc):
                                moles[i] = n_ref[index] + step * n2[index]
                                index += 1
                            if nc == 1:
                                if moles[0] < 0.0:
                                    return False
                            else:
                                if not entry['obj'].test_endmember_comp(moles):
                                    return False
                    if self.entropy:
                        t_est = n_ref[index] + step * n2[index]
                        if (t_est < 773.15) or (t_est > 2773.15):
                            return False
                        index += 1
                    if self.volume:
                        p_est = n_ref[index] + step * n2[index]
                        if (p_est < 1.0) or (p_est > 100000.0):
                            return False
                        index += 1
                    return True

                step_max = 2.0
                while not test_bounds_for_search(step_max, state, t, p, n_ref,
                                                 n2):
                    step_max *= 0.9
                    if step_max < sys.float_info.epsilon:
                        step_max = 0.0
                        break
                if debug > 1:
                    print('Maximum linear step search length; ', step_max)

                step_min = -2.0
                while not test_bounds_for_search(step_min, state, t, p, n_ref,
                                                 n2):
                    step_min *= 0.9
                    if step_min > -sys.float_info.epsilon:
                        if step_max == 0.0:
                            step_min = 0.0
                            break
                        else:
                            step_min = sys.float_info.epsilon
                            while not test_bounds_for_search(step_min, state, t, p, n_ref,
                                                            n2):
                                step_min *= 1.1
                                if step_min > step_max:
                                    step_min = step_max
                                    break
                        # step_min = 0.0
                        # break
                if debug > 1:
                    print('Minimum linear step search length; ', step_min)
                
                result = sp.optimize.minimize_scalar(linear_search,
                                                    bounds=(step_min, step_max),
                                                    args=(state, t, p, n_ref, n2), method='Bounded',
                                                    options={'maxiter': self.max_linear_iters,
                                                            'disp': debug, 'xatol': self.eps_linear})
                step_val = result.x[0] if isinstance(result.x, np.ndarray
                                                    ) else result.x
                func_val = result.fun[0] if isinstance(result.fun, np.ndarray
                                                    ) else result.fun
                if stats:
                    print(' Lin ({0:03d}) step: {1:20.13e} func: {2:20.13e}'.format(
                        result.nfev, step_val, func_val))
                # insure that state is evaluated at the last estimate for step
                linear_search(step_val, state, t, p, n_ref, n2)
                quad_energies.append(func_val)
                
                if self.entropy and self.volume:
                    t_est = n_ref[-2] + step_val * n2[-2]
                    p_est = n_ref[-1] + step_val * n2[-1]
                    t = t_est[0] if isinstance(t_est, np.ndarray) else t_est
                    p = p_est[0] if isinstance(p_est, np.ndarray) else p_est
                elif self.entropy:
                    t_est = n_ref[-1] + step_val * n2[-1]
                    t = t_est[0] if isinstance(t_est, np.ndarray) else t_est
                elif self.volume:
                    p_est = n_ref[-1] + step_val * n2[-1]
                    p = p_est[0] if isinstance(p_est, np.ndarray) else p_est

                if debug > 0:
                    print('Optimal linear search length:', step_val)
                    print('Optimal function value', func_val)
                    state.print_state()
                    if debug > 1:
                        print(result)
                    if self.entropy:
                        print('Approximate corrected temperature', t)
                    if self.volume:
                        print('Approximate corrected pressure', p)

                if self.reactions is not None:
                    blk_cmp = state.tot_moles_elements()
                    diff = blk_cmp - self.bulk_comp
                    self._lagrange_moles += np.matmul(self.CTf, diff)
                    if debug > 1:
                        print('Change in bulk composition:', diff.shape)
                        print(diff)
                        print('lagrange_moles', self.lagrange_moles.shape)
                        print(self.lagrange_moles)
                if self.entropy or self.volume:
                    tp_l = [t, p]
                    self._apply_non_linear_constraints(tp_l, state, debug)
                    (t, p) = tuple(tp_l)
                    state.temperature = t
                    state.pressure = p
                    if debug > 0 and self.entropy:
                        print('Final corrected temperature', t)
                    if debug > 0 and self.volume:
                        print('Final corrected pressure', p)
                if con_deltaNNO is not None:
                    omni_moles = state.phase_d['Liquid']['moles']
                    omni_oxide = self.kc_ferric_ferrous(t, p, omni_moles,
                                                        compute='oxides', deltaNNO=con_deltaNNO, debug=debug)
                    omni_moles, oxide_res = \
                        state.phase_d['Liquid']['obj'].calc_endmember_comp(
                            mol_oxide_comp=omni_oxide, method='intrinsic',
                            output_residual=True)
                    state.set_phase_comp(omni_phase, omni_moles)

                quad_iter += 1
                if quad_iter > self.max_quad_iters:
                    quad_normal_exit = False
                    break
                if len(quad_energies) > self.equil_linear_min:
                    last_three = quad_energies[-3:]
                    average = (last_three[0] + last_three[1] + last_three[2]) / 3.0
                    sumsqr = (last_three[0] - average) ** 2 + (last_three[1] -
                                                               average) ** 2 + (last_three[2] - average) ** 2
                    if debug > 0:
                        print('Quad pot:', quad_energies[-1], 'Ave:',
                              average, 'SS:', sumsqr)
                        if debug > 2:
                            print(last_three)
                    if np.sqrt(sumsqr) < np.fabs(average) * self.eps_minimal_energy:
                        quad_normal_exit = True
                        if stats or debug > 0:
                            print('Minimal energy termination of quadratic loop.')
                        break

                rem_phases_l = []
                for name, entry in state.phase_d.items():
                    mol_p = state.tot_moles_phase(name)
                    if mol_p > 0.0 and mol_p < self.moles_out:
                        self._remove_phase_from_system(state, name)
                        if stats:
                            print('Remove:', name)
                        if debug > 0:
                            print('----------')
                            print('--> Phase', name, 'removed from the system.')
                            print('----------')
                        if name[-2:-1] == '_':
                            rem_phases_l.append(name)
                        else:
                            if name + '_1' in state.phase_d:
                                state.phase_d[name] = state.phase_d[name+'_1']
                                rem_phases_l.append(name+'_1')
                        quad_iter = 0
                for name in rem_phases_l:
                    del state.phase_d[name]
            # termination of quad_loop
            if quad_normal_exit:
                if stats and debug>0:
                    print('')
                for phase_name in state.phase_d.keys():
                    result, result_comp = self._determine_phase_stability(state,
                                                                          t, p, phase_name, debug=debug)
                    if result:
                        break
                if result:
                    if stats:
                        print('Unmixing:', phase_name)
                    if debug > 0:
                        print('Phase separation found for', phase_name)
                        print('   composition:', result_comp)
                    new_phase = phase_name + '_1'
                    ident = 1
                    while new_phase in state.phase_d:
                        ident += 1
                        new_phase = phase_name + '_' + str(ident)
                    state.phase_d[new_phase] = state.phase_d[phase_name].copy()
                    state.phase_d[new_phase]['allow_unmixing'] = False
                    entry = state.phase_d[new_phase]
                    entry['moles'] = np.zeros(entry['nc'])
                    entry['mole_frac'] = np.zeros(entry['nc'])
                    for i in range(0, entry['nc']):
                        entry['mole_frac'][i] = result_comp[i]
                    state._c_matrix = np.vstack(
                        (state._c_matrix, state.c_matrix_for_phase(phase_name)))
                    if debug > 2:
                        print('Old phase dictionary ...')
                        print(state.phase_d[phase_name])
                        print('New phase dictionary ...')
                        print(state.phase_d[new_phase])
                    okay = self._add_phase_to_system(state, new_phase)
                    if okay and debug > 1:
                        print('... phase successfully added.')
                    elif not okay:
                        print('... phase addition unsuccessful. Exiting.')
                    force_once_quad_pass = True
            else:
                print('No check made for phase separation.')
        # termination of phase saturation loop
        return state

class MELTSmodel:
    """Class for creating an instance of the Equilibrate PhaseObjC class that is tailored to
    calculate equilibrium phase assemblages using one of the MELTS calibrations.

    Valid initializers are version='1.0.2', '1.1.0', '1.2.0', '5.6.1', 'DEW', 'OnlyDEW'

    """

    def __init__(self, version='1.0.2'):
        if version == '1.0.2':
            MELTSclass = ObjCClass('EquilibrateUsingMELTSv102')
        elif version == '1.1.0':
            MELTSclass = ObjCClass('EquilibrateUsingMELTSv110')
        elif version == '1.2.0':
            MELTSclass = ObjCClass('EquilibrateUsingMELTSv120')
        elif version == '5.6.1':
            MELTSclass = ObjCClass('EquilibrateUsingpMELTSv561')
        elif version == 'DEW':
            MELTSclass = ObjCClass('EquilibrateUsingMELTSwithDEW')
        elif version == 'OnlyDEW':
            MELTSclass = ObjCClass('EquilibrateUsingMELTSandOnlyDEW')
        else:
            assert False, 'Unknown version of MELTS stipulated'
        self.melts = MELTSclass.alloc().init()
        self.melts.setDebugS_(0)
        self.melts.setDebugV_(0)

        oxide_names_NSArray = self.melts.oxideNames()
        self.noxides = len(oxide_names_NSArray)  # was .count converted to len() in going from 0.2.7 -> 0.2.10
        self.oxide_names_a = [str(oxide_names_NSArray.objectAtIndex_(i)) for i in range(self.noxides)]

        phase_names_NSArray = self.melts.phaseNames()
        self.nphases = len(phase_names_NSArray)  # was .count converted to len() in going from 0.2.7 -> 0.2.10
        self.phase_names_a = [str(phase_names_NSArray.objectAtIndex_(i)) for i in range(self.nphases)]

        locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8')

    def print_default_settings(self):
        """Prints list of options and tolerance settings for algorithm.

        """
        option_settings = {
            'PPOPTIONS_DISTRIBUTION': (
                'Make failure in element redistribution non-fatal', self.melts.PPOPTIONS_DISTRIBUTION),
            'PPOPTIONS_CHEM_POTENTIAL': (
                'Make chemical potential recast errors non-fatal', self.melts.PPOPTIONS_CHEM_POTENTIAL),
            'PPOPTIONS_RANK': ('Make rank deficient quadratic solutions fatal', self.melts.PPOPTIONS_RANK),
            'PPOPTIONS_ZERO_LINEAR': (
                'Make "zero" steplength linear searches non-fatal', self.melts.PPOPTIONS_ZERO_LINEAR),
            'PPOPTIONS_QUAD_ITERS': (
                'Terminate if quadratic iterations ever exceed maximum', self.melts.PPOPTIONS_QUAD_ITERS),
            'PPOPTIONS_FORCE_COMPONENTS': (
                'Never zero a component concentration in a system phase', self.melts.PPOPTIONS_FORCE_COMPONENTS),
            'PPOPTIONS_DETECT_MINIMAL': (
                'Prevent minimal energy test for convergence', self.melts.PPOPTIONS_DETECT_MINIMAL)
        }
        for key in option_settings:
            value = option_settings[key]
            print("{0:<60.60s} {1:<30.30s} {2:1d}".format(value[0], key, value[1]))

        tolerance_settings = {
            'PPPARAMETERS_CHEM_POTENTIAL': \
                ('Residual tolerance for acceptable recasting of phase chemical potentials to those of the elements', \
                 float(self.melts.PPPARAMETERS_CHEM_POTENTIAL)),
            'PPPARAMETERS_RANK': \
                ('Tolerance (relative to Hessian norm) for computation of pseudorank during quadratic search', \
                 self.melts.PPPARAMETERS_RANK),
            'PPPARAMETERS_QUAD_OPTIMAL': \
                ('Residual norm relative tolerance for optimal quadratic convergence', \
                 self.melts.PPPARAMETERS_QUAD_OPTIMAL),
            'PPPARAMETERS_QUAD_SUBOPTIMAL': \
                ('Residual norm relative tolerance for sub-optimal quadratic convergence', \
                 self.melts.PPPARAMETERS_QUAD_SUBOPTIMAL),
            'PPPARAMETERS_QUAD_MAX_ITERS': \
                ('Maximal number of quadratic iterations', \
                 self.melts.PPPARAMETERS_QUAD_MAX_ITERS),
            'PPPARAMETERS_LINEAR_MAX_ITERS': \
                ('Maximal number of linear search iterations', \
                 self.melts.PPPARAMETERS_LINEAR_MAX_ITERS),
            'PPPARAMETERS_LINEAR_RELTEST': \
                ('Relative convergence criteria for estimation of the minimum during a linear search step', \
                 self.melts.PPPARAMETERS_LINEAR_RELTEST),
            'PPPARAMETERS_LINEAR_MIN_STEPLENGTH': \
                ('Minimum allowed steplength in a linear search step', \
                 self.melts.PPPARAMETERS_LINEAR_MIN_STEPLENGTH),
            'PPPARAMETERS_COMPONENT_MINIMUM': \
                ('Minimum molar concentration of a component in a phase (absolute value)', \
                 self.melts.PPPARAMETERS_COMPONENT_MINIMUM),
            'PPPARAMETERS_MASSOUT': \
                ('Phase mass that triggers removal of that phase from the system', \
                 self.melts.PPPARAMETERS_MASSOUT),
            'PPPARAMETERS_MOLESIN': \
                ('Moles of phase added to system on detection of phase saturation', \
                 self.melts.PPPARAMETERS_MOLESIN),
            'PPPARAMETERS_LINEAR_MINIMAL': \
                ('Number of quadratic iterations over which to average potential for minimal energy convergence test', \
                 self.melts.PPPARAMETERS_LINEAR_MINIMAL),
            'PPPARAMETERS_CYCLES_MAX': \
                ('Maximal number of phase inclusion/phase removal cycles permitted', \
                 self.melts.PPPARAMETERS_CYCLES_MAX)
        }
        for key in tolerance_settings:
            value = tolerance_settings[key]
            print("{0:<70.70s} {1:<30.30s} {2:13.6e}".format(value[0], key, value[1]))

    def set_debug_state(self, debugS=False, debugV=False):
        """Sets debug output level for Equilibrate class and subclasses

        Parameters
        ==========
        debugS : boolean, optional
            Sets on or off low level debug output. Default is off (False).
        debugV : boolean, optional
            Sets on or off high level debug output. Default is off (False).

        """
        if debugS == False:
            self.melts.setDebugS_(False)
        else:
            self.melts.setDebugS_(True)
        if debugV == False:
            self.melts.setDebugV_(False)
        else:
            self.melts.setDebugV_(True)

    def get_oxide_names(self):
        """Retrieves a list of system oxides

        Composition of the system can be expressed in terms of these oxides.

        Returns
        -------
        array : strings
            An array of strings naming system components in terms of oxides

        """
        return self.oxide_names_a

    def get_phase_names(self):
        """Retrieves a list of system phases

        Names of phases known to the system

        Returns
        -------
        array : strings
            An array of strings naming system phases

        """
        return self.phase_names_a

    def get_phase_inclusion_status(self):
        """Retrieves a dictionary of the inclusion status of each phase

        Returns
        -------
        dict : dictionary
            A dictionary of boolean values indicating the inclusion status of each phase
            (key) known to the system

        """
        dict = {}
        state_NSdict = self.melts.getPermissablePhasesState()
        for phase in self.phase_names_a:
            if sys.platform == "darwin":
                value = state_NSdict.valueForKey_(phase).boolValue # removed in going from 0.2.7 -> 0.2.10
            else:
                value = state_NSdict.valueForKey_(phase)
            if value == 1:
                value = True
            else:
                value = False
            dict[phase] = value
        return dict

    def set_phase_inclusion_status(self, status_d):
        """Sets the inclusion status of specified phases

        Parameters
        ----------
        status_d : dictionary
            A dictionary of phase name keys and boolean values. True sets inclusion, and
            False prevents inclusion of a phase in the equilibrium assemblage.  Note that
            the chemical affinity of the phase will still be calculated even if the
            inclusion level is set to False.
        """
        state_NSdict = self.melts.getPermissablePhasesState()
        nsarray_cls = ObjCClass('NSMutableArray')
        nsarray = nsarray_cls.arrayWithCapacity_(self.nphases)
        for phase in self.phase_names_a:
            if sys.platform == "darwin":
                value = state_NSdict.valueForKey_(phase).boolValue  # removed in going from 0.2.7 -> 0.2.10
            else:
                value = state_NSdict.valueForKey_(phase)
            if phase in status_d:
                if status_d[phase] == True:
                    value = 1
                else:
                    value = 0
            nsarray.addObject_(ObjCClass('NSNumber').numberWithBool_(value))
        self.melts.setPermissablePhasesState_(nsarray)

    def set_bulk_composition(self, oxide_d={}):
        """Sets the bulk composition of the system

        This function first tests if the composition is feasible before setting the bulk
        composition of the system.  You should check to make sure that the composition is
        feasible before proceeding.

        Parameters
        ----------
        oxide_d : a python dictionary
            A dictionary of oxide names and values, e.g. {'SiO2':77.8, 'Al2O3':12.0, ..., 'H2O':3.74}

        Returns
        -------
        boolean : True or False

            True if the composition is feasible, in which case the composition of the
            system is defined.

            False if the composition is infeasible, in which case the composition of the
            system is undefined.

        Notes
        -----
        Feasibility call has yet to be implemented: (Objective-C method call:)
        -(BOOL)compositionIsFeasible:(NSArray \*)compositionInWtPercentOxides forSolution:(id <SolutionPhaseProtocol>)omniComponentPhase;

        """
        wt = (ctypes.c_double * self.noxides)()
        ctypes.cast(wt, ctypes.POINTER(ctypes.c_double))
        for i in range(0, self.noxides):
            wt[i] = 0.0
        for key, value in oxide_d.items():
            index = self.oxide_names_a.index(key)
            wt[index] = value
        self.melts.setComposition_(wt)

    def set_temperature(self, t_c=800.0):
        """Sets the temperature of the system and reinitializes a calculation sequence

        Parameters
        ----------
        t_c : float optional
            Float value to set the system temperature in degrees centigrade. Default is 800 °C.

        """
        self.t = t_c + 273.15
        self.melts.setTemperature_(self.t)

    def set_entropy(self, s):
        """Sets the entropy of the system and reinitializes a calculation sequence

        Parameters
        ----------
        s : float
            Float value to set the system entropy in J/K.

        """
        self.entropy = s
        self.melts.setEntropy_(self.entropy)

    def set_pressure(self, p_mpa=200.0):
        """Sets the pressure of the system and reinitializes a calculation sequence

        Parameters
        ----------
        p_mpa : float optional
            Float value to set the system pressure in mega-Pascals. Default is 200 MPa.

        """
        self.p = p_mpa * 10.0
        self.melts.setPressure_(self.p)

    def set_volume(self, v):
        """Sets the volume of the system and reinitializes a calculation sequence

        Parameters
        ----------
        v : float
            Float value to set the system volume in J/bar

        """
        self.volume = v
        self.melts.setVolume_(self.volume)

    def get_object_for_phase(self, phase_name):
        """Retrieve an object instance for the named phase.

        Parameters
        ----------
        phase_name: string
            Name of phase

        Returns
        -------
        object: instance of a phase class
            null if the phase is not in the stable assemblage
        """
        phase_list = self.melts.dictionaryOfPhasesInSystem  # dictionary of EquilibrateStatePhase instances
        if phase_name in phase_list:
            return phase_list[phase_name].phaseClassInstance
        else:
            return None

    def get_properties_of_DEWFluid(self, property_name='species', T=1000, P=1000):
        """Retrieves a dictionary of properties of the specified type

        DEWFluid must exist in the equilibrium assemblage; otherwise an empty dictionary
        is returned.

        Parameters
        ----------
        property_name: string, optional
            'species' (default) returns a dictionary of species mole fractions in the
            equilibrated solution.

            'mu' returns a dictionary of species chemical potentials in the equilibrated
            solution.

            'activity' returns a dictionary of species activities in the equilibrated
            solution.

        T : float optional
            Temperature in degrees centigrade (default is 1000°C)
        P : float optional
            Pressure in bars (default is 1000 bars)

        Returns
        -------
        dictionary: a python dictionary
            Keys are species names (strings).

            Values are species concentrations in mole fraction, or chemical potential,
            or thermodynamic activity, depending on the value of ``property_name``.
        """
        phase_list = self.melts.dictionaryOfPhasesInSystem  # dictionary of EquilibrateStatePhase instances
        if 'DEWFluid' in phase_list:
            DEW_object = phase_list['DEWFluid'].phaseClassInstance  # instantiated class
            elements = phase_list['DEWFluid'].bulkCompositionInElements
            moles = DEW_object.convertElementsToMoles_(elements.pointerToDouble()).pointerToDouble()
            if property_name == 'species':
                return DEW_object.getSpeciesMoleFractionsForBulkComposition_aT_andP_(moles, T + 273.15, P)
            elif property_name == 'mu':
                ns = DEW_object.numberOfSolutionSpecies()
                mu = DEW_object.chemicalPotentialsOfSpeciesFromMolesOfComponents_andT_andP_(moles, T + 273.15,
                                                                                            P).pointerToDouble()
                result = dict()
                for i in range(0, ns):
                    name = DEW_object.nameOfSolutionSpeciesAtIndex_(i)
                    result[name] = mu[i]
                return result
            elif property_name == 'activity':
                ns = DEW_object.numberOfSolutionSpecies()
                activity = DEW_object.activitiesOfSpeciesFromMolesOfComponents_andT_andP_(moles, T + 273.15,
                                                                                          P).pointerToDouble()
                result = dict()
                for i in range(0, ns):
                    name = DEW_object.nameOfSolutionSpeciesAtIndex_(i)
                    result[name] = activity[i]
                return result
            else:
                return dict()
        else:
            return dict()

    def equilibrate_tp(self, T_a, P_a, initialize=False):
        """Determines the equilibrium phase assemblage at a temperature-pressure point
        or along a series of T-P points.

        The bulk composition of the system must first be set by calling the function:

        ``set_bulk_composition``

        Parameters
        ----------
        t_a : float or numpy array of floats
            Temperature in degrees centigrade.  Either a scaler values or a numpy array
            of float values must be provided.
        p_a : float or numpy array of floats
            Pressure in mega-Pascals. Either a scaler values or a numpy array of float
            values must be provided.

            NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the
            same length.
        initialize : bool, optional
            True if this is a T-, P-point that starts a sequence of calculations.

            False if this is a continuation T-,P-pair or series of pairs.

        Returns
        -------
        output_a : an array of tuples
            tuple = (status, T, P, xmlout):

            * **status** is a string indicating the status of the calculation:
              success/failiure, Reason for success/failure.

            * **T** is a float value corresponding to the temperature in degrees centigrade.

            * **P** is a float value corresponding to the pressure in mega-Pascals.

            * **xmlout** is an xml document tree of the type xml.etree.ElementTree. The
              xml tree contains information on the masses and abundances of all phases in
              the system. ``xmlout`` is utilized as input for a number of functions in this
              package that retrieve properties of the equilibrium assemblage.

        Notes
        -----
        The ``xmlout`` document tree will be expanded to include thermodynamic properties
        of the phases and chemical affinities of phases not present in the equilibrium
        assemblage.

        """
        T_a, P_a = core.fill_array(T_a, P_a)
        output_a = []
        for ind, (T, P) in enumerate(zip(T_a, P_a)):
            if initialize:
                self.melts.setTemperature_(T + 273.15)
                self.melts.setPressure_(P * 10.0)

                nsarray_cls = ObjCClass('NSMutableArray')
                nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

                nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Temperature (°C)'))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Pressure (MPa)'))
                nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
                nsarrayvalues.addObject_(
                    ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

                nsdict_cls = ObjCClass('NSDictionary')
                nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
                self.melts.setCalculationOptions_(nsdict)

            else:
                self.melts.incrementTemperature_(T + 273.15)
                self.melts.incrementPressure_(P * 10.0)
            output_NSDictionary = self.melts.execute()
            xmlout = ET.fromstring(str(self.melts.equilibrateResultsAsXML()))
            output_a.append((str(output_NSDictionary.objectForKey_('status')), T, P, xmlout))
        return output_a

    def equilibrate_sp(self, S_a, P_a, initialize=False):
        """Determines the equilibrium phase assemblage at an entropy-pressure point or
        along a series of S-P points.

        The bulk composition of the system must first be set by calling the function:

        ``set_bulk_composition``

        Parameters
        ----------
        S_a : float or numpy array of floats
            Entropy in Joules per Kelvins.  Either a scaler values or a numpy array of
            float values must be provided.
        P_a : float or numpy array of floats
            Pressure in mega-Pascals. Either a scaler values or a numpy array of float
            values must be provided.

            NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the
            same length.
        initialize : bool, optional
            True if this is a S-, P-point that starts a sequence of calculations.

            False if this is a continuation S-,P-pair or series of pairs.

        Returns
        -------
        output_a : an array of tuples
            tuple = (status, T, P, xmlout):

            * **status** is a string indicating the status of the calculation:
              success/failiure, Reason for success/failure.

            * **T** is a float value corresponding to the temperature in degrees
              centigrade.

            * **P** is a float value corresponding to the pressure in mega-Pascals.

            * **xmlout** is an xml document tree of the type xml.etree.ElementTree. The
              xml tree contains information on the masses and abundances of all phases in
              the system. ``xmlout`` is utilized as input for a number of functions in
              this package that retrieve properties of the equilibrium assemblage.

        Notes
        -----
        The ``xmlout`` document tree will be expanded to include thermodynamic properties
        of the phases and chemical affinities of phases not present in the equilibrium
        assemblage.

        """
        S_a, P_a = core.fill_array(S_a, P_a)
        output_a = []
        for ind, (S, P) in enumerate(zip(S_a, P_a)):
            if initialize:
                self.melts.setEntropy_(S)
                self.melts.setPressure_(P * 10.0)

                nsarray_cls = ObjCClass('NSMutableArray')
                nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

                nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Entropy (J/K-kg)'))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Pressure (MPa)'))
                nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
                nsarrayvalues.addObject_(
                    ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

                nsdict_cls = ObjCClass('NSDictionary')
                nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
                self.melts.setCalculationOptions_(nsdict)

            else:
                self.melts.incrementEntropy_(S)
                self.melts.incrementPressure_(P * 10.0)
            output_NSDictionary = self.melts.execute()
            xmlout = ET.fromstring(str(self.melts.equilibrateResultsAsXML()))
            T = locale.atof(xmlout.find(".//Temperature").text)
            output_a.append((str(output_NSDictionary.objectForKey_('status')), T, P, xmlout))
        return output_a

    def equilibrate_tv(self, T_a, V_a, initialize=False):
        """Determines the equilibrium phase assemblage at a temperature-volume point
        or along a series of T-V points.

        The bulk composition of the system must first be set by calling the function:

        ``set_bulk_composition``

        Parameters
        ----------
        T_a : float or numpy array of floats
            Temperature in degrees centigrade.  Either a scaler values or a numpy array
            of float values must be provided.
        V_a : float or numpy array of floats
            Volume in Joules per bar. Either a scaler values or a numpy array of float
            values must be provided.

            NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the
            same length.
        initialize : bool, optional
            True if this is a T-, V-point that starts a sequence of calculations.

            False if this is a continuation T-,V-pair or series of pairs.

        Returns
        -------
        output_a : an array of tuples
            tuple = (status, T, P, xmlout):

            * **status** is a string indicating the status of the calculation:
              success/failiure, Reason for success/failure.

            * **T** is a float value corresponding to the temperature in degrees centigrade.

            * **P** is a float value correcponding to the pressure in mega-Pascals.

            * **xmlout** is an xml document tree of the type xml.etree.ElementTree.
              The xml tree contains information on the masses and abundances of all phases
              in the system. ``xmlout`` is utilized as input for a number of functions in
              this package that retrieve properties of the equilibrium assemblage.

        Notes
        -----
        The ``xmlout`` document tree will be expanded to include thermodynamic properties
        of the phases and chemical affinities of phases not present in the equilibrium
        assemblage.

        """
        T_a, V_a = core.fill_array(T_a, V_a)
        output_a = []
        for ind, (T, V) in enumerate(zip(T_a, V_a)):
            if initialize:
                self.melts.setTemperature_(T + 273.15)
                self.melts.setVolume_(V)

                nsarray_cls = ObjCClass('NSMutableArray')
                nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

                nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Temperature (°C)'))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Volume (cc/kg)'))
                nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
                nsarrayvalues.addObject_(
                    ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

                nsdict_cls = ObjCClass('NSDictionary')
                nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
                self.melts.setCalculationOptions_(nsdict)

            else:
                self.melts.incrementTemperature_(T + 273.15)
                self.melts.incrementVolume_(V)
            output_NSDictionary = self.melts.execute()
            xmlout = ET.fromstring(str(self.melts.equilibrateResultsAsXML()))
            output_a.append((str(output_NSDictionary.objectForKey_('status')), T, V, xmlout))
        return output_a

    def equilibrate_sv(self, S_a, V_a, initialize=False):
        """Determines the equilibrium phase assemblage at an entropy-volume point
        or along a series of S-V points.

        The bulk composition of the system must first be set by calling the function:

        ``set_bulk_composition``

        Parameters
        ----------
        S_a : float or numpy array of floats
            Entropy in Joules per Kelvins.  Either a scaler values or a numpy array of
            float values must be provided.
        V_a : float or numpy array of floats
            Volume in Joules per bar. Either a scaler values or a numpy array of float
            values must be provided.

            NOTE: If both ``t_a`` and ``p_a`` are arrays, then they must both be the
            same length.
        initialize : bool, optional
            True if this is a S-, V-point that starts a sequence of calculations.

            False if this is a continuation S-,V-pair or series of pairs.

        Returns
        -------
        output_a : an array of tuples
            tuple = (status, T, P, xmlout):

            * **status** is a string indicating the status of the calculation:
              success/failiure, Reason for success/failure.

            * **T** is a float value corresponding to the temperature in degrees centigrade.

            * **P** is a float value corresponding to the pressure in mega-Pascals.

            * **xmlout** is an xml document tree of the type xml.etree.ElementTree.
              The xml tree contains information on the masses and abundances of all phases
              in the system. ``xmlout`` is utilized as input for a number of functions in
              this package that retrieve properties of the equilibrium assemblage.

        Notes
        -----
        The ``xmlout`` document tree will be expanded to include thermodynamic properties
        of the phases and chemical affinities of phases not present in the equilibrium
        assemblage.

        """
        S_a, V_a = core.fill_array(S_a, V_a)
        output_a = []
        for ind, (S, V) in enumerate(zip(S_a, V_a)):
            if initialize:
                self.melts.setEntropy_(S)
                self.melts.setVolume_(V)

                nsarray_cls = ObjCClass('NSMutableArray')
                nsarraykeys = nsarray_cls.arrayWithCapacity_(5)
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('ordinate'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('abscissa'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('imposeBuffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('buffer'))
                nsarraykeys.addObject_(ObjCClass('NSString').stringWithString_('bufferOffset'))

                nsarrayvalues = nsarray_cls.arrayWithCapacity_(5)
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Entropy (J/K-kg)'))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('Volume (cc/kg)'))
                nsarrayvalues.addObject_(ObjCClass('NSNumber').numberWithBool_(0))
                nsarrayvalues.addObject_(ObjCClass('NSString').stringWithString_('QFM'))
                nsarrayvalues.addObject_(
                    ObjCClass('NSNumber').numberWithBool(0))  # Double initialization does not work (?)

                nsdict_cls = ObjCClass('NSDictionary')
                nsdict = nsdict_cls.dictionaryWithObjects_forKeys_(nsarrayvalues, nsarraykeys)
                self.melts.setCalculationOptions_(nsdict)

            else:
                self.melts.incrementEntropy_(S)
                self.melts.incrementVolume_(V)
            output_NSDictionary = self.melts.execute()
            xmlout = ET.fromstring(str(self.melts.equilibrateResultsAsXML()))
            output_a.append((str(output_NSDictionary.objectForKey_('status')), S, V, xmlout))
        return output_a

    def get_list_of_phases_in_assemblage(self, root):
        """Returns a list of phases in the specified equilibrium assemblage.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``

        Returns
        -------
        list : list
            A Python list of all phases in the equilibrium assemblage

        """
        list = []
        for phase in self.phase_names_a:
            if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                list.append(phase)
        return list

    def get_dictionary_of_default_fractionation_coefficients(self, fracLiq=False, fracSolid=True, fracFluid=True,
                                                             fracCoeff=1.0):
        """Returns a dictionary of default coefficients for phase fractionation.

        These coefficients may be modified by the user. They are used when setting
        the extent to which a phase will fractionate from the equilibrium assemblage.

        Parameters
        ----------
        fracLiq : bool, optional
            Flag to indicate if liquid phases should be fractionated from the system.
            Default is False.
        fracSolids : bool, optional
            Flag to indicate if solid phases should be fractionated from the system.
            Default is True.
        fracFluids : bool, optional
            Flag to indicate if fluid phases should be fractionated from the system.
            Default is True.
        fracCoeff : float, optional
            Fractionation coefficient, which gives the fractional extend to which the
            mass of the phase is extracted during phase fractionation.
            Default is 1.0 (100%).

        Returns
        -------
        dict : dictionary
            A Python dictionary keyed on phases with values giving the extent (in fractional
            units) that a phase will fractionation mass.

        Notes
        -----
        This dictionary is provided as input to the function ``fractionate_phases()``.
        The default configuration fractionates all solid/fluid phases and retains liquid.

        """
        dict = {}
        for phase in self.phase_names_a:
            if phase == 'Liquid':
                if fracLiq:
                    dict[phase] = fracCoeff
                else:
                    dict[phase] = 0.0
            elif phase == 'Water' or phase == 'Fluid':
                if fracFluid:
                    dict[phase] = fracCoeff
                else:
                    dict[phase] = 0.00
            else:
                if fracSolid:
                    dict[phase] = fracCoeff
                else:
                    dict[phase] = 0.0
        return dict

    def get_mass_of_phase(self, root, phase_name='System'):
        """Returns the mass of a phase in the specified equilibrium assemblage.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``
        phase_name : string, optional
            The name of the phase whose abundance is requested, or the string 'System',
            which returns the combined mass of all phases in the system.  Default value
            is 'System'.

        Returns
        -------
        value : float
            The mass of the phase in the equilibrium assemblage specified by ``root``, in grams.
            If the specified phase is not in the equilibrium assemblage, a value of zero is retuned.

        """
        if phase_name == 'System':
            value = 0.0
            for phase in self.phase_names_a:
                if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                    value += float(root.find(".//System/Phase[@Type='" + phase + "']/Mass").text)
        else:
            if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                value = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Mass").text)
            else:
                value = 0.0
        return value

    def get_composition_of_phase(self, root, phase_name='System', mode='oxide_wt'):
        """Returns the composition of a phase in the specified equilibrium assemblage
        as a dictionary, with composition tabluated in the specified mode.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``
        phase_name : string, optional
            The name of the phase whose abundance is requested, or the string 'System',
            which returns the combined mass of all phases in the system.  Default value
            is 'System'.
        mode : string, optional
            Controls the contents of the returned dictionary.

            * 'formula' returns a dictionary with the string 'formula' as key and value
              set to a string representation of the phase formula. For pure component
              phases, this is the standard phase formula. For solutions, this is the
              actual formula constructed by weighting the endmember components by their
              mole fractions.

            * 'oxide_wt' returns a dictionary of oxide string keys with values in wt%.
              This is a valid ``mode`` for all ``phase_name`` entries.

            * 'component' returns a dictionary of endmember component keys with values in
              mole fraction. The length of this dictionary will vary dependening on the
              number of components that describe the solution. Pure phases return an empty
              dictionary, as does ``phase_name`` set to 'System'.

            The default value of ``mode`` is 'oxide_wt'.

        Returns
        -------
        dict : dictionary
            A dictionary describing the composition of ``phase_name`` according to the
            ``mode`` specified. The dictionary will be empty if ``phase_name`` is not
            present in the equilibrium assemblage.  It will also be empty for certain
            cases described above under ``mode``.

        """
        dict = {}
        if phase_name == 'System':
            if mode == 'oxide_wt':
                oxides = list(root.findall(".//Composition/Oxide"))
                for oxide in oxides:
                    key = oxide.attrib['Type']
                    value = float(oxide.text)
                    dict[key] = value
        else:
            phase = root.find(".//System/Phase[@Type='" + phase_name + "']")
            if phase != None:
                if mode == 'formula':
                    dict['formula'] = phase.find("Formula").text
                elif mode == 'oxide_wt':
                    oxides = list(phase.findall("Oxide"))
                    for oxide in oxides:
                        key = oxide.attrib['Type']
                        value = float(oxide.text)
                        dict[key] = value
                elif mode == 'component':
                    components = list(phase.findall("Component"))
                    if not components:
                        dict['formula'] = phase.find("Formula").text
                    else:
                        for component in components:
                            key = component.attrib['Name']
                            value = float(component.text)
                            dict[key] = value
        return dict

    def fractionate_phases(self, root, frac_coeff):
        """Fractionates phases from the system.
        Partitions and maintains an internal dictionary of fractionates and automatically
        modifies system bulk composition to reflect fractionation.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``
        frac_coeff : dictionary
            A dictionary keyed on phase names with values that indicate the fraction of
            each phase that should fractionate.

            See get_dictionary_of_default_fractionation_coefficients().

        Returns
        -------
        dict : dictionary
            A dictionary keyed on phase names with values corresponding to a dictionary
            of phase properties. Keys are property names.
        """
        dict = {}

        bc = {}
        bcFactor = locale.atof(root.find(".//Mass").text) / 100.0
        oxides = list(root.findall(".//Composition/Oxide"))
        for oxide in oxides:
            key = oxide.attrib['Type']
            value = locale.atof(oxide.text)
            bc[key] = value

        phases = list(root.findall(".//System/Phase"))
        for phase in phases:
            phase_name = phase.attrib['Type']
            dict[phase_name] = {}
            fraction = frac_coeff[phase_name]
            if fraction > 0.0:
                mass = locale.atof(phase.find("Mass").text)
                oxides = list(phase.findall("Oxide"))
                for oxide in oxides:
                    key = oxide.attrib['Type']
                    value = locale.atof(oxide.text)
                    dict[phase_name][key] = value * fraction * mass / 100.0
                    bc[key] -= value * fraction * mass / 100.0
                    if bc[key] < 0.0:
                        bc[key] = 0.0
                        print("Zeroed: " + key)

        self.set_bulk_composition(bc)
        return dict

    def get_thermo_properties_of_phase_components(self, root, phase_name, mode='mu'):
        """Returns a dictionary of the specified component thermodynamic properties of
        the designated phase.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``.
        phase_name : string
            The name of the phase whose abundance is requested.
        mode : string, optional
            Controls the contents of the returned dictionary.

            * 'mu' returns a dictionary of endmember component keys with values of
              chemical potential. The length of this dictionary will vary depending on
              the number of components that describe the solution. Pure phases return a
              dictionary of unit length, with the phase name as key and their specific
              Gibbs free energy as value (J/g).

            * 'excess' returns a dictionary of endmember component keys with values of
              excess chemical potential. The length of this dictionary will vary
              depending on the number of components that describe the solution.
              Pure phases return a dictionary of unit length, with the phase name as key
              and zero as value.

            * 'activity' returns a dictionary of endmember component keys with values of
              component activity. The length of this dictionary will vary depending on
              the number of components that describe the solution. Pure phases return a
              dictionary of unit length, with the phase name as key and unity as value.

        Returns
        -------
        dict : dictionary
            A dictionary describing the thermodynamic properties of components in
            ``phase_name`` according to the ``mode`` specified. The dictionary will be
            empty if ``phase_name`` is not present in the equilibrium assemblage.

        """
        dict = {}
        phase = root.find(".//System/Phase[@Type='" + phase_name + "']")
        if phase != None:
            if mode == 'mu':
                mus = list(phase.findall("ChemicalPotential"))
                if not mus:
                    value = locale.atof(phase.find("GibbsFreeEnergy").text)
                    mass = float(phase.find("Mass").text)
                    dict[phase_name] = value / mass
                else:
                    for mu in mus:
                        key = mu.attrib['Name']
                        value = locale.atof(mu.text)
                        dict[key] = value

            elif mode == 'excess':
                mus = list(phase.findall("ExcessChemicalPotential"))
                if not mus:
                    dict[phase_name] = 0.0
                else:
                    for mu in mus:
                        key = mu.attrib['Name']
                        value = locale.atof(mu.text)
                        dict[key] = value

            elif mode == 'activity':
                mus = list(phase.findall("ExcessChemicalPotential"))
                if not mus:
                    dict[phase_name] = 1.0
                else:
                    t = locale.atof(root.find(".//Temperature").text) + 273.15
                    for mu in mus:
                        key = mu.attrib['Name']
                        value = locale.atof(mu.text)
                        dict[key] = np.exp(value / 8.3143 / t)

        return dict

    def get_list_of_properties(self):
        """Returns a list of properties reported for each phase in an equilibrium assemblage.

        Returns
        -------
        list : list
            A Python list of all properties of phases in an equilibrium assemblage

        """
        list = ['Mass', 'GibbsFreeEnergy', 'Enthalpy', 'Entropy', 'HeatCapacity', 'DcpDt', 'Volume', 'DvDt', 'DvDp',
                'D2vDt2', 'D2vDtDp', 'D2vDp2', \
                'Density', 'Alpha', 'Beta', 'K', "K'", 'Gamma']
        return list

    def get_units_of_property(self, prop='Mass'):
        """Returns the units of a specified property.

        Returns
        -------
        string : string
            The units of the specified property.  Returns 'none' if property is invalid.

        """
        dict = {'Mass': 'g', 'GibbsFreeEnergy': 'J', 'Enthalpy': 'J', 'Entropy': 'J/K', 'HeatCapacity': 'J/K',
                'DcpDt': 'J/K^2', \
                'Volume': 'J/bar', 'DvDt': 'J/bar-K', 'DvDp': 'J/bar^2', 'D2vDt2': 'J/bar-K^2', 'D2vDtDp': 'J/bar^2-K',
                'D2vDp2': 'J/bar^3', \
                'Density': 'g/cm^3', 'Alpha': '1/K', 'Beta': '1/bar', 'K': 'GPa', "K'": 'none', 'Gamma': 'none'}
        return dict.get(prop)

    def get_property_of_phase(self, root, phase_name='System', property_name='Mass'):
        """Returns the specified property of a phase in the specified equilibrium assemblage.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``
        phase_name : string, optional
            The name of the phase whose property is requested, or the string 'System',
            which returns the combined property of all phases in the system.  Default
            value is 'System'.
        property_name : string, optional
            The name of the property to be returned. Default value is 'Mass'.

        Returns
        -------
        value : float
            The property of the phase in the equilibrium assemblage specified by ``root``,
            in standard units.
            If the specified phase is not in the equilibrium assemblage, a value of zero
            is returned.
            If the property is not in the standard list, a value of zero is returned.

        """
        standard = ['Mass', 'GibbsFreeEnergy', 'Enthalpy', 'Entropy', 'HeatCapacity', 'DcpDt', 'Volume', 'DvDt', 'DvDp',
                    'D2vDt2', 'D2vDtDp', 'D2vDp2']
        derived = ['Density', 'Alpha', 'Beta', 'K', "K'", 'Gamma']

        if property_name in standard:
            if phase_name == 'System':
                value = 0.0
                for phase in self.phase_names_a:
                    if root.find(".//System/Phase[@Type='" + phase + "']/" + property_name) != None:
                        value += locale.atof(root.find(".//System/Phase[@Type='" + phase + "']/" + property_name).text)
            else:
                if root.find(".//System/Phase[@Type='" + phase_name + "']/" + property_name) != None:
                    value = locale.atof(root.find(".//System/Phase[@Type='" + phase_name + "']/" + property_name).text)
                else:
                    value = 0.0
        elif property_name in derived:
            if property_name == 'Density':
                if phase_name == 'System':
                    volume = 0.0
                    mass = 0.0
                    for phase in self.phase_names_a:
                        if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                            volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
                            mass += float(root.find(".//System/Phase[@Type='" + phase + "']/Mass").text)
                    value = mass / volume / 10.0  # g/cc
                else:
                    if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                        volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
                        mass = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Mass").text)
                        value = mass / volume / 10.0  # g/cc
                    else:
                        value = 0.0
            elif property_name == 'Alpha':
                if phase_name == 'System':
                    volume = 0.0
                    dvdt = 0.0
                    for phase in self.phase_names_a:
                        if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                            volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
                            dvdt += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDt").text)
                    value = dvdt / volume
                else:
                    if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                        volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
                        dvdt = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDt").text)
                        value = dvdt / volume
                    else:
                        value = 0.0
            elif property_name == 'Beta':
                if phase_name == 'System':
                    volume = 0.0
                    dvdp = 0.0
                    for phase in self.phase_names_a:
                        if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                            volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
                            dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
                    value = -dvdp / volume
                else:
                    if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                        volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
                        dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
                        value = -dvdp / volume
                    else:
                        value = 0.0
            elif property_name == 'K':
                if phase_name == 'System':
                    volume = 0.0
                    dvdp = 0.0
                    for phase in self.phase_names_a:
                        if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                            volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
                            dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
                    value = -volume / dvdp / 10000.0
                else:
                    if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                        volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
                        dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
                        value = -volume / dvdp / 10000.0
                    else:
                        value = 0.0
            elif property_name == "K'":
                if phase_name == 'System':
                    volume = 0.0
                    dvdp = 0.0
                    d2vdp2 = 0.0
                    for phase in self.phase_names_a:
                        if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                            volume += float(root.find(".//System/Phase[@Type='" + phase + "']/Volume").text)
                            dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
                            d2vdp2 += float(root.find(".//System/Phase[@Type='" + phase + "']/D2vDp2").text)
                    value = (volume * d2vdp2 / dvdp / dvdp - 1.0)
                else:
                    if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                        volume = float(root.find(".//System/Phase[@Type='" + phase_name + "']/Volume").text)
                        dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
                        d2vdp2 = float(root.find(".//System/Phase[@Type='" + phase_name + "']/D2vDp2").text)
                        value = (volume * d2vdp2 / dvdp / dvdp - 1.0)
                    else:
                        value = 0.0
            elif property_name == 'Gamma':
                dvdp = 0.0
                dvdt = 0.0
                cp = 0.0
                t = locale.atof(root.find(".//Temperature").text) + 273.15
                if phase_name == 'System':
                    for phase in self.phase_names_a:
                        if root.find(".//System/Phase[@Type='" + phase + "']/Mass") != None:
                            dvdp += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDp").text)
                            dvdt += float(root.find(".//System/Phase[@Type='" + phase + "']/DvDt").text)
                            cp += float(root.find(".//System/Phase[@Type='" + phase + "']/HeatCapacity").text)
                    value = -dvdt / (cp * dvdp + t * dvdt * dvdt)
                else:
                    if root.find(".//System/Phase[@Type='" + phase_name + "']/Mass") != None:
                        dvdp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDp").text)
                        dvdt = float(root.find(".//System/Phase[@Type='" + phase_name + "']/DvDt").text)
                        cp = float(root.find(".//System/Phase[@Type='" + phase_name + "']/HeatCapacity").text)
                        value = -dvdt / (cp * dvdp + t * dvdt * dvdt)
                    else:
                        value = 0.0
        else:
            value = 0.0

        return value

    def get_dictionary_of_affinities(self, root, sort=True):
        """Returns an ordered dictionary of tuples of chemical affinity and phase formulae for
        undersaturated phases in the system.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_xx``
        sort : boolean
            A flag when set to sort the dictionary in order of ascending affinities

        Returns
        -------
        dict : OrderedDict
            A Python ordered dictionary. Dictionary keys are strings naming the phases
            not in the equilibrium assemblage but known to the system.  These phases are
            by definition undersaturated. Dictionary values are tuples consisting of a
            float value and a string: (affinity, formula). For a solution phase, the
            formula is the composition closest to equilibrium with the reported phase
            assemblage.  Dictionary ordering corresponds to array order in
            get_phase_names(), unless ``sorted`` is set to True; then entries are
            ordered by ascending chemical affinity.

        """
        dict = OrderedDict()
        for phase in self.phase_names_a:
            if root.find(".//Potential/Phase[@Type='" + phase + "']") != None:
                affinity = locale.atof(root.find(".//Potential/Phase[@Type='" + phase + "']/Affinity").text)
                formulae = root.find(".//Potential/Phase[@Type='" + phase + "']/Formula").text
                if affinity == 0.0:
                    affinity = 999999.0
                dict[phase] = (affinity, formulae)
        if sort == True:
            return OrderedDict(sorted(dict.items(), key=lambda t: t[1]))
        else:
            return dict

    def output_summary(self, root, printT=True, printP=True, printMass=False, printSysWt=False, printSysM=False,
                       printPhs=True, printPhsWt=False, printPhsM=False):
        """Prints information about the specified equilibrium phase assemblage.

        Parameters
        ----------
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_tp``
        printT : bool, optional, default=True
            Print the system temperature in degrees centigrade
        printP : bool, optional, default=True
            Print the system pressure in mega-Pascals
        printMass ; bool, optional, default=False
            Print the mass of the system in grams
        printSysWt : bool, optional, default=False
            Print the composition of the system in wt% oxides
        printSysM : bool, optional, default=False
            Print the composition of the system in moles of elements
        printPhs : bool, optional, default=True
            Print the phases present in the system, their masses (in grams) and their
            chemical formulas
        printPhsWt : bool, optional, default=False
            Print the composition of each phase in wt% oxides (most often used in
            conjunction with printPhs=True)
        printPhsM : bool, optional, default=False
            Print the composition of each phase in moles of endmember components
            (most often used in conjunction with printPhs=True)

        """
        if printT:
            print("{0:<10s} {1:>8.2f}".format("T (°C)", locale.atof(root.find(".//Temperature").text)))
        if printP:
            print("{0:<10s} {1:>8.2f}".format("P (MPa)", locale.atof(root.find(".//Pressure").text) * 1000.0))
        if printMass:
            print("{0:<10s} {1:>8.3f}".format("Mass (g)", float(root.find(".//Mass").text)))

        if printSysM:
            print("Bulk composition in elemental abundances (moles):")
            bcElements = list(root.findall(".//Composition/Element"))
            for element in bcElements:
                print("   {0:>2s} {1:>8.5f}".format(element.attrib['Type'], float(element.text)))

        if printSysWt:
            print("Bulk composition in oxide abundances (wt %):")
            bcOxides = list(root.findall(".//Composition/Oxide"))
            for oxide in bcOxides:
                print("   {0:>5s} {1:>8.4f}".format(oxide.attrib['Type'], float(oxide.text)))

        phases = list(root.findall(".//System/Phase"))
        for phase in phases:
            if printPhs:
                formula = phase.find("Formula").text.split(" ")
                if len(formula) == 1:
                    print("{0:<15.15s} {1:>8.4f} (g)  {2:<60s}".format(phase.attrib['Type'],
                                                                       float(phase.find("Mass").text), formula[0]))
                else:
                    # this formula consists of oxide, value pairs (liquid phase)
                    n_oxides = int(len(formula) / 2)
                    n_oxides_1 = int(n_oxides / 2)
                    n_oxides_2 = int(n_oxides - n_oxides_1)
                    if n_oxides <= 8:
                        n_oxides_1 = n_oxides
                        n_oxides_2 = 0
                    line_1 = ""
                    for i in range(n_oxides_1):
                        line_1 = line_1 + ' ' + formula[2 * i] + ' ' + formula[2 * i + 1]
                    print(
                        "{0:<15.15s} {1:>8.4f} (g) {2:<s}".format(phase.attrib['Type'], float(phase.find("Mass").text),
                                                                  line_1))
                    line_2 = ""
                    for i in range(n_oxides_2):
                        line_2 = line_2 + ' ' + formula[n_oxides_1 * 2 + 2 * i] + ' ' + formula[
                            n_oxides_1 * 2 + 2 * i + 1]
                    print("{0:<33.33s}{1:<s}".format(" ", line_2))
            if printPhsWt:
                oxides = list(phase.findall("Oxide"))
                for oxide in oxides:
                    value = float(oxide.text)
                    if (value != 0.0):
                        print("   {0:<5s} {1:>8.4f}".format(oxide.attrib['Type'], float(oxide.text)))
            if printPhsM:
                components = list(phase.findall("Component"))
                for component in components:
                    value = float(component.text)
                    if (value != 0.0):
                        print("   {0:<15s} {1:>9.6f}".format(component.attrib['Name'], float(component.text)))

    # =================================================================
    # Block of Excel workbook functions for output
    # =================================================================

    def start_excel_workbook_with_sheet_name(self, sheetName="Summary"):
        """Create an Excel workbook with one named sheet.

        Parameters
        ----------
        sheetName : string
            Sheet name in the new empty Excel workbook

        Returns
        -------
        wb : type Workbook
            A pointer to an Excel workbook

        Notes
        -----
        Part of the Excel workbook functions subpackage

        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheetName
        self.row = 0
        return wb

    def add_sheet_to_workbook_named(self, wb, sheetName):
        """Creates a new sheet in an existing Excel workbook.

        Parameters
        ----------
        wb : type Workbook
            A pointer to an existing Excel workbook
        sheetName : string
            New sheet name in the specified Excel workbook

        Returns
        -------
        ws : type Worksheet
            A pointer to an Excel worksheet in the specified workbook, ``wb``

        Notes
        -----
        Part of the Excel workbook functions subpackage

        """
        ws = wb.create_sheet(title=sheetName)
        return ws

    def write_to_cell_in_sheet(self, ws, col, row, value, format='general'):
        """Writes information into the specified row, col on the specified worksheet.

        Parameters
        ----------
        ws : type Worksheet
            A pointer to a previously created Excel worksheet
            (see add_sheet_to_workbook_named)
        col : int
            Column number to write entry to.  Numbering starts at column one.
        row : int
            Row number to write entry to. Numbering starts at row one.
        value : float
            Value to place at entry
        format : string, optional
            Format to use for value

            * 'general' is the default; no formatting applied
            * 'number' formats as '0.00'
            * 'scientific' formats as '0.00E+00'

        Notes
        -----
        Part of the Excel workbook functions subpackage

        """
        if format == 'number':
            ws.cell(column=col, row=row, value=float(value)).number_format = '0.00'
        elif format == 'scientific':
            ws.cell(column=col, row=row, value=float(value)).number_format = '0.00E+00'
        else:
            ws.cell(column=col, row=row, value=value)

    def write_excel_workbook(self, wb, fileName="junk.xlsx"):
        """Writes the specified Excel workbook to disk.

        Parameters
        ----------
        wb : type Workbook
            A pointer to an existing Excel workbook
        fileName : string, optional
            Name of the file that will contain the specified Excel notebook.
            Default file name is 'junk.xlsx'.

        Notes
        -----
        Part of the Excel workbook functions subpackage

        """
        wb.save(filename=fileName)

    def update_excel_workbook(self, wb, root):
        """Writes the specified equilibrium system state to the specified Excel workbook.

        Parameters
        ----------
        wb : type Workbook
            A pointer to an existing Excel workbook
        root : type xml.etree.ElementTree
            An xml document tree returned by the function ``equilibrate_tp``

        Notes
        -----
        The workbook is structured with a Summary worksheet and one worksheet for each
        equilibrium phase.  The evolution of the system is recorded in successive rows.
        Row integrity is maintained across all sheets. This function may be called
        repeatedly using different ``root`` objects. Part of the Excel workbook functions
        subpackage.

        """
        t = locale.atof(root.find(".//Temperature").text)
        p = locale.atof(root.find(".//Pressure").text) * 1000.0
        bcElements = list(root.findall(".//Composition/Element"))
        bcOxides = list(root.findall(".//Composition/Oxide"))

        wsSummary = wb["Summary"]  # .get_sheet_by_name("Summary") upgraded
        if (self.row == 0):
            col = 1
            self.row = 1
            self.write_to_cell_in_sheet(wsSummary, col, self.row, "T °C")
            col += 1
            self.write_to_cell_in_sheet(wsSummary, col, self.row, "P MPa")
            col += 1
            self.write_to_cell_in_sheet(wsSummary, col, self.row, "Mass g")
            col += 1
            for element in bcElements:
                self.write_to_cell_in_sheet(wsSummary, col, self.row, element.attrib['Type'])
                col += 1
            for oxide in bcOxides:
                self.write_to_cell_in_sheet(wsSummary, col, self.row, oxide.attrib['Type'])
                col += 1

        self.row += 1
        col = 1
        self.write_to_cell_in_sheet(wsSummary, col, self.row, t, format='number')
        col += 1
        self.write_to_cell_in_sheet(wsSummary, col, self.row, p, format='number')
        col += 1
        self.write_to_cell_in_sheet(wsSummary, col, self.row, root.find(".//Mass").text, format='number')
        col += 1
        for element in bcElements:
            self.write_to_cell_in_sheet(wsSummary, col, self.row, element.text, format='scientific')
            col += 1
        for oxide in bcOxides:
            self.write_to_cell_in_sheet(wsSummary, col, self.row, oxide.text, format='number')
            col += 1

        phases = list(root.findall(".//System/Phase"))
        for phase in phases:
            phaseType = phase.attrib['Type']
            oxides = list(phase.findall("Oxide"))
            components = list(phase.findall("Component"))

            try:
                wsPhase = wb[phaseType]  # .get_sheet_by_name(phaseType) upgraded
            except KeyError:
                wsPhase = wb.create_sheet(phaseType)
                col = 1
                self.write_to_cell_in_sheet(wsPhase, col, 1, "T °C")
                col += 1
                self.write_to_cell_in_sheet(wsPhase, col, 1, "P MPa")
                col += 1
                self.write_to_cell_in_sheet(wsPhase, col, 1, "Mass g")
                col += 1
                self.write_to_cell_in_sheet(wsPhase, col, 1, "Formula")
                col += 1
                for oxide in oxides:
                    self.write_to_cell_in_sheet(wsPhase, col, 1, oxide.attrib['Type'])
                    col += 1
                for component in components:
                    self.write_to_cell_in_sheet(wsPhase, col, 1, component.attrib['Name'])
                    col += 1

            col = 1
            self.write_to_cell_in_sheet(wsPhase, col, self.row, t, format='number')
            col += 1
            self.write_to_cell_in_sheet(wsPhase, col, self.row, p, format='number')
            col += 1
            self.write_to_cell_in_sheet(wsPhase, col, self.row, phase.find("Mass").text, format='number')
            col += 1
            self.write_to_cell_in_sheet(wsPhase, col, self.row, phase.find("Formula").text)
            col += 1
            for oxide in oxides:
                self.write_to_cell_in_sheet(wsPhase, col, self.row, oxide.text, format='number')
                col += 1
            for component in components:
                self.write_to_cell_in_sheet(wsPhase, col, self.row, component.text, format='scientific')
                col += 1

class PhaseLibrary:
    _available_phases: List[Type[phs.Phase]]
    _available_phase_names: List[str]
    _available_phase_abbrevs: List[str]

    def __init__(self, phase_models: Optional[List[Type[phs.Phase]]] = None):
        self._available_phases = UnorderedList()
        self._available_phase_names = UnorderedList()
        self._available_phase_abbrevs = UnorderedList()
        if phase_models is not None:
            self.add(phase_models)

    def add(self, phase_models: List[Type[phs.Phase]]):
        [self.add_phase(phase) for phase in phase_models]
        return self

    def add_phase(self, phase: phs.Phase):
        self._available_phases.append(phase)
        self._available_phase_names.append(phase.phase_name)
        self._available_phase_abbrevs.append(phase.abbrev)

    @property
    def available_phases(self) -> List[Type[phs.Phase]]:
        return self._available_phases

    @property
    def available_phase_num(self) -> int:
        return len(self.available_phases)

    @property
    def available_phase_names(self) -> List[str]:
        return self._available_phase_names

    @property
    def available_phase_abbrevs(self) -> List[str]:
        return self._available_phase_abbrevs

    @property
    def sample_assemblage(self) -> Assemblage:
        samples = []
        for phs in self.available_phases:
            samples.extend(SampleMaker.get_sample_endmembers(phs))

        return Assemblage(samples)

class GibbsMinimizer:
    class Method(Enum):
        GRIDDED = auto()
        FIXEDGRID = auto()
        NONLINEAR = auto()

        @classmethod
        def get_valid_names(cls):
            return cls.__dict__['_member_names_']

    GRIDDED = Method.GRIDDED
    FIXEDGRID = Method.FIXEDGRID
    NONLINEAR = Method.NONLINEAR

    def __new__(cls, *args, method=None, phase_library: PhaseLibrary = None,
                **kwargs) -> 'BaseGibbsMinimizer':
        if method is None:
            method = cls.GRIDDED

        if phase_library is None:
            raise cls.MissingPhaseLibraryError

        kwargs['phase_library'] = phase_library

        method_factory = GibbsMinimizer._init_method_factory()
        cls._validate_method(method, method_factory)
        return method_factory[method](*args, **kwargs)

    @classmethod
    def _validate_method(cls, method, method_factory):
        if method not in method_factory:
            raise cls.InvalidMinimizerMethod(minimizer_method=method)

    @classmethod
    def _init_method_factory(cls):
        method_factory = {
            cls.GRIDDED: GriddedSolutionMinimizer,
            cls.FIXEDGRID: FixedGridSolutionMinimizer
        }
        return method_factory

    @classmethod
    def get_valid_methods(cls):
        return ['GibbsMinimizer.' + nm
                for nm in cls.Method.get_valid_names()]

    class InvalidMinimizerMethod(Exception):
        """Raised when type is not valid"""

        def __init__(self, minimizer_method=None):
            self.minimizer_method = minimizer_method
            valid_types_str = '[' + ', '.join(GibbsMinimizer.get_valid_methods()) + ']'
            message = (f"'{minimizer_method}' is not valid. Method must be selected from "
                       + valid_types_str + ".")
            super().__init__(message)

    class MissingPhaseLibraryError(Exception):
        pass

class BaseGibbsMinimizer(abc.ABC):
    def __init__(self, phase_library: Optional[PhaseLibrary]=None,
                 aff_thresh: float=1e3, use_sample_pool=True,
                 XTOL: float=1e-6):
        self.phase_library = phase_library
        if self.phase_library is None:
            raise self.MissingPhaseLibraryError

        sample_library = self.get_sample_library(phase_library)
        self._initial_sample_library = sample_library.remove_redundant_endmembers()
        self._sample_library = SampleLibrary(self._initial_sample_library, safe_copy=False)

        self._current_assemblage = None
        self._current_sample_pool_ind = []

        self.aff_thresh=aff_thresh
        self._AFF_TOL = 1e-3
        self._XTOL = XTOL

        self._use_sample_pool = use_sample_pool

    def init_sample_pool(self, T: float, P: float, comp):

        if self.phase_library is None:
            raise self.MissingPhaseLibraryError

        filtered_sample_library = self._filter_sample_library_by_sys_comp(comp)
        self._reset_sample_library_to_initial_grid(filtered_sample_library)
        self._sample_library.update_conditions(T=T, P=P)

        updated_sample_pool = self._sample_library.samples
        self._current_assemblage = Assemblage(updated_sample_pool)

    def _filter_sample_library_by_sys_comp(self, comp):
        init_sample_comps = self._initial_sample_library.elem_comps
        all_elems = init_sample_comps.columns
        elems_present = comp.elem_comp.components
        elems_missing = [elem for elem in all_elems if elem not in elems_present]

        missing_components_mask = np.any(init_sample_comps[elems_missing] > 0, axis=1)
        ind_filtered_samples = np.where((~missing_components_mask).values)[0]
        if len(ind_filtered_samples)==0:
            raise self.CompositionNotViable()

        return self._initial_sample_library.get_subset(ind_filtered_samples)

    def _reset_sample_library_to_initial_grid(self, sample_library):
        self._sample_library = SampleLibrary(sample_library, safe_copy=False)
        self._current_sample_pool_ind = [i for i in range(len(sample_library))]


    @abc.abstractmethod
    def get_sample_library(self, phase_library: PhaseLibrary) -> Assemblage:
        pass

    @property
    def XTOL(self):
        return self._XTOL

    @property
    def initial_sample_pool(self):
        return SampleLibrary(self._initial_sample_library.samples)

    @property
    def sample_library(self):
        return SampleLibrary(self._sample_library.samples)

    @property
    def current_assemblage(self) -> Assemblage:
        # TODO (sample_pool_assemblage): promote sample library to assemblage????
        return Assemblage(self.current_sample_pool)

    @property
    def current_sample_pool(self):
        if self._use_sample_pool:
            return self._sample_library.get_subset(self._current_sample_pool_ind)
        else:
            return self._current_assemblage

    @property
    def stable_assemblage(self) -> Assemblage:
        return Assemblage(self.current_sample_pool)

    @property
    def nearly_stable_phases(self) -> List[_PhaseSample]:
        # TODO (aphabetize nearly stable phases)???
        ind = np.where(self._sample_library.metastability)[0]
        return self._sample_library.get_subset(ind)

    @property
    def nearly_stable_phase_names(self) -> UnorderedList[str]:
        return UnorderedList(np.unique(self.nearly_stable_phases.names))

    @property
    def nearly_stable_samples(self) -> List[_PhaseSample]:
        return self._sample_library.get_nearly_stable_samples(self.aff_thresh)

    @abc.abstractmethod
    def equilibrate(self, T: float, P: float, comp: chemistry.Comp) -> BaseGibbsMinimizer:
        pass

    def filter_phases_by_comp(self, comp) -> BaseGibbsMinimizer:
        sample_pool_ind = []
        for ind, phase in enumerate(self.current_sample_pool):
            if phase.comp == comp:
                sample_pool_ind.append(ind)

        if len(sample_pool_ind) == 0:
            raise self.CompositionNotViable

        self._current_sample_pool_ind = sample_pool_ind

        return self

    def find_stable_phase(self, T: float, P: float, comp):
        self.init_sample_pool(T, P, comp)
        self.filter_phases_by_comp(comp)
        # TODO(selected_phases): find new way to select only phases with exact comp
        energies = self.get_phase_energies()
        self._calc_most_stable_phase(energies)
        return self.current_sample_pool[0]

    def _calc_most_stable_phase(self, energies):
        ind_current_stable = np.argmin(energies)
        ind_stable = self._current_sample_pool_ind[ind_current_stable]
        self._current_sample_pool_ind = [ind_stable]

    def get_phase_energies(self):
        return self.current_sample_pool.energies

    def calc_stable_assemblage(self, energies: NDArray, sample_comps: NDArray,
                               system_comp: NDArray):

        assem_wts = self.get_new_assemblage_sample_wts(
            energies, sample_comps.values, system_comp.values)

        self._calc_metastable_sample_affinities(assem_wts, energies, sample_comps, system_comp)
        self._update_sample_pool(assem_wts)

    def _process_comps(self, sys_comp: chemistry.Comp) -> NDArray[Shape["N, N"], Float]:
        # TODO (refactor sample_comps): generalize process comps to any sample set
        system_comp = self._format_system_comp(sys_comp)
        sample_comps = self._format_sample_comps()
        self._drop_redundant_oxygen_component(system_comp)
        self._drop_redundant_oxygen_component(sample_comps)

        if not self._system_comp_is_viable(sample_comps, system_comp):
            raise self.CompositionNotViable()

        return sample_comps, system_comp

    def _update_sample_affinities(self, chempot_current: pd.Series):
        sample_comps = self._sample_library.elem_comps
        self._drop_redundant_oxygen_component(sample_comps)
        G_vals = self._sample_library.energies
        Gref_vals = sample_comps.dot(chempot_current)
        self._sample_library.update_affinities(G_vals-Gref_vals)

    class _SampleSubsetQueue(collections.UserList):
        def __init__(self, current_sample_pool, assem_wts, energies, sample_comps, XTOL):
            self.current_sample_pool = current_sample_pool
            self.sample_num = assem_wts.size
            self.data = [[ind] for ind, wt in enumerate(assem_wts) if wt >0]
            self.XTOL = XTOL
            self.ind_subset = None
            self.ind_excluded = None
            self._all_energies = energies
            self._all_sample_comps = sample_comps


        def next_sample_subset(self):
            ind_excluded = self.data.pop(-1)
            ind_subset = list(range(self.sample_num))
            [ind_subset.remove(ind) for ind in ind_excluded]
            self.ind_subset = ind_subset
            self.ind_excluded = ind_excluded

        def add_sample_subset(self, assem_wts_subset):
            ind_exclude_next = []
            ind_exclude_next.extend(self.ind_excluded)

            ind_metastable = np.array(self.ind_subset)[assem_wts_subset > self.XTOL]
            ind_exclude_next.extend(ind_metastable)
            self.data.append(ind_exclude_next)

        def has_no_available_samples(self):
            return len(self.ind_subset) == 0

        @property
        def energies(self):
            return self._all_energies[self.ind_subset]

        @property
        def sample_comps(self):
            return self._all_sample_comps.iloc[self.ind_subset]


        def get_trial_assem_ind(self, iassem_wts):
            return np.array(self.ind_subset)[iassem_wts > self.XTOL]


    def _calc_metastable_sample_affinities(self, assem_wts, energies, sample_comps, system_comp):
        # TODO: What about rest of sample library???

        assem_wts[assem_wts <= self.XTOL] = 0
        stable_assem_energy = np.dot(assem_wts, energies)

        self._init_metastable_props_for_current_assem(assem_wts)
        excluded_samples_queued = self._SampleSubsetQueue(
            self.current_sample_pool, assem_wts, energies, sample_comps, self.XTOL)

        while excluded_samples_queued:
            excluded_samples_queued.next_sample_subset()
            try:
                ind_trial_assem = self._find_next_metastable_trial_assem(
                    excluded_samples_queued, stable_assem_energy, system_comp)
            except (self.NoAvailableSamples,
                    self.CompositionNotViable,
                    self.AffinityExceedsThreshold):
                continue

            self._mark_pure_phases_as_metastable(ind_trial_assem)
            self._distinguish_independent_metastable_solutions(energies, ind_trial_assem, sample_comps)
            # self.current_sample_pool.update_affinities(aff)

        self._cleanup_near_zero_affinities()

    def _cleanup_near_zero_affinities(self):
        # TODO: Need test to validate Afftol ???
        for samp in self.current_sample_pool:
            if samp.aff < self._AFF_TOL:
                samp.aff = 0

    def _init_metastable_props_for_current_assem(self, assem_wts):
        for samp, assem_wt in zip(self.current_sample_pool, assem_wts):
            if assem_wt > 0:
                samp.metastable = True
                samp.aff = 0
            else:
                samp.metastable = False
                samp.aff = np.inf

    def _mark_pure_phases_as_metastable(self, ind_metastable):
        metastable_samples = self._sample_library.get_subset(ind_metastable)
        ind_pure_samps = [ind for ind, samp in
                          zip(ind_metastable, metastable_samples.samples)
                          if not samp.is_solution]

        for samp in self.current_sample_pool.get_subset(ind_pure_samps):
            samp.metastable = True

    def _find_next_metastable_trial_assem(self, excluded_samples_queued, stable_assem_energy, system_comp):
        metastable_assem_wts = self.get_new_assemblage_sample_wts(
            excluded_samples_queued.energies,
            excluded_samples_queued.sample_comps.values, system_comp.values)
        imetastable_assem_energy = np.dot(metastable_assem_wts, excluded_samples_queued.energies)
        metastable_assem_aff = imetastable_assem_energy - stable_assem_energy
        if metastable_assem_aff > self.aff_thresh:
            raise self.AffinityExceedsThreshold()

        metastable_samples = excluded_samples_queued.sample_comps.iloc[metastable_assem_wts > 0]
        print(f'Aff = {metastable_assem_aff}')
        print(metastable_samples )
        print('---')


        excluded_samples_queued.add_sample_subset(metastable_assem_wts)

        ind_trial_assem = excluded_samples_queued.get_trial_assem_ind(metastable_assem_wts)
        for ind in ind_trial_assem:
            samp = self.current_sample_pool[ind]
            print(samp.name, samp.X)
            if metastable_assem_aff < samp.aff:
                samp.aff = metastable_assem_aff

        return ind_trial_assem

    def _distinguish_independent_metastable_solutions(self, energies, ind_trial_assem,
                                                      sample_comps):

        metastable_trial_chempot, metastable_trial_samples = self._calc_metastable_chempot(
            ind_trial_assem, energies, sample_comps)
        if metastable_trial_chempot is None:
            return

        for soln in metastable_trial_samples.solution_samples:
            prior_metastable_soln_samps = self._get_prior_metastable_soln_samples(soln)

            if prior_metastable_soln_samps is None:
                soln.metastable = True
                continue
            if soln in prior_metastable_soln_samps:
                continue

            soln_mesh = self._get_soln_mesh(soln, energies, sample_comps)

            ref_energy = soln_mesh['comp'].dot(metastable_trial_chempot)
            local_energy_landscape = soln_mesh['energy'] - ref_energy
            energy_pathways = self._get_energy_pathway_to_prior_metastable_samples(
                soln, prior_metastable_soln_samps, soln_mesh['endmem_comp'],
                local_energy_landscape)

            if self._soln_lies_in_distinct_energy_basin(energy_pathways):
                soln.metastable = True

    def _soln_lies_in_distinct_energy_basin(self, energy_path, TOL = 1e-3):
        dG_step = np.diff(energy_path, axis=0)
        shared_energy_basin = np.all(dG_step <= TOL, axis=0)
        is_metastable = False
        if not np.any(shared_energy_basin):
            is_metastable = True
        return is_metastable

    def _get_prior_metastable_soln_samples(self, soln):
        ind = np.where(self.current_sample_pool.metastability)[0]
        prior_metastable_samples = self.current_sample_pool.get_subset(ind)
        if soln.name not in prior_metastable_samples.unique_phase_names:
            return None

        prior_metastable_soln_samps = prior_metastable_samples.get_subset_for_phase(
            soln.name)
        return prior_metastable_soln_samps

    def _get_soln_mesh(self, soln, energies, sample_comps):
        # TODO (refactor sample_comps): remove inputs and write convenience function
        ind_soln_samp_mesh = [ind for ind, samp in enumerate(self._sample_library)
                              if samp.name == soln.name]

        elem_comps = self._sample_library.elem_comps
        sample_comps = pd.DataFrame(
            data=elem_comps, columns=self._sample_library.elems).fillna(0)
        self._drop_redundant_oxygen_component(sample_comps)

        energies = self._sample_library.energies

        soln_mesh = {}
        endmem_comp = self._sample_library.get_subset(ind_soln_samp_mesh).sample_endmem_comps
        soln_mesh['ind'] = ind_soln_samp_mesh
        soln_mesh['comp'] = sample_comps.iloc[ind_soln_samp_mesh]
        soln_mesh['energy'] = energies[ind_soln_samp_mesh]
        soln_mesh['endmem_comp'] = pd.DataFrame(endmem_comp, columns=soln.endmember_names)
        return soln_mesh

    def _calc_metastable_chempot(self, ind_trial_assem, energies, sample_comps):
        metastable_trial_samples = self._sample_library.get_subset(ind_trial_assem)
        metastable_trial_comps = sample_comps.iloc[ind_trial_assem]
        metastable_trial_energies = energies[ind_trial_assem]

        contains_solns = [samp.is_solution for samp in metastable_trial_samples.samples]
        if not np.any(contains_solns):
            metastable_trial_chempot = None
        else:
            results = np.linalg.lstsq(metastable_trial_comps, metastable_trial_energies)
            metastable_trial_chempot = pd.Series(results[0], index=metastable_trial_comps.columns)

        return metastable_trial_chempot, metastable_trial_samples

    def _get_energy_pathway_to_prior_metastable_samples(
            self, soln, prior_metastable_soln_samps, soln_samp_mesh_comp,
            local_energy_landscape):

        col_is_constant = np.all(soln_samp_mesh_comp.diff(axis=0).iloc[1:] == 0, axis=0)
        variable_col_ind = np.where(~col_is_constant)[0]

        constant_cols = [col for col, is_const in col_is_constant.items() if is_const]

        soln_samp_mesh_comp.drop(columns=constant_cols, inplace=True)

        # TODO: Only built for 1D interpolation, needs to be expanded to ND
        component_num = len(soln_samp_mesh_comp.columns)
        X = soln_samp_mesh_comp.iloc[:, :-1].squeeze()
        y = local_energy_landscape.squeeze()
        if component_num == 2:
            interp_energy_landscape = interp.interp1d(X, y, kind='linear')
        else:
            interp_energy_landscape = interp.LinearNDInterpolator(X, y)

        frac = np.linspace(0, 1, 30)
        paths = (frac[:, np.newaxis, np.newaxis] * prior_metastable_soln_samps.sample_endmem_comps[np.newaxis, :, :]
                 + (1 - frac[:, np.newaxis, np.newaxis]) * soln.X[np.newaxis, np.newaxis, :])

        X_paths = paths[:,:,variable_col_ind[:-1]]
        energy_paths = interp_energy_landscape(X_paths).squeeze()
        return energy_paths

    def _drop_redundant_oxygen_component(self, comp: Union[pd.Series, pd.DataFrame]):
        """
        Drops oxygen component to avoid unconstrained chemical potential of oxygen

        Only relevant if every element is monovalent in sample library.
        Should be turned off if both FeO and Fe2O3 are present.
        """
        axis = comp.ndim - 1
        comp.drop('O', axis=axis, inplace=True)

    def _format_sample_comps(self):
        # TODO (refactor sample_comps):
        elem_comps = self.current_sample_pool.elem_comps
        sample_comps = pd.DataFrame(
            data=elem_comps, columns=self._sample_library.elems).fillna(0)


        return sample_comps

    def _format_system_comp(self, sys_comp):
        normalized_sys_comp = sys_comp.elem_comp.normalize().data
        elems = self._sample_library.elems
        system_comp = pd.Series(name='system', data=normalized_sys_comp,
                                index=elems).fillna(0)


        return system_comp

    def get_new_assemblage_sample_wts(self, energies: NDArray, sample_comps,
                                      sys_comp) -> NDArray:

        if len(energies) == 0:
            raise self.NoAvailableSamples()
        
        if not self._system_comp_is_viable(sample_comps, sys_comp):
            raise self.CompositionNotViable()

        output = opt.linprog(energies, A_eq=sample_comps.T, b_eq=sys_comp)
        assem_wts = output.x
        assem_wts[assem_wts < self.XTOL] = 0
        return assem_wts

    def _approx_current_chempot(self, assem_wts: NDArray, energies:NDArray,
                                sample_comps):
        stable_samp_comps = sample_comps[assem_wts > 0]
        stable_energies = energies[assem_wts > 0]
        output = np.linalg.lstsq(stable_samp_comps, stable_energies)
        chempot_current = pd.Series(output[0], index=sample_comps.columns)
        return chempot_current

    ##############################
    ### sample_pool management ###
    ##############################
    def _update_sample_pool(self, assem_wts: NDArray):


        # TODO: updating sample amounts may fail under some circumstances...
        #  Need to see if assem_wts always refers to entire sample library...
        # self._sample_library.update_amounts(assem_wts)
        self.current_sample_pool.update_amounts(assem_wts)
        samp_present_ind = np.where(assem_wts > 0)[0]

        if self._use_sample_pool:
            ind_updated_sample_pool = [self._current_sample_pool_ind[ind]
                                       for ind in samp_present_ind]

            self._current_sample_pool_ind = ind_updated_sample_pool
        else:
            self._current_assemblage.update_amounts(assem_wts)
            updated_sample_pool = [self.current_assemblage[ind] for ind in samp_present_ind]
            self._current_assemblage = Assemblage(samples=updated_sample_pool)

    def _system_comp_is_viable(self, sample_comps: NDArray, system_comp: NDArray):
        mol_samp, rnorm = opt.nnls(sample_comps.T, system_comp)

        squared_error_tol = (self.XTOL) ** 2
        if rnorm > squared_error_tol:
            return False

        return True

    class MissingPhaseLibraryError(Exception):
        pass

    class SelectedPhasesEmptyError(Exception):
        pass

    class CompositionNotViable(Exception):
        pass

    class NoAvailableSamples(Exception):
        pass

    class AffinityExceedsThreshold(Exception):
        pass


class BaseGridMinimizer(BaseGibbsMinimizer):
    def __init__(self, phase_library: Optional[PhaseLibrary] = None,
                 grid_spacing=0.1, **kwargs):
        self.grid_spacing = grid_spacing
        super().__init__(phase_library=phase_library, **kwargs)

    def get_sample_library(self, phase_library: PhaseLibrary):
        sample_library = []
        for phase in phase_library.available_phases:
            if phase.phase_type == 'pure':
                sample_grid = [SampleMaker.get_sample(phase)]
            else:
                sample_grid = SampleMaker.get_sample_grid(
                    phase, grid_spacing=self.grid_spacing)

            sample_library.extend(sample_grid)

        return SampleLibrary(sample_library)

    def _merge_unresolved_phase_samples(self):
        """Merge unresolved sample clusters into mixed-sample if stable"""

        current_assem = Assemblage(self.current_sample_pool, safe_copy=False)
        for phasenm in current_assem.multi_sample_phase_names:
            self._merge_unresolved_samples_for_phase(phasenm)

    def _merge_unresolved_samples_for_phase(self, phasenm: str):
        # TODO: fix repetition with Merge_unresolved phase samples
        #  should not recreate assem multiple times!
        current_assem = Assemblage(self.current_sample_pool, safe_copy=False)
        monophase_assem = current_assem.get_subset_for_phase(phasenm)
        assem_groups = monophase_assem.segregate_resolved_samples(self.grid_spacing)
        for sample_cluster in assem_groups:
            if len(sample_cluster) > 1:
                self._replace_with_mixed_phase_if_stable(sample_cluster)

    def _replace_with_mixed_phase_if_stable(self, sample_cluster: MonophaseAssemblage):
        # TODO: update this method to work directly on sample_library by adding
        # TODO:   samples to the end of the list and adjusting amounts

        mixed_assem = sample_cluster.get_mixed_subset_for_phase()

        self._extend_sample_pool(mixed_assem)

        if mixed_assem.total_energy < sample_cluster.total_energy:
            self._replace_members_of_current_sample_pool(sample_cluster, mixed_assem)

    def _replace_members_of_current_sample_pool(self, old_samples, new_samples):
        energy_drop = old_samples.total_energy - new_samples.total_energy

        # TODO separate sample removal and sample addition actions
        if not self._use_sample_pool:
            [self.current_assemblage.remove(isamp) for isamp in old_samples]
            [self.current_assemblage.append(isamp) for isamp in new_samples]
        else:
            ind_sample_cluster = [
                ind for (ind, samp) in zip(
                    self._current_sample_pool_ind, self.current_sample_pool)
                if samp in old_samples]

            for samp in self.current_sample_pool:
                if samp in old_samples:
                    samp.metastable = False

            for ind in ind_sample_cluster:
                samp = self._sample_library[ind]
                samp.amount = 0
                samp.aff = energy_drop

            # update sample amounts to zero
            [self._current_sample_pool_ind.remove(ind)
             for ind in ind_sample_cluster]

    ##############################
    ### sample_pool management ###
    ##############################
    def _extend_sample_pool(self, assem):
        current_samp_lib_size = len(self._sample_library)
        self._sample_library.extend(assem)
        if self._use_sample_pool:
            ind_mixed_samples = [i + current_samp_lib_size
                                 for i in range(len(assem))]
            self._current_sample_pool_ind.extend(ind_mixed_samples)




class FixedGridSolutionMinimizer(BaseGridMinimizer):
    def __init__(self, phase_library: Optional[PhaseLibrary] = None,
                 grid_spacing=0.1, **kwargs):
        super().__init__(phase_library=phase_library,
                         grid_spacing=grid_spacing, **kwargs)

    def equilibrate(self, T: float, P: float, comp: chemistry.Comp):
        self.DGTOL = 1e-1

        self.init_sample_pool(T, P, comp)
        sample_comps, system_comp = self._process_comps(comp)

        energies = self.get_phase_energies()

        self.calc_stable_assemblage(energies, sample_comps, system_comp)
        self._merge_unresolved_phase_samples()

        return self


class GriddedSolutionMinimizer(BaseGridMinimizer):
    """
    Equilibrates system by gridding solution phases including local refinement.

    Adopts an initial compositional grid-spacing with is further improved by
    successive local refinement for stable phases.
    """

    def __init__(self, phase_library: Optional[PhaseLibrary] = None,
                 grid_spacing=0.1, **kwargs):
        super().__init__(phase_library=phase_library,
                         grid_spacing=grid_spacing, **kwargs)
        self.grid_spacing_refined = self.grid_spacing


    def init_sample_pool(self, T: float, P: float, comp):
        super().init_sample_pool(T, P, comp)
        self.grid_spacing_refined = self.grid_spacing

    def equilibrate(self, T: float, P: float, comp: chemistry.Comp):
        self.DGTOL = 1e-1

        self.init_sample_pool(T, P, comp)
        sample_comps, system_comp = self._process_comps(comp)

        energies = self.get_phase_energies()

        self.calc_stable_assemblage(energies, sample_comps, system_comp)

        self._equilibrate_locally_refined_solutions(T, P, comp)

        self._merge_unresolved_phase_samples()
        return self

    def _equilibrate_locally_refined_solutions(self, T: float, P: float,
                                               comp: chemistry.Comp):
        self._refine_local_solution_grids(T, P, comp)
        energies = self.get_phase_energies()
        sample_comps, system_comp = self._process_comps(comp)
        self.calc_stable_assemblage(energies, sample_comps, system_comp)

    def _refine_local_solution_grids(self, T:float, P:float, comp:chemistry.Comp):

        stable_solution_phase_names = self.stable_assemblage.unique_solution_phase_names
        stable_solutions = self.get_all_solution_samples(
            stable_solution_phase_names, self.stable_assemblage)
        solution_grids = self.get_all_solution_samples(
            stable_solution_phase_names, self._sample_library)

        for phase_name in solution_grids:
            soln_grid = solution_grids[phase_name]
            stable_soln_samples = stable_solutions[phase_name]
            self._refine_local_soln_grid(T, P, stable_soln_samples, soln_grid)

        self.grid_spacing_refined /=2


    def _refine_local_soln_grid(self, T: float, P: float,
                                stable_soln_samples: MonophaseAssemblage,
                                soln_grid: MonophaseAssemblage):

        samp = soln_grid.samples[0]
        phase_model = samp.phase_model

        X0_samples = stable_soln_samples.sample_endmem_comps

        Xlocal0, Xlocal_refined = SampleMesh.refine_mesh_for_multiple_samples(
            X0_samples, self.grid_spacing_refined)

        local_samples = SampleMaker._get_fixed_sample_set_from_X_grid(
            phase_model, Xlocal0, T, P, amount=0)
        refined_samples = SampleMaker._get_fixed_sample_set_from_X_grid(
            phase_model, Xlocal_refined, T, P, amount=0)

        self._extend_sample_pool(refined_samples)
        self._extend_sample_pool(local_samples)

    def get_all_solution_samples(self, phase_names, grid_assem):
        solution_grids = {}
        phase_names = np.unique(phase_names)
        for name in phase_names:
            sub_assem = MonophaseAssemblage(
                [samp for samp in grid_assem.solution_samples if samp.name == name])
            if sub_assem is not None:
                solution_grids[name] = sub_assem

        return solution_grids

class SystemSummary:
    """Summary report of state of system for printing to screen.

    Provides information about stable assemblage and nearly stable phases.
    Includes:
        * energetically favorable composition for solution phases
        * amount of each stable phase in moles
        * affinity of nearly stable phases in Joules
    Note: the affinity reflects the size of the energetic barrier to stability

    """
    def __init__(self, sys: System):
        self._XTOL = sys.XTOL
        # TODO: Need affinity tolerance for nearly stable phases vs stable ones...
        self.report = self.init_report(sys)

    def init_report(self, sys:System) -> str:
        report = []
        fmt_T = '6.2f'
        fmt_P = '5.1f'
        report.append(f'T = {sys.T:{fmt_T}} K, P = {sys.P / units.GPA:{fmt_P}} GPa')

        report_stable = []
        report_nearly_stable = []

        samples = sys.nearly_stable_phases

        for samp in samples.pure_samples:
            phs_report = []
            phs_report.append(self.get_phase_header(samp))
            if self._sample_is_stable(samp):
                report_stable.extend(phs_report)
            else:
                report_nearly_stable.extend(phs_report)

        for samp in samples.solution_samples:
            phs_report = []
            phs_report.append(self.get_phase_header(samp))
            [phs_report.append(' '+line) for line in samp.summary.mol_comp]
            if self._sample_is_stable(samp):
                report_stable.extend(phs_report)
            else:
                report_nearly_stable.extend(phs_report)

        report.extend(report_stable)
        report.extend(report_nearly_stable)

        return report

    def get_phase_header(self, samp: _PhaseSample) -> str:
        phase_name = samp.summary.title
        if self._sample_is_stable(samp):
            phase_header = f'{phase_name:16} amt:{samp.amount:11.6f} mol'
        else:
            phase_header = f'... {phase_name:16} affn:{samp.aff:11.6f} J'
        return phase_header

    def _sample_is_stable(self, samp):
        return samp.amount > self._XTOL

    def __str__(self) -> str:
        return '\n'.join(self.report)

class System:
    """
    Represents an equilibrium thermodynamic system

    Attributes
    ----------
    T: float
    P: float
    comp: chemistry.OxideMolComp
    phase_library: PhaseLibrary
    equil_method: GibbsMinimizer.Method
    
    """
    gibbs_minimizer: BaseGibbsMinimizer
    equil_method: GibbsMinimizer.Method
    resolution: Optional[float]

    def __init__(self, T: float = 1000, P: float = 1,
                 comp: np.ndarray = None,
                 phase_library: PhaseLibrary = PhaseLibrary(),
                 equil_method: GibbsMinimizer.Method = GibbsMinimizer.GRIDDED,
                 affinity_thresh: float = 100,
                 options: Dict[str, float] = None):
        self.T = T
        self.P = P
        if comp is None:
            self.comp = chemistry.OxideMolComp()
        elif type(comp) is dict:
            self.comp = chemistry.OxideMolComp(comp)
        else:
            self.comp = comp

        if options is None:
            options = {}

        options['aff_thresh'] = affinity_thresh
        self.phase_library = phase_library

        self.equil_method = equil_method
        self._affinity_thresh = affinity_thresh
        self.gibbs_minimizer = GibbsMinimizer(method=equil_method,
                                              phase_library=phase_library,
                                              **options)

        if equil_method is GibbsMinimizer.GRIDDED:
            self.resolution = self.gibbs_minimizer.grid_spacing
        else:
            self.resolution = None

        if not self.sample_library_is_empty() and not self.comp_is_empty():
            self.update()

    def sample_library_is_empty(self):
        return len(self.gibbs_minimizer._sample_library) == 0

    def comp_is_empty(self):
        return self.comp == chemistry.OxideMolComp()

    def __eq__(self, other):
        T_match = self.T == other.T
        P_match = self.P == other.P
        comp_match = (self.comp == other.comp)
        if type(comp_match) is np.ndarray:
            comp_match = comp_match.all()
        return T_match & P_match & comp_match

    @property
    def affinity_thresh(self) -> float:
        return self._affinity_thresh

    @property
    def XTOL(self) -> float:
        return self.gibbs_minimizer.XTOL

    @property
    def available_phases(self) -> List[str]:
        return self.phase_library._available_phase_abbrevs

    @property
    def stable_phase_names(self) -> List[str]:
        return [samp.name for samp in self.stable_assemblage]

    @property
    def nearly_stable_phase_names(self) -> List[str]:
        return list(np.unique([samp.name for samp in self.nearly_stable_phases]))

    @property
    def nearly_stable_phases(self) -> List[_PhaseSample]:
        """Get stable and metastable phases within energy tolerance"""
        return self.gibbs_minimizer.nearly_stable_phases

    @property
    def nearly_stable_samples(self) -> List[_PhaseSample]:
        """Get complete list of every gridded samples within energy tolerance"""
        return self.gibbs_minimizer.nearly_stable_samples

    @property
    def stable_assemblage(self) -> Assemblage:
        return self.gibbs_minimizer.stable_assemblage

    @property
    def spans_miscibility_gap(self):
        """reflects if any phases have multiple samples separated by a miscibility gap"""
        assem = self.gibbs_minimizer.stable_assemblage
        sample_phase_names = [samp.name for samp in assem]
        return self._has_multiple_samples_per_phase(sample_phase_names)

    def _has_multiple_samples_per_phase(self, sample_phase_names):
        unique_phase_names = set(sample_phase_names)
        sample_num = len(sample_phase_names)
        uniq_phase_num = len(unique_phase_names)
        has_multiple_samples_per_phase = sample_num > uniq_phase_num
        return has_multiple_samples_per_phase

    def resolves_any_exsolved_samples(self):
        """Tells if exsolved samples are resolved w/in limits of minimization method """
        if not self.spans_miscibility_gap:
            return True

        if self.resolution is None:
            return True

        assem = self.gibbs_minimizer.stable_assemblage
        immiscible_phases = assem.multi_sample_phase_names

        for phase_name in immiscible_phases:
            monophase_assem = assem.get_subset_for_phase(phase_name)
            if not monophase_assem.all_samples_resolved(self.resolution):
                return False

        return True

    def update(self, T: float=None, P: float=None, comp=None):
        if T is not None:
            self.T = T

        if P is not None:
            self.P = P

        if comp is not None:
            self.comp = comp

        self.gibbs_minimizer.equilibrate(self.T, self.P, self.comp)
        return self

    @property
    def summary(self) -> SystemSummary:
        return SystemSummary(self)
