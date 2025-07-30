"""
Export utilities for data export in various formats.
"""

import csv
import io
import json
from datetime import datetime, date
from typing import List, Dict, Any, Optional

import logging

logger = logging.getLogger(__name__)


def export_to_csv(
    data: List[Dict[str, Any]],
    filename: Optional[str] = None,
    encoding: str = 'utf-8-sig'
) -> bytes:
    """
    Export data to CSV format.
    
    Args:
        data: List of dictionaries to export
        filename: Optional filename (not used in bytes output)
        encoding: Character encoding (utf-8-sig includes BOM for Excel)
        
    Returns:
        CSV data as bytes
    """
    if not data:
        return b""
    
    # Get all unique keys from all records
    all_keys = set()
    for record in data:
        all_keys.update(record.keys())
    
    # Sort keys for consistent output
    fieldnames = sorted(all_keys)
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.DictWriter(
        output, 
        fieldnames=fieldnames, 
        extrasaction='ignore'
    )
    
    writer.writeheader()
    
    # Process each record
    for record in data:
        # Convert special types to strings
        processed_record = {}
        for key, value in record.items():
            if isinstance(value, (datetime, date)):
                processed_record[key] = value.strftime('%d/%m/%Y %H:%M' if isinstance(value, datetime) else '%d/%m/%Y')
            elif isinstance(value, (list, dict)):
                processed_record[key] = json.dumps(value, ensure_ascii=False)
            elif value is None:
                processed_record[key] = ''
            else:
                processed_record[key] = str(value)
        
        writer.writerow(processed_record)
    
    # Convert to bytes
    csv_data = output.getvalue().encode(encoding)
    
    logger.info(f"Exported {len(data)} records to CSV ({len(csv_data)} bytes)")
    
    return csv_data


def export_to_excel(
    data: List[Dict[str, Any]],
    sheet_name: str = "Dados",
    filename: Optional[str] = None
) -> bytes:
    """
    Export data to Excel format.
    
    Note: This requires openpyxl or xlsxwriter to be installed.
    For now, returns CSV as Excel can open CSV files.
    
    Args:
        data: List of dictionaries to export
        sheet_name: Name of the Excel sheet
        filename: Optional filename
        
    Returns:
        Excel data as bytes
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            # Return empty workbook
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        
        # Get headers
        headers = list(data[0].keys())
        
        # Write headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, record in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = record.get(header)
                
                # Handle special types
                if isinstance(value, datetime):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = 'DD/MM/YYYY HH:MM'
                elif isinstance(value, date):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.number_format = 'DD/MM/YYYY'
                elif isinstance(value, (int, float)):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                elif isinstance(value, (list, dict)):
                    ws.cell(row=row_idx, column=col_idx, value=json.dumps(value, ensure_ascii=False))
                else:
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_data = output.getvalue()
        
        logger.info(f"Exported {len(data)} records to Excel ({len(excel_data)} bytes)")
        
        return excel_data
        
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV export")
        # Fallback to CSV if openpyxl is not available
        return export_to_csv(data, filename)


def export_to_pdf(
    content: str,
    title: str = "Relatório",
    filename: Optional[str] = None,
    orientation: str = 'portrait'
) -> bytes:
    """
    Export content to PDF format.
    
    Note: This requires reportlab to be installed.
    For now, returns the content as text.
    
    Args:
        content: Text content to export
        title: PDF title
        filename: Optional filename
        orientation: Page orientation (portrait/landscape)
        
    Returns:
        PDF data as bytes
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
        
        # Create PDF in memory
        output = io.BytesIO()
        
        # Set page size
        pagesize = landscape(letter) if orientation == 'landscape' else letter
        
        # Create document
        doc = SimpleDocTemplate(
            output,
            pagesize=pagesize,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Container for elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.5 * inch))
        
        # Add date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(
            Paragraph(
                f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                date_style
            )
        )
        elements.append(Spacer(1, 0.5 * inch))
        
        # Process content
        content_style = ParagraphStyle(
            'ContentStyle',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_JUSTIFY,
            spaceAfter=12
        )
        
        # Split content into paragraphs
        for paragraph in content.split('\n\n'):
            if paragraph.strip():
                elements.append(Paragraph(paragraph, content_style))
                elements.append(Spacer(1, 0.2 * inch))
        
        # Build PDF
        doc.build(elements)
        
        pdf_data = output.getvalue()
        logger.info(f"Exported content to PDF ({len(pdf_data)} bytes)")
        
        return pdf_data
        
    except ImportError:
        logger.warning("reportlab not installed, returning text content")
        # Fallback to text if reportlab is not available
        return content.encode('utf-8')


def export_patients_csv(patients: List[Any]) -> bytes:
    """
    Export patients list to CSV.
    
    Args:
        patients: List of patient objects
        
    Returns:
        CSV data as bytes
    """
    data = []
    for patient in patients:
        data.append({
            'Nome': patient.full_name,
            'CPF': patient.cpf or '',
            'Data de Nascimento': patient.bdate,
            'Idade': patient.age,
            'Telefone': patient.phone or '',
            'Email': patient.email or '',
            'Endereço': patient.address or '',
            'Ativo': 'Sim' if getattr(patient, 'active', True) else 'Não'
        })
    
    return export_to_csv(data)


def export_financial_summary_excel(summary: Dict[str, Any]) -> bytes:
    """
    Export financial summary to Excel.
    
    Args:
        summary: Financial summary dictionary
        
    Returns:
        Excel data as bytes
    """
    # Prepare data for multiple sheets
    sheets_data = {}
    
    # Summary sheet
    summary_data = [{
        'Período': f"{summary['period']['start']} a {summary['period']['end']}",
        'Receita Total': summary['revenue']['total'],
        'Receita Paga': summary['revenue']['paid'],
        'Receita Pendente': summary['revenue']['pending'],
        'Despesa Total': summary['expenses']['total'],
        'Despesa Paga': summary['expenses']['paid'],
        'Despesa Pendente': summary['expenses']['pending'],
        'Balanço Bruto': summary['balance']['gross'],
        'Balanço Líquido': summary['balance']['net']
    }]
    
    # Revenue by service
    revenue_data = []
    for service, data in summary['revenue']['by_service'].items():
        revenue_data.append({
            'Serviço': service,
            'Quantidade': data['count'],
            'Total': data['total'],
            'Pago': data['paid'],
            'Pendente': data['total'] - data['paid']
        })
    
    # Expenses by category
    expense_data = []
    for category, data in summary['expenses']['by_category'].items():
        expense_data.append({
            'Categoria': category,
            'Quantidade': data['count'],
            'Total': data['total'],
            'Pago': data['paid'],
            'Pendente': data['total'] - data['paid']
        })
    
    # For now, combine all data in one sheet
    all_data = []
    all_data.append({'Tipo': '=== RESUMO ==='})
    all_data.extend(summary_data)
    all_data.append({})  # Empty row
    all_data.append({'Tipo': '=== RECEITAS POR SERVIÇO ==='})
    all_data.extend(revenue_data)
    all_data.append({})  # Empty row
    all_data.append({'Tipo': '=== DESPESAS POR CATEGORIA ==='})
    all_data.extend(expense_data)
    
    return export_to_excel(all_data, sheet_name="Resumo Financeiro")