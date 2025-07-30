#!/usr/bin/env python3

from pathlib import Path

from openpyxl import load_workbook

from clanguru.object_analyzer import ObjectData, ObjectsDataExcelReportGenerator, Symbol, SymbolLinkage


def test_excel_report_generator(tmp_path: Path) -> None:
    """Create a sample Excel report with mock object data."""
    # Create sample object data
    obj1 = ObjectData(Path("main.o"))
    obj1.symbols = [
        Symbol("main", SymbolLinkage.LOCAL),
        Symbol("printf", SymbolLinkage.EXTERN),
        Symbol("strlen", SymbolLinkage.EXTERN),
        Symbol("my_function", SymbolLinkage.LOCAL),
    ]

    obj2 = ObjectData(Path("utils.o"))
    obj2.symbols = [
        Symbol("my_function", SymbolLinkage.EXTERN),
        Symbol("utility_func", SymbolLinkage.LOCAL),
        Symbol("malloc", SymbolLinkage.EXTERN),
        Symbol("free", SymbolLinkage.EXTERN),
    ]

    obj3 = ObjectData(Path("math.o"))
    obj3.symbols = [
        Symbol("add", SymbolLinkage.LOCAL),
        Symbol("subtract", SymbolLinkage.LOCAL),
        Symbol("multiply", SymbolLinkage.LOCAL),
        Symbol("sin", SymbolLinkage.EXTERN),
        Symbol("cos", SymbolLinkage.EXTERN),
    ]

    # Create the report generator
    generator = ObjectsDataExcelReportGenerator([obj1, obj2, obj3])

    # Generate the Excel report
    output_file = tmp_path / "objects_report.xlsx"
    generator.generate_report(output_file)

    assert output_file.exists(), f"Report file {output_file} was not created."

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == ["Objects", "Dependency Matrix"]
