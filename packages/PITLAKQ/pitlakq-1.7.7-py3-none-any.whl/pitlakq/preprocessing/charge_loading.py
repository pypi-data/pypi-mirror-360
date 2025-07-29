"""Charge balance the input for loadings from pit walls.

The specie inputs for the loading are typically not charge balanced.
This script helps to charge balance them.
"""

from __future__ import print_function

import os
import shutil
import subprocess

from openpyxl.reader.excel import load_workbook
import xlsxwriter

from pitlakq.commontools.input.xlsx_reader import read_xslx_table
from pitlakq.commontools.input.resources import Resources
from pitlakq.metamodel.configuration.getconfig import get_config


def read_conc(file_name='charge_loading.xlsx'):
    """Read data from xlsx input file.
    """
    workbook = load_workbook(filename=file_name)
    worksheet = workbook.get_sheet_by_name('Uncharged')
    return read_xslx_table(worksheet, xlsx_file_name=file_name)


def write_conc(data, file_name='charge_loading_out.xlsx'):
    """Write the charge balanced values back into an output file.
    """
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet('Charged')
    bold = workbook.add_format()
    bold.set_bold()
    for index, head in enumerate(data.keys()):
        worksheet.write(0, index, head, bold)
    for line_number, line_index in enumerate(range(len(data['zone'])), 1):
        for index, head in enumerate(data.keys()):
            style = None
            if index == 0:
                style = bold
            worksheet.write(line_number, index, data[head][line_index], style)
    workbook.close()


class PhreeqcCalc(object):
    """PHREEQC calculation.
    """
    def __init__(self, project_name, equi_species={}):
        self.equi_species = equi_species
        self.config = get_config(project_name)
        self.config.phreeqc_original_database
        self.base_dir = os.path.join(self.config.ram_path, 'charge_loading')
        if not os.path.exists(self.base_dir):
            os.mkdir(self.base_dir)

        self.input_file_name = os.path.join(self.base_dir, 'charge_input.txt')
        self.punch_file_name = os.path.join(self.base_dir, 'charge_punch.txt')
        self.output_file_name = os.path.join(self.base_dir,
                                             'charge_output.txt')
        self.phreeqc_exe = os.path.join(self.base_dir, 'phreeqc.exe')
        self.phreeqc_db = os.path.join(self.base_dir, 'phreeqc.dat')
        shutil.copy(self.config.phreeqc_exe_original, self.phreeqc_exe)
        shutil.copy(self.config.phreeqc_original_database, self.phreeqc_db)
        resources = Resources(self.config)
        self.name_map = dict((entry['name'], entry['phreeqc_name'])
                              for entry in resources.phreeqc_species)
        self.name_map_reverse = dict((value, key) for key, value in
                                     self.name_map.items())
        self.molar_weight = dict((entry['name'], entry['molar_weight'])
                              for entry in resources.phreeqc_species)

    def write_input(self, data):
        """Write PHREEQC input file.
        """
        valid_names = set(data.keys()) - set(['zone', 'water_volume [L]',
                                              'charge_by', 'target_ph',
                                              'time [s]'])
        valid_names = sorted(valid_names)
        self.total_names = [
            ''.join(char for char in self.name_map[name] if char not in('+-'))
                            for name in valid_names]
        self.valid_names = valid_names
        fobj = open(self.input_file_name, 'w')
        fobj.write('Title charge balancing of loading\n')
        fobj.write('\nSELECTED_OUTPUT\n-high_precision\n')
        fobj.write('-file %s\n' % self.punch_file_name)
        fobj.write('-high_precision\n')
        if self.equi_species:
            fobj.write('-equilibrium_phases ')
            for name in self.equi_species:
                fobj.write(name)
            fobj.write('\n')
        species = ' '.join(self.total_names)
        fobj.write('-totals %s\n-percent_error true\n\n' % species)
        count = range(len(data['zone']))
        for counter in count:
            fobj.write('Solution %03d\n' % counter)
            fobj.write('# Zone %03s\n' % data['zone'][counter])
            fobj.write('units              mmol/l\ntemp           3\n')
            fobj.write('%20s %25.20f\n' % ('pH', data['target_ph'][counter]))
            fobj.write('%20s %25.20f\n' % ('O(0)', 100))
            for name in valid_names:
                post_fix = ''
                if name == data['charge_by'][counter]:
                    post_fix = 'charge'
                value = data[name][counter] / data['water_volume [L]'][counter]
                value *= 1e3
                value /= self.molar_weight[name]
                value *= data['time [s]'][counter]
                fobj.write('%20s %20.15f %s\n' % (self.name_map[name], value,
                                                 post_fix))
            fobj.write('END\n\n')
        if self.equi_species:
            for counter in count:
                fobj.write('EQUILIBRIUM_PHASES %03d\n' % counter)
                for name, sat_index in self.equi_species.items():
                    fobj.write('%s %10.8f 0\n' % (name, sat_index))
                fobj.write('END\n\n')
        for counter in count:
            fobj.write('REACTION %03d\n' % counter)
            fobj.write('H2O 1.0\n0.0 moles\nEND\n\n')
        for counter in count:
            fobj.write('USE SOLUTION %03d\n' % counter)
            if self.equi_species:
                fobj.write('USE EQUILIBRIUM_PHASES %03d\n' % counter)
            fobj.write('END\n\n')
        fobj.close()

    def read_punch(self, data):
        """Read the PHREEQC output.
        """
        with open(self.punch_file_name) as fobj:
            header_pos = dict((name, index) for index, name in
                              enumerate(next(fobj).split()))
            if self.equi_species:
                # Skip non-reaction output lines.
                count = range(len(data['zone']))
                for counter in count:
                    next(fobj)
            for counter, raw_line in enumerate(fobj):
                line = raw_line.split()
                for index, total_name in enumerate(self.total_names):
                    name = self.valid_names[index]
                    value = float(line[header_pos[total_name]])
                    value *= self.molar_weight[name]
                    value *= data['water_volume [L]'][counter]
                    value /= data['time [s]'][counter]
                    data[name][counter] = value
        for key in ['water_volume [L]', 'charge_by', 'target_ph', 'time [s]']:
            data.pop(key)
        return data

    def run_phreeqc(self):
        """Run PHREEQC
        """
        cmd = '%s %s %s %s' % (self.phreeqc_exe,
                               self.input_file_name,
                               self.output_file_name,
                               self.phreeqc_db)
        print(cmd)
        subprocess.call(cmd)


def main(project_name, equi_species={}, file_name='charge_loading.xlsx'):
    """Run the charge balance.
    """
    data = (read_conc(file_name))
    phreeqc_calc = PhreeqcCalc(project_name, equi_species)
    phreeqc_calc.write_input(data)
    phreeqc_calc.run_phreeqc()
    phreeqc_calc.read_punch(data)
    write_conc(data)
