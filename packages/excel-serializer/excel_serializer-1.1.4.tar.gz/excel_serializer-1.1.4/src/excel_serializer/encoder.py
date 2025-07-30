__all__ = ["ExcelEncoder"]

import random
import string
import warnings
from typing import Callable, Iterable, Optional, Sequence, Sized

from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import KNOWN_TYPES, Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .defaults import LINK_FONT, ROOT_SHEET_NAME, SEPARATOR, TABLE_STYLE_INFO


class ExcelEncoder:
    """
    An Excel encoder that writes a Python value to an Excel workbook.
    Methods are overridable to support custom types or behaviors.
    """

    def __init__(
        self,
        workbook: Workbook,
        skipkeys: bool = False,
        check_circular: bool = True,
        sort_keys: bool = False,
        default: Optional[Callable[[object], object]] = None,
        separator: str = SEPARATOR,
        root_sheet_name: str = ROOT_SHEET_NAME,
        link_font: Font = LINK_FONT,
        table_style_info: TableStyleInfo = TABLE_STYLE_INFO,
    ) -> None:
        """
        :param workbook: an openpyxl workbook object.
        :param skipkeys: If True, dict keys that are not of an openpyxl-compatible type will be skipped.
        :param check_circular: If True, circular references will raise an exception.
        :param sort_keys: If True, dict keys will be sorted.
        :param default: a function that should return a serializable version of obj or raise TypeError. The default
        simply raises TypeError.
        :param separator: Separator used in the sheet names.
        :param link_font: The openpyxl Font object to use for hyperlinks.
        :param table_style_info: The openpyxl TableStyleInfo object to use for tables.
        """
        self.workbook = workbook
        self.skipkeys = skipkeys
        self.check_circular = check_circular
        self.sort_keys = sort_keys
        self._default = lambda obj: (default or (lambda x: x))(self.default(obj))
        self.separator = separator
        self.root_sheet_name = root_sheet_name
        self.link_font = link_font
        self.table_style_info = table_style_info
        self._obj_ids = set()

    def default(self, obj: object) -> object:
        """
        Default method to convert an object to a serializable object.
        :param obj: Object to convert.
        :return: Serializable object.
        """
        return obj

    def gen_uid(self) -> str:
        """
        Generates a unique identifier for a sheet.
        :return: The unique identifier.
        """
        uid = "".join(
            random.choice(string.ascii_uppercase + string.digits) for _ in range(31)
        )
        return uid if uid not in self.workbook.sheetnames else self.gen_uid()

    def sheet_name_truncate(self, name: str, n: int) -> str:
        """
        Safely truncates a sheet name to n characters.
        :param name: The name to sanitize.
        :param n: The maximum length of the name.
        :return: The truncated name.
        """
        name = name[:n]
        last_ampersand = name.rfind("&")
        if last_ampersand != -1 and name.rfind(";", last_ampersand) == -1:
            return "&".join(name.split("&")[:-1])
        return name

    def sheet_name_encode(self, name: Iterable[str] | str) -> str:
        """
        Returns a valid sheet name from a list of names (likely a json-path).
        :param name: List of names (likely a json-path).
        :return: A valid sheet name, guaranteed to not already exist in the workbook.
        """
        if isinstance(name, str):
            name = (name,)
        name = self.sheet_name_truncate(
            self.separator.join(
                "".join(filter(lambda c: c not in "\\/?*[](), ", part)) for part in name
            )
            .replace("&", "&amp;")
            .replace(">", "&gt;")
            .replace("<", "&lt;"),
            31
        )
        if name not in self.workbook.sheetnames:
            return name
        name = self.sheet_name_truncate(name, 27)
        for i in range(2, 1000):
            if f"{name}_{i}" not in self.workbook.sheetnames:
                return f"{name}_{i}"
        return self.gen_uid()

    def write_sheet(self, origin: str, sheet: Worksheet, value: object) -> str:
        """
        Writes a value to a sheet.
        :param origin: The origin sheet name, where this value comes from. This is used to create a hyperlinks between
        sheets.
        :param sheet: The sheet to write to.
        :param value: The value to write.
        :return: The type of the value written, as found in the top left cell (the type cell).
        """
        # Check circular references problems
        if self.check_circular:
            obj_id = id(value)
            if obj_id in self._obj_ids:
                raise ValueError(f"Circular reference detected")
            self._obj_ids.add(obj_id)

        # Prepare the type cell
        if origin:
            type_cell = WriteOnlyCell(sheet, f'from {origin.split("!")[0]}')
            type_cell.hyperlink = f"#{origin}"
            type_cell.font = self.link_font
        else:
            type_cell = WriteOnlyCell(sheet, "(root)")

        if (
            not isinstance(value, dict)
            and not isinstance(value, list)
            and not isinstance(value, tuple)
        ):
            value = self._default(value)

        # Write the value in a new sheet
        if isinstance(value, dict):
            end_row, end_col, cols = self.write_dict(sheet, type_cell, value)
        elif isinstance(value, Sequence):
            first = value[0] if value else None
            if first and all(
                isinstance(e, dict) and e.keys() == first.keys() for e in value
            ):
                end_row, end_col, cols = self.write_dict_list(sheet, type_cell, value)
            else:
                end_row, end_col, cols = self.write_simple_list(sheet, type_cell, value)
        else:
            end_row, end_col, cols = self.write_custom_type(sheet, type_cell, value)

        # This does not work in write-only mode:
        # sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

        # Make the table containing the value
        if value:
            table = Table(
                displayName=sheet.title, ref=f"A2:{get_column_letter(end_col)}{end_row}"
            )
            for i, col_name in enumerate(cols):
                table.tableColumns.append(TableColumn(id=i + 1, name=col_name))
            table.autoFilter = AutoFilter(ref=table.ref)
            table.tableStyleInfo = self.table_style_info
            with warnings.catch_warnings(action="ignore"):
                sheet.add_table(table)

        return type_cell.value.split(" ")[0]

    def write_custom_type(
        self, sheet: Worksheet, type_cell: WriteOnlyCell, value: object
    ) -> tuple[int, int, Iterable[str]]:
        """
        Override this method to write additional types to a sheet. Call this super method if the type is truly unknown.
        :param sheet: The sheet to write to.
        :param type_cell: The type cell to write to (its value must be prefixed with the name of your custom type).
        :param value: The value to write.
        :return: The end row (1-based), end column (1-based), and columns' name of the sheet.
        """
        raise TypeError(f"Cannot insert {type(value)}:{value} into sheet")

    def write_dict(
        self, sheet: Worksheet, type_cell: WriteOnlyCell, dct: dict
    ) -> tuple[int, int, Iterable[str]]:
        """
        Writes a dictionary to a sheet.
        :param sheet: The sheet to write to.
        :param type_cell: The type cell to write to (its value must be prefixed with 'Dict').
        :param dct: The dictionary to write.
        :return: The end row (1-based), end column (1-based), and columns' name of the sheet.
        """
        cols = ("Key", "Value")
        type_cell.value = f"Dict {type_cell.value}"
        sheet.append((type_cell,))
        sheet.append(cols)
        items = sorted(dct.items()) if self.sort_keys else dct.items()
        for i, (key, value) in enumerate(items):
            if type(key) not in KNOWN_TYPES:
                key = self._default(key)
                if type(key) not in KNOWN_TYPES:
                    if not self.skipkeys:
                        raise TypeError(
                            f"Cannot convert Dict key {type(key)}:{key} to Excel"
                        )
                    continue
            sheet.append((key, self.encode(sheet, i + 3, 2, key, value)))
        return 2 + len(dct), 2, cols

    def write_simple_list(
        self, sheet: Worksheet, type_cell: WriteOnlyCell, lst: Sequence[object]
    ) -> tuple[int, int, Iterable[str]]:
        """
        Writes a list to a sheet.
        :param sheet: The sheet to write to.
        :param type_cell: The type cell to write to (its value must be prefixed with 'List').
        :param lst: The list to write.
        :return: The end row (1-based), end column (1-based), and columns' name of the sheet.
        """
        cols = ("Value",)
        type_cell.value = (
            f"{'List' if isinstance(lst, list) else 'Tuple'} {type_cell.value}"
        )
        sheet.append((type_cell,))
        sheet.append(cols)
        for i, e in enumerate(lst):
            sheet.append((self.encode(sheet, i + 3, 1, str(i + 1), e),))
        return 2 + len(lst), 1, cols

    def write_dict_list(
        self, sheet: Worksheet, type_cell: WriteOnlyCell, dct_lst: Sequence[dict]
    ) -> tuple[int, int, Iterable[str]]:
        """
        Writes a list of dictionaries to a sheet. All dictionaries must have the same keys.
        :param sheet: The sheet to write to.
        :param type_cell: The type cell to write to (its value must be prefixed with 'DictList').
        :param dct_lst: The list of dictionaries to write.
        :return: The end row (1-based), end column (1-based), and columns' name of the sheet.
        """
        first = dct_lst[0] if dct_lst else {}
        keys = sorted(first.keys()) if self.sort_keys else list(first.keys())
        for key in filter(lambda k: type(k) not in KNOWN_TYPES, first.keys()):
            key = self._default(key)
            if key not in KNOWN_TYPES:
                if not self.skipkeys:
                    raise TypeError(
                        f"Cannot convert DictList key {type(key)}:{key} to Excel"
                    )
                keys.remove(key)
        cols = [
            str(k) if type(k) in KNOWN_TYPES else str(self._default(k)) for k in keys
        ]
        type_cell.value = f"DictList {type_cell.value}"
        sheet.append((type_cell,))
        sheet.append(cols)
        i = 0
        for e in dct_lst:
            i += 1
            sheet.append(
                [
                    self.encode(sheet, i + 3, j + 1, f"{i}{self.separator}{k}", e[k])
                    for j, k in enumerate(keys)
                ]
            )
        return 2 + i, len(keys), cols

    def encode(
        self, sheet: Worksheet, row: int, col: int, name: str, value: object
    ) -> Optional[Cell | object]:
        """
        Encodes a value to a cell in a sheet.
        :param sheet: The sheet to write to.
        :param row: The row to write to (1-based).
        :param col: The column to write to (1-based).
        :param name: The name of the value.
        :param value: The value to write.
        :return: An openpyxl-writable value. Either it is a known type, or a Cell object.
        """
        if type(value) in KNOWN_TYPES:
            return value
        if not sheet:
            new_sheet_name = self.root_sheet_name
        else:
            curr_sheet_name = sheet.title.split(SEPARATOR)
            new_name = (
                [] if curr_sheet_name == [self.root_sheet_name] else curr_sheet_name
            )
            new_name.append(name)
            new_sheet_name = self.sheet_name_encode(new_name)
        value_type = self.write_sheet(
            f"{sheet.title}!{get_column_letter(col)}{row}" if sheet else None,
            self.workbook.create_sheet(new_sheet_name),
            value,
        )
        if not sheet:
            return None
        str_value = str(value)[:21]
        size = f"[{len(value)}]" if isinstance(value, Sized) else ""
        cell = WriteOnlyCell(
            sheet,
            f"{value_type}{size}{f': {str_value}' if len(str_value) < 20 else ''}",
        )
        cell.hyperlink = f"#{new_sheet_name}!A1"
        cell.font = self.link_font
        return cell
