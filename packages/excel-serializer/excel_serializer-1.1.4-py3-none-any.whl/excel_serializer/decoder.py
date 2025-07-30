__all__ = ["ExcelDecoder", "ExcelDecodeError"]

from typing import Iterator, Type

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .defaults import ROOT_SHEET_NAME


class ExcelDecodeError(ValueError):
    """
    Subclass of ValueError with the following additional properties:
    msg: The unformatted error message
    w_b: The Excel workbook being parsed
    w_s: The Excel sheet name being parsed
    row: The row index (1-based) of doc where parsing failed
    col: The col index (1-based) of doc where parsing failed
    """

    # Note that this exception is used from _json
    def __init__(self, msg: str, w_b: Workbook, w_s: str, row: int, col: int) -> None:
        errmsg = f"{msg}: #{w_s}!{get_column_letter(col)}{row}"
        ValueError.__init__(self, errmsg)
        self.msg = msg
        self.w_b = w_b
        self.w_s = w_s
        self.row = row
        self.col = col

    def __reduce__(self) -> tuple[Type, tuple[str, Workbook, str, int, int]]:
        return self.__class__, (self.msg, self.w_b, self.w_s, self.row, self.col)


class ExcelDecoder:
    """
    An Excel decoder that reads an Excel workbook and returns the decoded object.
    Methods are overridable to support custom types or behaviors.
    """

    def __init__(
        self, workbook: Workbook, root_sheet_name: str = ROOT_SHEET_NAME
    ) -> None:
        """
        :param workbook: an openpyxl workbook object
        :param root_sheet_name: Name of the root sheet to look at.
        """
        self.workbook = workbook
        self.root_sheet_name = root_sheet_name

    def read_sheet(self, sheet: Worksheet) -> object:
        """
        Reads a sheet and returns the decoded object.
        :param sheet: an openpyxl worksheet object
        :return: Decoded object.
        """
        rows = sheet.iter_rows()
        first_row = next(rows)
        type_cell = first_row[0] if first_row else None
        if not type_cell:
            raise ExcelDecodeError(
                "Missing type cell", self.workbook, sheet.title, 1, 1
            )
        sheet_type = type_cell.value.split(" ")[0]
        return {
            "Dict": self.read_dict,
            "List": self.read_simple_list,
            "Tuple": self.read_simple_list,
            "DictList": self.read_dict_list,
        }.get(sheet_type, self.read_custom_type)(sheet.title, rows)

    def read_custom_type(
        self, sheet_type: str, sheet_name: str, rows: Iterator
    ) -> object:
        """
        Override this method to decode additional types to a Python object. Call this super method if the type is truly
        unknown.
        :param sheet_type: The data type of the sheet, as read from the top-left cell of the sheet.
        :param sheet_name: The name of the sheet.
        :param rows: An iterator of rows in the sheet, starting after the type cell row.
        :return: Decoded object.
        """
        raise ExcelDecodeError(
            f'Unknown type "{sheet_type}"', self.workbook, sheet_name, 1, 1
        )

    def read_simple_list(self, sheet_name: str, rows: Iterator) -> list[object]:
        """
        Reads a simple list from a rows iterator and returns it.
        :param sheet_name: The name of the sheet.
        :param rows: An iterator of rows in the sheet, starting after the type cell row.
        :return: Decoded list.
        """
        headers = next(rows)
        if len(headers) != 1:
            raise ExcelDecodeError(
                f"Invalid list headers. Expected 1, found {len(headers)}",
                self.workbook,
                sheet_name,
                2,
                len(headers) + 1,
            )
        if headers[0].value != "Value":
            raise ExcelDecodeError(
                f'Invalid list headers. Expected "Value", found "{headers[0].value}"',
                self.workbook,
                sheet_name,
                2,
                1,
            )
        return [self.read_value(row[0]) for row in rows]

    def read_dict_list(
        self, sheet_name: str, rows: Iterator
    ) -> list[dict[str, object]]:
        """
        Reads a list of dictionaries from a rows iterator and returns it.
        e.g. [{'Key1': 'Value1', 'Key2': 'Value2'}, {'Key1': 'Value3', 'Key2': 'Value4'}, ...]
        :param sheet_name: The name of the sheet.
        :param rows: An iterator of rows in the sheet, starting after the type cell row.
        :return: Decoded list of dictionaries.
        """
        headers = next(rows)
        keys = [header.value for header in headers]
        return [
            {key: self.read_value(row[j]) for j, key in enumerate(keys)} for row in rows
        ]

    def read_dict(self, sheet_name: str, rows: Iterator) -> dict[str, object]:
        """
        Reads a dictionary from a rows iterator and returns it.
        :param sheet_name: The name of the sheet.
        :param rows: An iterator of rows in the sheet, starting after the type cell row.
        :return: Decoded dictionary.
        """
        headers = next(rows)
        if len(headers) != 2:
            raise ExcelDecodeError(
                f"Invalid dict headers. Expected 2, found {len(headers)}",
                self.workbook,
                sheet_name,
                2,
                len(headers) + 1,
            )
        if headers[0].value != "Key":
            raise ExcelDecodeError(
                f'Invalid dict headers. Expected "Key", found "{headers[0].value}"',
                self.workbook,
                sheet_name,
                2,
                1,
            )
        if headers[1].value != "Value":
            raise ExcelDecodeError(
                f'Invalid dict headers. Expected "Value", found "{headers[1].value}"',
                self.workbook,
                sheet_name,
                2,
                2,
            )
        return {row[0].value: self.read_value(row[1]) for row in rows}

    def read_value(self, cell: Cell) -> object:
        """
        Reads a cell and returns its value.
        :param cell: an openpyxl cell object
        :return: Decoded value.
        """
        if cell.hyperlink:
            return self.read_sheet(
                self.workbook[
                    "!".join(
                        (cell.hyperlink.target or "").split("!")[:-1]
                    ).removeprefix("#")
                ]
            )
        return cell.value

    def decode(self) -> object:
        """
        Reads the root sheet and returns the decoded object.
        :return: Decoded object.
        """
        if self.root_sheet_name not in self.workbook.sheetnames:
            raise ValueError(
                f"Workbook does not contain a {self.root_sheet_name} sheet"
            )
        return self.read_sheet(self.workbook[self.root_sheet_name])
