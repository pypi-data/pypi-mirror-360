"""
Export utilities for different file formats
"""

import logging
from typing import List, Optional
from datetime import datetime

from ..core.data_models import FinancialData
from .compatibility import HAS_OPENPYXL

if HAS_OPENPYXL:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel export functionality using openpyxl"""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export functionality")
    
    def export_data(self, data_list: List[FinancialData], filename: Optional[str] = None) -> str:
        """
        Export financial data to Excel file
        
        Args:
            data_list: List of FinancialData objects
            filename: Output filename (optional)
            
        Returns:
            Path to created Excel file
        """
        if not filename:
            filename = f"financial_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Data"
        
        if not data_list:
            # Create empty workbook with headers
            ws.cell(row=1, column=1, value="No data available")
            wb.save(filename)
            return filename
        
        # Get headers from first item
        first_item = data_list[0]
        headers = list(first_item.to_dict().keys())
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=self._format_header(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, data in enumerate(data_list, 2):
            data_dict = data.to_dict()
            for col, header in enumerate(headers, 1):
                value = data_dict[header]
                # Format certain fields
                if header == "timestamp" and value:
                    try:
                        # Format timestamp to be more readable
                        dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        value = dt.strftime('%Y-%m-%d %H:%M')
                    except:
                        pass
                
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Add a summary sheet if we have enough data
        if len(data_list) > 1:
            self._add_summary_sheet(wb, data_list)
        
        wb.save(filename)
        logger.info(f"Exported {len(data_list)} records to Excel file: {filename}")
        return filename
    
    def _format_header(self, header: str) -> str:
        """Format header text to be more readable"""
        # Convert snake_case to Title Case
        formatted = header.replace('_', ' ').title()
        
        # Special cases
        replacements = {
            'Pe Ratio': 'P/E Ratio',
            'Ps Ratio': 'P/S Ratio', 
            'Pb Ratio': 'P/B Ratio',
            'Roa': 'ROA',
            'Roe': 'ROE',
            'Roi': 'ROI',
            'Eps': 'EPS',
            'Ttm': 'TTM',
            '5Y': '5-Year'
        }
        
        for old, new in replacements.items():
            formatted = formatted.replace(old, new)
        
        return formatted
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set minimum width of 12, maximum of 50
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_sheet(self, wb: Workbook, data_list: List[FinancialData]):
        """Add a summary sheet with basic statistics"""
        ws = wb.create_sheet(title="Summary")
        
        # Title
        ws.cell(row=1, column=1, value="Financial Data Summary")
        title_cell = ws["A1"]
        title_cell.font = Font(bold=True, size=16)
        
        # Basic stats
        row = 3
        ws.cell(row=row, column=1, value="Total Tickers:")
        ws.cell(row=row, column=2, value=len(data_list))
        row += 1
        
        ws.cell(row=row, column=1, value="Export Date:")
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
        row += 2
        
        # Sector breakdown
        sectors = {}
        for data in data_list:
            sector = data.sector if data.sector != "N/A" else "Unknown"
            sectors[sector] = sectors.get(sector, 0) + 1
        
        if sectors:
            ws.cell(row=row, column=1, value="Sector Breakdown:")
            ws["A" + str(row)].font = Font(bold=True)
            row += 1
            
            for sector, count in sorted(sectors.items()):
                ws.cell(row=row, column=1, value=f"  {sector}:")
                ws.cell(row=row, column=2, value=count)
                row += 1
        
        # Auto-adjust columns for summary sheet
        self._auto_adjust_columns(ws)


class CSVExporter:
    """Enhanced CSV export functionality"""
    
    @staticmethod
    def export_data_with_metadata(data_list: List[FinancialData], filename: Optional[str] = None) -> str:
        """
        Export data to CSV with metadata header
        
        Args:
            data_list: List of FinancialData objects
            filename: Output filename (optional)
            
        Returns:
            Path to created CSV file
        """
        import csv
        
        if not filename:
            filename = f"financial_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            # Write metadata header
            f.write(f"# Financial Data Export\n")
            f.write(f"# Generated: {datetime.now().isoformat()}\n")
            f.write(f"# Total Records: {len(data_list)}\n")
            f.write(f"# \n")
            
            if not data_list:
                f.write("ticker\n")
                return filename
            
            # Write CSV data
            fieldnames = list(data_list[0].to_dict().keys())
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for data in data_list:
                writer.writerow(data.to_dict())
        
        logger.info(f"Exported {len(data_list)} records to CSV file: {filename}")
        return filename 